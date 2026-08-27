# pyrefly: ignore-errors
# type: ignore
"""
MFM ORINOCO – Backend Flask + SoftPLC "Todo en Uno"
=====================================================
Arquitectura de Memoria Compartida:
  - Hilo 1 (ScanEngine): ciclo PLC a 100 ms — lee/escribe el objeto global V
  - Hilo 2 (websocket_updater): lee V pasivamente cada 500 ms y emite vía SocketIO
  - Hilo Flask/SocketIO: sirve la API REST y WebSocket al frontend

Inicia con: python app.py
"""

import sys
import os
import time
import threading
import csv
import io
import logging
from datetime import datetime

from flask import Flask, jsonify, request, send_from_directory, Response, send_file
from flask_socketio import SocketIO, emit
from flask_cors import CORS
import mysql.connector
from mysql.connector import pooling

# openpyxl imports for native Excel reports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# PUENTE: inyectar python_migration en sys.path.
# Necesario para que las fases del motor (fase1..fase9, config,
# plc_timers, etc.) se resuelvan entre sí con imports relativos
# durante el runtime, sin modificar esos archivos.
# ─────────────────────────────────────────────────────────────
_MIGRATION_DIR = os.path.join(os.path.dirname(__file__), "python_migration")
if _MIGRATION_DIR not in sys.path:
    sys.path.insert(0, _MIGRATION_DIR)

# Importar usando notación de paquete para satisfacer el linter.
# python_migration/ ya contiene __init__.py, por lo que es un
# paquete Python válido reconocido por el analizador estático.
from global_vars import V                          # Memoria compartida única
from scan_engine import ScanEngine, PHASE_REGISTRY # Motor de ciclos + fases

# ─────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("orinoco.web")

# ─────────────────────────────────────────────────────────────
# Flask + SocketIO
# ─────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = "mfm_orinoco_2024"
CORS(app, resources={r"/*": {"origins": "*"}})
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

import json
HART_CONFIG_FILE = "hart_config.json"
HART_CONFIG = {
    "mode": "tcp",
    "ip": "192.168.255.1",
    "port": 502,
    "com_port": "COM3",
    "baudrate": 9600,
    "slave_id": 1,
    "start_address": 618
}
try:
    if os.path.exists(HART_CONFIG_FILE):
        with open(HART_CONFIG_FILE, "r") as f:
            loaded = json.load(f)
            # Garantizar que slave_id sea al menos 1 (nunca 0)
            if int(loaded.get('slave_id', 1)) == 0:
                loaded['slave_id'] = 1
            HART_CONFIG.update(loaded)
except Exception:
    pass

# ── Caché del último resultado HART + hilo de polling ────────────
_HART_CACHE_LOCK   = threading.Lock()
_HART_LATEST_RESULTS = {
    i: {
        "connected": False,
        "error": "Deshabilitado" if i > 0 else "Iniciando...",
        "status": 0,
        "pv_current": 0.0,
        "pv1": {"value": 0.0, "unit": "-"},
        "pv2": {"value": 0.0, "unit": "-"},
        "pv3": {"value": 0.0, "unit": "-"},
        "pv4": {"value": 0.0, "unit": "-"},
    }
    for i in range(15)
}
_HART_GLOBAL = {"last_success": None, "last_error": "", "last_attempt": 0.0}

_HART_POLL_INTERVAL = 3.0   # segundos entre lecturas HART

def _hart_background_poller():
    """Hilo daemon que lee los canales HART habilitados de la BD secuencialmente.

    Arquitectura HRT-711 (confirmada con HG Tool v1.6):
      - UN SOLO slave Modbus = el gateway (slave_id de HART_CONFIG, FIJO).
      - Cada instrumento se diferencia por la DIRECCION DE REGISTRO:
          addr = 1300 + (hart_device_index x 10)
      - hart_device_index = N en 'HART Device N' del HG Tool (0-based).
      - hart_device_address = direccion HART fisica en el bus (informativo).

      Ejemplos:
        HART Device 0 -> addr 1300 | HART Device 2 -> addr 1320
    """
    from comunicacion_hart import leer_instrumento_hart
    hart_logger = logging.getLogger("orinoco.hart.poller")
    hart_logger.info("Poller HART iniciado (intervalo %.1fs)" % _HART_POLL_INTERVAL)

    while True:
        try:
            # 1. Cargar configuracion de canales desde la BD
            channels = db_exec("SELECT * FROM hart_channel_config ORDER BY channel_idx")
            if not channels:
                # Fallback: solo Device 0 habilitado
                channels = [
                    {
                        "channel_idx":         i,
                        "v_name":              f"HART_CH{i}",
                        "description":         f"Instrumento HART {i+1}",
                        "hart_device_index":   i,
                        "hart_device_address": i + 1,
                        "enabled":             1 if i == 0 else 0
                    }
                    for i in range(15)
                ]

            # 2. Pollear cada canal habilitado secuencialmente
            for ch in channels:
                idx = int(ch["channel_idx"])

                if not ch.get("enabled", 1):
                    with _HART_CACHE_LOCK:
                        _HART_LATEST_RESULTS[idx] = {
                            "connected": False, "error": "Deshabilitado",
                            "status": 0, "pv_current": 0.0,
                            "pv1": {"value": 0.0, "unit": "-"},
                            "pv2": {"value": 0.0, "unit": "-"},
                            "pv3": {"value": 0.0, "unit": "-"},
                            "pv4": {"value": 0.0, "unit": "-"},
                        }
                    continue

                # hart_device_index = N en 'HART Device N' del HG Tool
                # Si la columna no existe en la BD, usar channel_idx como fallback
                dev_index = ch.get('hart_device_index')
                if dev_index is None:
                    dev_index = idx   # fallback
                dev_index = int(dev_index)

                # slave_id = slave del GATEWAY (el mismo para TODOS los canales)
                gateway_slave = max(1, int(HART_CONFIG.get('slave_id', 1)))

                # instrument_type = rol fijo del slot (determina inyeccion en V)
                instrument_type = ch.get('instrument_type', 'NONE') or 'NONE'

                cfg = {
                    'mode':                HART_CONFIG.get('mode', 'tcp'),
                    'ip':                  HART_CONFIG.get('ip', '192.168.255.1'),
                    'port':                int(HART_CONFIG.get('port', 502)),
                    'com_port':            HART_CONFIG.get('com_port', 'COM3'),
                    'baudrate':            int(HART_CONFIG.get('baudrate', 9600)),
                    'slave_id':            gateway_slave,
                    'hart_device_index':   dev_index,
                    'hart_device_address': ch.get('hart_device_address', dev_index + 1),
                    'instrument_type':     instrument_type,   # Rol fijo: LAMINAR_A, WEDGE_LIQ...
                }

                hart_logger.debug(
                    f"[Poller] CH{idx} '{ch.get('description','')}' "
                    f"-> Device{dev_index} addr={1300 + dev_index * 10} slave={gateway_slave}"
                )

                result = leer_instrumento_hart(cfg)

                with _HART_CACHE_LOCK:
                    _HART_LATEST_RESULTS[idx] = result
                    _HART_GLOBAL["last_attempt"] = time.monotonic()
                    if result.get("connected"):
                        _HART_GLOBAL["last_success"] = time.monotonic()
                        _HART_GLOBAL["last_error"] = ""
                    elif result.get("error") and "Deshabilitado" not in result.get("error", ""):
                        _HART_GLOBAL["last_error"] = result["error"]

                # Retardo entre lecturas para no saturar el bus HART
                time.sleep(0.5)

        except Exception as ex:
            hart_logger.error(f"[Poller] Error inesperado en loop HART: {ex}")
            with _HART_CACHE_LOCK:
                _HART_GLOBAL["last_error"] = f"Poller error: {ex}"

        time.sleep(_HART_POLL_INTERVAL)


# ─────────────────────────────────────────────────────────────
# MySQL Pool (XAMPP → puerto 3306)
# ─────────────────────────────────────────────────────────────
DB_CONFIG = dict(
    host="localhost", port=3306,
    user="root", password="",
    database="x4", charset="utf8mb4",
    connection_timeout=10,
)

try:
    db_pool = pooling.MySQLConnectionPool(pool_name="mfm", pool_size=5, **DB_CONFIG)
    logger.info("✅  MySQL pool OK")
except Exception as _e:
    logger.warning(f"⚠️  MySQL no disponible: {_e}")
    db_pool = None


def get_conn():
    global global_db_ok
    if db_pool:
        try:
            conn = db_pool.get_connection()
            global_db_ok = True
            return conn
        except Exception:
            pass
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        global_db_ok = True
        return conn
    except Exception as e:
        global_db_ok = False
        raise e


def db_exec(sql, params=None, fetch=True):
    """Ejecuta SQL; retorna filas si fetch=True, lastrowid si fetch=False."""
    global global_db_ok
    conn = cur = None
    try:
        conn = get_conn()
        cur  = conn.cursor(dictionary=True)
        cur.execute(sql, params or ())
        if fetch:
            return cur.fetchall()
        conn.commit()
        global_db_ok = True
        return cur.lastrowid
    except Exception as e:
        logger.debug(f"DB error: {e}")
        global_db_ok = False
        if conn and not fetch:
            try: conn.rollback()
            except Exception: pass
        return None
    finally:
        if cur:  cur.close()
        if conn: conn.close()

# Estado de BD global para el frontend
global_db_ok = False

def ping_db():
    global global_db_ok
    try:
        conn = get_conn()
        conn.ping(reconnect=True, attempts=1, delay=0)
        global_db_ok = True
        conn.close()
    except Exception:
        global_db_ok = False



# ─────────────────────────────────────────────────────────────
# HILO 1: Motor del SoftPLC (100 ms)
# El ScanEngine ya gestiona su propio hilo internamente mediante
# start() → _scan_loop() en modo daemon. Solo lo instanciamos y
# registramos las fases antes de arrancarlo.
# ─────────────────────────────────────────────────────────────
plc_engine = ScanEngine()
plc_engine.register_phases(PHASE_REGISTRY)


# ─────────────────────────────────────────────────────────────
# Carga de configuración DAQ desde BD
# Se ejecuta al iniciar: restaura el último puerto/baudrate guardado.
# ─────────────────────────────────────────────────────────────
def _load_daq_connection_from_db():
    """Lee la tabla daq_connection_config y aplica los parámetros al módulo modbus_daq."""
    import modbus_daq as _mdaq
    rows = db_exec("SELECT * FROM daq_connection_config WHERE id=1")
    if not rows:
        logger.info("🔌 DAQ: usando parámetros por defecto (tabla vacía)")
        return
    r = rows[0]
    _mdaq.DAQ_PORT     = str(r.get("port",     _mdaq.DAQ_PORT))
    _mdaq.DAQ_BAUDRATE = int(r.get("baudrate", _mdaq.DAQ_BAUDRATE))
    _mdaq.DAQ_SLAVE_ID = int(r.get("slave_id", _mdaq.DAQ_SLAVE_ID))
    timeout_ms = int(r.get("timeout_ms", 80))
    _mdaq.DAQ_TIMEOUT  = timeout_ms / 1000.0
    logger.info(f"✅ DAQ config cargada: {_mdaq.DAQ_PORT} @ {_mdaq.DAQ_BAUDRATE} baud, slave={_mdaq.DAQ_SLAVE_ID}")


def _load_instrument_selection_from_db():
    """Restaura la selección de instrumentos y estado de lazos desde la BD a la memoria (objeto V)."""
    try:
        # Asegurar de forma robusta que la columna b_DESHABILITA_PID y otras columnas existan en la tabla instrument_selection_config
        additional_cols = {
            "b_DESHABILITA_PID": "BOOLEAN DEFAULT FALSE",
            "b_sel_tipo_instrum_dil": "BOOLEAN DEFAULT FALSE",
            "b_AUTO_GAS_01": "BOOLEAN DEFAULT FALSE",
            "b_SEL_VLV_GAS_01": "BOOLEAN DEFAULT FALSE"
        }
        for col, col_type in additional_cols.items():
            try:
                db_exec(f"ALTER TABLE instrument_selection_config ADD COLUMN IF NOT EXISTS {col} {col_type}", fetch=False)
            except Exception:
                # En algunos motores de BD que no soporten IF NOT EXISTS en ALTER TABLE,
                # o si la columna ya existe, esto podría arrojar un error que ignoramos.
                pass

        rows = db_exec("SELECT * FROM instrument_selection_config WHERE id=1")
        if not rows:
            logger.info("ℹ️ instrument_selection_config: tabla vacía en la BD, usando valores en memoria")
            return
        
        r = rows[0]
        mapping = {
            "b_Control_PID_Gas": "b_Control_PID_Gas",
            "b_PID_POSIC_SW": "b_PID_POSIC_SW",
            "b_Sw_Wedge_Gas": "b_Sw_Wedge_Gas",
            "b_SW_DIL_MEDIDO_CALC": "b_SW_DIL_MEDIDO_CALC",
            "b_Sw_Wedge_Gas_2": "b_Sw_Wedge_Gas_2",
            "b_SEL_LAMINAR": "b_SEL_LAMINAR",
            "b_SEL_T_baja": "b_SEL_T_baja",
            "b_sw_AM_Laminar_Wedge_x": "b_sw_AM_Laminar_Wedge_x",
            "b_sw_AM_Laminar_Wedge_y": "b_sw_AM_Laminar_Wedge_y",
            "b_sel_tipo_instrum_dil": "b_sel_tipo_instrum_dil",
            "b_AUTO_GAS_01": "b_AUTO_GAS_01",
            "b_SEL_VLV_GAS_01": "b_SEL_VLV_GAS_01",
            "b_DESHABILITA_PID": "b_DESHABILITA_PID"
        }
        
        for db_col, v_var in mapping.items():
            if db_col in r and r[db_col] is not None:
                setattr(V, v_var, bool(r[db_col]))
                
        # Sincronizar b_Sel_T_baja si corresponde
        if hasattr(V, "b_Sel_T_baja") and hasattr(V, "b_SEL_T_baja"):
            V.b_Sel_T_baja = V.b_SEL_T_baja

        logger.info("✅ Selección de instrumentos y estado de lazos (b_DESHABILITA_PID) restaurados desde BD")
    except Exception as e:
        logger.error(f"❌ Error al restaurar instrument_selection_config desde BD: {e}")


def _load_configuracion_actual_from_db():
    """Carga los parámetros de los lazos PID (SP, modo, CV manual, Kp, Ki, Kd) desde la BD a V."""
    try:
        rows = db_exec("SELECT * FROM configuracion_actual")
        for r in (rows or []):
            inst = r.get("instrumento")
            modo = r.get("modo", "Auto")
            sp = r.get("SP", 0.0)
            cv_man = r.get("CV_manual", 0.0)
            kp = r.get("Kp", 1.0)
            ki = r.get("Ki", 0.1)
            kd = r.get("Kd", 0.0)
            
            if inst == "LIC-01":
                V.b_MAN_LC = (modo == "Manual")
                V.r_LEVEL_PID_SP = float(sp)
                V.r_LEVEL_PID_03_CVOverride = float(cv_man)
                V.r_LEVEL_PID_03_CVOper = float(cv_man)
                V.r_LEVEL_PID_03_KP = float(kp)
                V.r_LEVEL_PID_03_KI = float(ki)
                V.r_LEVEL_PID_03_KD = float(kd)
                logger.info(f"⚙️ PID LIC-01 restaurado desde BD: modo={modo}, SP={sp}, CV_man={cv_man}, Kp={kp}, Ki={ki}, Kd={kd}")
            elif inst == "PIC-01":
                V.b_MAN_PC = (modo == "Manual")
                V.r_PRESS_PID_SP = float(sp)
                V.r_PRESS_PID_03_CVOverride = float(cv_man)
                V.r_PRESS_PID_03_CVOper = float(cv_man)
                V.r_PRESS_PID_03_KP = float(kp)
                V.r_PRESS_PID_03_KI = float(ki)
                V.r_PRESS_PID_03_KD = float(kd)
                logger.info(f"⚙️ PID PIC-01 restaurado desde BD: modo={modo}, SP={sp}, CV_man={cv_man}, Kp={kp}, Ki={ki}, Kd={kd}")
    except Exception as e:
        logger.error(f"❌ Error al restaurar configuracion_actual desde BD: {e}")


def _init_and_load_prueba_config():
    """Crea la tabla prueba_configuracion si no existe y restaura sus valores en V."""
    global _ACTIVE_PRUEBA_ID, _LAST_PRUEBA_EN_PROGRESO
    try:
        # 1. Crear tabla prueba_configuracion si no existe
        db_exec("""
            CREATE TABLE IF NOT EXISTS prueba_configuracion (
                id INT PRIMARY KEY,
                lugar VARCHAR(16) DEFAULT '',
                pozo VARCHAR(6) DEFAULT '',
                metodo VARCHAR(6) DEFAULT '',
                rpm VARCHAR(4) DEFAULT '',
                inyeccion VARCHAR(4) DEFAULT '',
                temp_yac FLOAT DEFAULT 0.0,
                api_formacion FLOAT DEFAULT 0.0,
                api_mezcla FLOAT DEFAULT 0.0,
                api_diluente FLOAT DEFAULT 0.0,
                caudal_diluente FLOAT DEFAULT 0.0,
                duracion_horas INT DEFAULT 0,
                combo_metodo INT DEFAULT 0,
                combo_inyeccion INT DEFAULT 0,
                fecha_dd INT DEFAULT 0,
                fecha_mm INT DEFAULT 0,
                fecha_aaaa INT DEFAULT 0,
                hora_hh INT DEFAULT 0,
                hora_mm INT DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """, fetch=False)

        # 1b. Crear tabla historico_pruebas si no existe
        db_exec("""
            CREATE TABLE IF NOT EXISTS historico_pruebas (
                id INT AUTO_INCREMENT PRIMARY KEY,
                codigo_pozo VARCHAR(64) DEFAULT '',
                lugar VARCHAR(64) DEFAULT '',
                pozo VARCHAR(32) DEFAULT '',
                metodo VARCHAR(32) DEFAULT '',
                rpm VARCHAR(32) DEFAULT '',
                inyeccion VARCHAR(32) DEFAULT '',
                temp_yac FLOAT DEFAULT 0.0,
                api_formacion FLOAT DEFAULT 0.0,
                api_mezcla FLOAT DEFAULT 0.0,
                api_diluente FLOAT DEFAULT 0.0,
                caudal_diluente FLOAT DEFAULT 0.0,
                duracion_horas FLOAT DEFAULT 0.0,
                fecha_inicio DATETIME NOT NULL,
                fecha_fin DATETIME DEFAULT NULL,
                estado VARCHAR(20) DEFAULT 'En progreso'
            )
        """, fetch=False)

        # 1c. Migrar tabla lecturas_proceso para agregar columna prueba_id si no existe
        try:
            cols = db_exec("SHOW COLUMNS FROM lecturas_proceso LIKE 'prueba_id'")
            if not cols:
                db_exec("ALTER TABLE lecturas_proceso ADD COLUMN prueba_id INT DEFAULT NULL", fetch=False)
                logger.info("MIGRATION: Added column 'prueba_id' to table 'lecturas_proceso'")
        except Exception as em:
            logger.error(f"Error en migracion column prueba_id: {em}")
        
        # 2. Intentar cargar el registro id=1
        rows = db_exec("SELECT * FROM prueba_configuracion WHERE id = 1")
        if rows:
            row = rows[0]
            # Escribir en V
            V.as_Codigo_pozo_17 = str(row.get("lugar", ""))
            V.as_Codigo_pozo_03 = str(row.get("pozo", ""))
            V.as_Codigo_pozo_06 = str(row.get("metodo", ""))
            V.as_Codigo_pozo_08 = str(row.get("rpm", ""))
            V.as_Codigo_pozo_18 = str(row.get("inyeccion", ""))
            
            V.r_T_Yac_C = float(row.get("temp_yac", 0.0))
            V.r_API_formacion_BM = float(row.get("api_formacion", 0.0))
            V.r_API_2 = float(row.get("api_mezcla", 0.0))
            V.r_API_1 = float(row.get("api_diluente", 0.0))
            V.r_caudal_dil_BM = float(row.get("caudal_diluente", 0.0))
            
            V.i_duracion_prueba_horas = int(row.get("duracion_horas", 0))
            V.i_posicion_combo_box_1 = int(row.get("combo_metodo", 0))
            V.i_posicion_combo_box_2 = int(row.get("combo_inyeccion", 0))
            
            # Cargar fecha/hora de inicio
            arr = list(getattr(V, "ad_IHM_HORA_inicio", [0]*8))
            arr[2] = int(row.get("fecha_dd", 0))
            arr[1] = int(row.get("fecha_mm", 0))
            arr[0] = int(row.get("fecha_aaaa", 0))
            arr[3] = int(row.get("hora_hh", 0))
            arr[4] = int(row.get("hora_mm", 0))
            V.ad_IHM_HORA_inicio = arr
            
            print("  [OK] Configuracion de prueba de pozo restaurada desde BD")
        else:
            print("  [INFO] No hay configuracion de prueba previa en la BD")

        # 3. Limpiar o recuperar pruebas en progreso
        pep = bool(getattr(V, "b_Prueba_en_Progreso", False))
        if pep:
            running = db_exec("SELECT * FROM historico_pruebas WHERE estado = 'En progreso' ORDER BY id DESC LIMIT 1")
            if running:
                _ACTIVE_PRUEBA_ID = running[0]["id"]
                _LAST_PRUEBA_EN_PROGRESO = True
                print(f"  [OK] Re-conectado a prueba activa en BD ID: {_ACTIVE_PRUEBA_ID}")
            else:
                # Crear una nueva si no existe
                codigo = str(getattr(V, "as_Codigo_pozo_16", ""))
                lugar = str(getattr(V, "as_Codigo_pozo_17", ""))
                pozo = str(getattr(V, "as_Codigo_pozo_03", ""))
                metodo = str(getattr(V, "as_Codigo_pozo_06", ""))
                rpm = str(getattr(V, "as_Codigo_pozo_08", ""))
                inyeccion = str(getattr(V, "as_Codigo_pozo_18", ""))
                temp_yac = float(getattr(V, "r_T_Yac_C", 0.0))
                api_form = float(getattr(V, "r_API_formacion_BM", 0.0))
                api_mezc = float(getattr(V, "r_API_2", 0.0))
                api_dil = float(getattr(V, "r_API_1", 0.0))
                caud_dil = float(getattr(V, "r_caudal_dil_BM", 0.0))
                duracion = float(getattr(V, "i_duracion_prueba_horas", 0.0))
                
                _ACTIVE_PRUEBA_ID = db_exec("""
                    INSERT INTO historico_pruebas (
                        codigo_pozo, lugar, pozo, metodo, rpm, inyeccion,
                        temp_yac, api_formacion, api_mezcla, api_diluente, caudal_diluente,
                        duracion_horas, fecha_inicio, estado
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), 'En progreso')
                """, (codigo, lugar, pozo, metodo, rpm, inyeccion,
                      temp_yac, api_form, api_mezc, api_dil, caud_dil, duracion), fetch=False)
                _LAST_PRUEBA_EN_PROGRESO = True
                print(f"  [OK] Creado registro de prueba activa detectada al arrancar: ID {_ACTIVE_PRUEBA_ID}")
        else:
            db_exec("UPDATE historico_pruebas SET estado = 'Interrumpida', fecha_fin = NOW() WHERE estado = 'En progreso'", fetch=False)
            _ACTIVE_PRUEBA_ID = None
            _LAST_PRUEBA_EN_PROGRESO = False

    except Exception as e:
        logger.error(f"Error inicializando/cargando configuracion de prueba: {e}")


def _init_and_load_pvt_balance_config():
    """Crea la tabla pvt_balance_config si no existe y restaura sus valores en V."""
    try:
        db_exec("""
            CREATE TABLE IF NOT EXISTS pvt_balance_config (
                id             INT PRIMARY KEY DEFAULT 1,
                pvt_mode       INT DEFAULT 0,
                temp_yac       FLOAT DEFAULT 0.0,
                rso            FLOAT DEFAULT 0.0,
                bo             FLOAT DEFAULT 1.0,
                api_form_real  FLOAT DEFAULT 0.0,
                api_form_teo   FLOAT DEFAULT 0.0,
                api_mez_real   FLOAT DEFAULT 0.0,
                api_mez_teo    FLOAT DEFAULT 0.0,
                api_dil_real   FLOAT DEFAULT 0.0,
                api_dil_teo    FLOAT DEFAULT 0.0,
                q_dil_real     FLOAT DEFAULT 0.0,
                q_dil_teo      FLOAT DEFAULT 0.0,
                q_net_real     FLOAT DEFAULT 0.0,
                q_net_teo      FLOAT DEFAULT 0.0,
                q_net_dil_real FLOAT DEFAULT 0.0,
                q_net_dil_teo  FLOAT DEFAULT 0.0,
                q_agua_real    FLOAT DEFAULT 0.0,
                q_agua_teo     FLOAT DEFAULT 0.0,
                q_total_real   FLOAT DEFAULT 0.0,
                q_total_teo    FLOAT DEFAULT 0.0,
                updated_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """, fetch=False)

        rows = db_exec("SELECT * FROM pvt_balance_config WHERE id = 1")
        if rows:
            row = rows[0]
            mode = int(row.get("pvt_mode", 0))
            V.b_PB_PVT = bool(mode == 1)
            rso = float(row.get("rso", 0.0))
            bo = float(row.get("bo", 1.0))
            V.r_Rso_PT2 = rso
            V.r_Bo2 = bo if bo > 0 else 1.0
            if V.b_PB_PVT:
                V.r_Rso_PT = V.r_Rso_PT2
                V.r_Bo = V.r_Bo2
            print(f"  [OK] Configuracion PVT y Balance de Masa restaurada desde BD (modo={'INGRESADA' if mode==1 else 'CALCULADA'}, RSO={rso}, BO={bo})")
        else:
            db_exec("INSERT INTO pvt_balance_config (id, pvt_mode, bo) VALUES (1, 0, 1.0)", fetch=False)
            print("  [INFO] Configuracion inicial de PVT y Balance de Masa creada en BD")
    except Exception as e:
        logger.error(f"Error inicializando/cargando pvt_balance_config: {e}")


def _load_daq_channels_from_db():
    """Lee la tabla daq_channel_config y aplica los parámetros a los módulos fase2_entradas."""
    try:
        import sys
        rows = db_exec("SELECT * FROM daq_channel_config ORDER BY channel_addr")
        if not rows:
            logger.info("🔌 DAQ Canales: usando mapa por defecto (tabla vacía)")
            return
        
        new_map = []
        for r in rows:
            if r.get("enabled", 1):
                phys_addr = r.get("modbus_addr")
                if phys_addr is None:
                    phys_addr = r["channel_addr"]
                new_map.append((
                    r["v_name"],
                    int(phys_addr),
                    float(r.get("scale", 1000.0)),
                    r["description"]
                ))
        
        if new_map:
            import fase2_entradas as _f2
            
            # Actualizar en el módulo importado por Flask
            _f2._INPUT_MAP = new_map
            _f2.V._daq_channel_snapshot = [
                {
                    "ch":        entry[1],
                    "var":       entry[0],
                    "desc":      entry[3],
                    "raw":       None,
                    "ma":        None,
                    "open_wire": True,
                }
                for entry in new_map
            ]
            
            # Actualizar en el módulo importado por ScanEngine
            _f2_se = sys.modules.get('fase2_entradas')
            if _f2_se:
                _f2_se._INPUT_MAP = new_map
                if hasattr(_f2_se, 'V'):
                    _f2_se.V._daq_channel_snapshot = _f2.V._daq_channel_snapshot
            
            logger.info(f"✅ DAQ Canales cargados/actualizados de BD: {new_map}")
    except Exception as e:
        logger.error(f"⚠️ Error cargando canales de la BD: {e}")

def _load_daq_ao_from_db():
    """Lee la tabla daq_ao_config y aplica los parámetros a fase8_salidas."""
    try:
        import sys
        rows = db_exec("SELECT * FROM daq_ao_config ORDER BY channel_addr")
        if not rows:
            logger.info("🔌 DAQ AO: usando mapa por defecto (tabla vacía)")
            return
        
        new_map = []
        for r in rows:
            if r.get("enabled", 1):
                phys_addr = r.get("modbus_addr")
                if phys_addr is None:
                    phys_addr = r["channel_addr"]
                new_map.append((
                    r["v_name"],
                    int(phys_addr),
                    float(r.get("scale_min", 4000.0)),
                    float(r.get("scale_max", 20000.0)),
                    r["description"]
                ))
        
        if new_map:
            import fase8_salidas as _f8
            
            # Actualizar en el módulo importado por Flask
            _f8._OUTPUT_MAP = new_map
            
            # Actualizar en el módulo importado por ScanEngine
            _f8_se = sys.modules.get('fase8_salidas')
            if _f8_se:
                _f8_se._OUTPUT_MAP = new_map
            
            logger.info(f"✅ DAQ AO cargados/actualizados de BD: {new_map}")
    except Exception as e:
        logger.error(f"⚠️ Error cargando canales AO de la BD: {e}")



# ── Estado de seguimiento de prueba de pozo ───────────────────
_ACTIVE_PRUEBA_ID = None
_LAST_PRUEBA_EN_PROGRESO = False

# ─────────────────────────────────────────────────────────────
# HILO 2: WebSocket Updater (500 ms)
# Lee pasivamente el objeto V (memoria compartida) y emite los
# datos de proceso al frontend. NO escribe en V.
# ─────────────────────────────────────────────────────────────
def websocket_updater():
    """
    Publica el estado del proceso hacia el frontend cada 500 ms.
    Lee variables directamente del objeto global V del SoftPLC.

    Mapeo de variables V → Tags de display:
      V.r_T_Gas       → TI_02   (Temperatura Gas)
      V.r_Q_GAT       → GAS_01  (Caudal GAT)
      V.r_WC          → VI_01   (Corte de agua)
    """
    global _LAST_PRUEBA_EN_PROGRESO, _ACTIVE_PRUEBA_ID
    logger.info("  WebSocket Updater activo (500 ms)")
    loop_count = 0

    while True:
        try:
            # Monitorear transiciones de prueba de pozo
            pep = bool(getattr(V, "b_Prueba_en_Progreso", False))
            if pep and not _LAST_PRUEBA_EN_PROGRESO:
                # Transición False -> True: Inicio de prueba
                try:
                    codigo = str(getattr(V, "as_Codigo_pozo_16", ""))
                    lugar = str(getattr(V, "as_Codigo_pozo_17", ""))
                    pozo = str(getattr(V, "as_Codigo_pozo_03", ""))
                    metodo = str(getattr(V, "as_Codigo_pozo_06", ""))
                    rpm = str(getattr(V, "as_Codigo_pozo_08", ""))
                    inyeccion = str(getattr(V, "as_Codigo_pozo_18", ""))
                    
                    temp_yac = float(getattr(V, "r_T_Yac_C", 0.0))
                    api_form = float(getattr(V, "r_API_formacion_BM", 0.0))
                    api_mezc = float(getattr(V, "r_API_2", 0.0))
                    api_dilu = float(getattr(V, "r_API_1", 0.0))
                    caud_dil = float(getattr(V, "r_caudal_dil_BM", 0.0))
                    duracion = float(getattr(V, "i_duracion_prueba_horas", 0.0))
                    
                    insert_id = db_exec("""
                        INSERT INTO historico_pruebas (
                            codigo_pozo, lugar, pozo, metodo, rpm, inyeccion,
                            temp_yac, api_formacion, api_mezcla, api_diluente, caudal_diluente,
                            duracion_horas, fecha_inicio, estado
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), 'En progreso')
                    """, (codigo, lugar, pozo, metodo, rpm, inyeccion,
                          temp_yac, api_form, api_mezc, api_dilu, caud_dil, duracion), fetch=False)
                    
                    _ACTIVE_PRUEBA_ID = insert_id
                    logger.info(f"💾 Historico: Prueba iniciada en BD. ID = {insert_id}")
                except Exception as ex_st:
                    logger.error(f"Error al registrar inicio de prueba: {ex_st}")
                _LAST_PRUEBA_EN_PROGRESO = True
                
            elif not pep and _LAST_PRUEBA_EN_PROGRESO:
                # Transición True -> False: Fin de prueba
                try:
                    if _ACTIVE_PRUEBA_ID:
                        is_aborted = bool(getattr(V, "b_IHM_Abortar_Prueba", False))
                        estado = 'Abortada' if is_aborted else 'Completada'
                        codigo = str(getattr(V, "as_Codigo_pozo_16", ""))
                        
                        db_exec("""
                            UPDATE historico_pruebas 
                            SET fecha_fin = NOW(), estado = %s, codigo_pozo = %s
                            WHERE id = %s
                        """, (estado, codigo, _ACTIVE_PRUEBA_ID), fetch=False)
                        logger.info(f"💾 Historico: Prueba finalizada en BD. ID = {_ACTIVE_PRUEBA_ID} con estado = {estado}")
                        _ACTIVE_PRUEBA_ID = None
                except Exception as ex_end:
                    logger.error(f"Error al registrar fin de prueba: {ex_end}")
                _LAST_PRUEBA_EN_PROGRESO = False

            # ── Leer datos de Coriolis desde base de datos ──
            coriolis_data = {}
            try:
                coriolis_rows = db_exec(
                    "SELECT var_name, current_val FROM modbus_rtu_variables WHERE var_name IN ('Coriolis_Density', 'Coriolis_Temp', 'Coriolis_Vol_flow_Rate')"
                )
                if coriolis_rows:
                    for row in coriolis_rows:
                        coriolis_data[row["var_name"]] = float(row["current_val"]) if row["current_val"] is not None else 0.0
            except Exception as e:
                logger.error(f"Error reading Coriolis vars: {e}")

            # Overrides manuales para Coriolis si están activos
            overrides = getattr(V, 'instrument_overrides', {})
            c_dens = overrides.get("Coriolis_Density", coriolis_data.get("Coriolis_Density", 0.0))
            c_temp = overrides.get("Coriolis_Temp", coriolis_data.get("Coriolis_Temp", 0.0))
            c_flow = overrides.get("Coriolis_Vol_flow_Rate", coriolis_data.get("Coriolis_Vol_flow_Rate", 0.0))

            # Sincronizar en V para alarmas y cálculos
            setattr(V, "Coriolis_Density", float(c_dens))
            setattr(V, "Coriolis_Temp", float(c_temp))
            setattr(V, "Coriolis_Vol_flow_Rate", float(c_flow))

            # ── Leer datos de proceso desde V (solo lectura) ──
            process_data = {
                "Coriolis_Density": round(float(c_dens), 3),
                "Coriolis_Temp": round(float(c_temp), 2),
                "Coriolis_Vol_flow_Rate": round(float(c_flow), 2),
                "r_Q_gas_STD": round(V.r_Q_gas_STD,           3),   # Caudal de gas standard
                "r_P_Gas":     round(V.r_P_Gas,               2),   # Presión de gas
                "r_T_Gas":     round(V.r_T_Gas,               2),   # Temp gas
                "r_LIT_001":   round(V.r_LIT_001,             2),   # Nivel Separador
                "PDI_01":      round(V.r_PDT_01,              2),   # Laminar A (PDI-01)
                "r_PDT_02":    round(V.r_PDT_02,              2),   # Wedge liquido en inh2o
                "PDI_03":      round(V.r_PDT_03,              2),   # Laminar B
                "r_P_Oil":     round(V.r_P_Oil,               2),   # Presión de crudo (Wedge liquido)
                "r_T_Oil_C":   round(V.r_T_Oil_C,             2),   # Temp crudo °C (Wedge liquido temp)
                "r_T_oil_F":   round(V.r_T_Oil_F,             2),   # Temp crudo °F
                "r_GVoidF":    round(V.r_GVoidF,              3),   # % Fracción de gas
                "r_v_oil_medida": round(V.r_v_oil_medida,     3),   # Viscosidad
                "r_WC":        round(V.r_WC,                  2),   # WC Corte de Agua
                "r_nivel_aux": round(V.r_nivel_aux,           2),   # Nivel aux
                "r_Q_gas":     round(V.r_Q_gas,               3),   # Caudal de gas calculado
                "r_Transmisor_Gas": round(getattr(V, "r_Transmisor_Gas", 0.0), 2), # DP Wedge Gas
                # ── Tablas inferiores ─────────────────────────────────
                "Est_Q_Liq":   round(V.r_Qb_Liquido_Estimado, 3),   # Est.Qliq
                "Est_Q_Crudo": round(V.r_Q_Crudo_Estimado,    3),   # Est.Q.Crudo
                "Est_Q_Neto":  round(V.r_Q_Crudo_Neto_Estimado, 3), # Est.Q.Neto
                "Est_Q_Dil":   round(V.r_Qb_Dil_Estimado, 3),       # Est.Q.Diluente
                "Est_Q_Agua":  round(V.r_Q_W_Estimado,        3),   # Est.QAgua
                "Est_Q_Gas":   round(V.r_Q_gas_sc_Estimado,   3),   # Est.Q.Gas Total STD

                # ── Datos Prueba Progreso ──
                "Vol_Liquido": round(float(getattr(V, "r_Vol_Liquido_Total", 0.0)), 3),
                "Vol_Crudo": round(float(getattr(V, "r_Vol_Crudo_Total", 0.0)), 3),
                "Vol_Crudo_Neto": round(float(getattr(V, "r_Vol_Crudo_Total_neto", 0.0)), 3),
                "Vol_Diluente": round(float(getattr(V, "r_Vol_Dil_Total", 0.0)), 3),
                "Vol_Agua": round(float(getattr(V, "r_Vol_W_Total", 0.0)), 3),
                "Vol_Gas_Arr": round(float(getattr(V, "r_Vol_gat_Total", 0.0)) * 1000.0, 3),
                "Vol_Gas_Total": round(float(getattr(V, "r_Vol_gas_Total", 0.0)), 3),

                "Vol_Liquido_sc": round(float(getattr(V, "r_Vol_Liquido_Total_sc", 0.0)), 3),
                "Vol_Crudo_sc": round(float(getattr(V, "r_Vol_Crudo_Total_sc", 0.0)), 3),
                "Vol_Crudo_Neto_sc": round(float(getattr(V, "r_Vol_Crudo_Total_neto_sc", 0.0)), 3),
                "Vol_Diluente_sc": round(float(getattr(V, "r_Vol_Dil_Total_sc", 0.0)), 3),
                "Vol_Agua_sc": round(float(getattr(V, "r_Vol_W_Total_sc", 0.0)), 3),
                "Vol_Gas_Arr_sc": round(float(getattr(V, "r_Vol_gat_Total_sc", 0.0)) * 1000.0, 3),
                "Vol_Gas_Total_sc": round(float(getattr(V, "r_Vol_gas_Total_sc", 0.0)), 3),

                "Est_Q_gat": round(float(getattr(V, "r_Q_gat_Estimado", 0.0)) * 1000.0, 3),
                "Est_Q_gas_line": round(float(getattr(V, "r_Q_gas_Estimado", 0.0)), 3),
                "Est_Q_Liq_sc": round(float(getattr(V, "r_Qb_Liquido_sc_Estimado", 0.0)), 3),
                "Est_Q_Crudo_sc": round(float(getattr(V, "r_Q_Crudo_sc_Estimado", 0.0)), 3),
                "Est_Q_Crudo_Neto_sc": round(float(getattr(V, "r_Q_Crudo_Neto_Estimado_sc", 0.0)), 3),
                "Est_Q_Dil_sc": round(float(getattr(V, "r_Qb_Dil_Estimado_sc", 0.0)), 3),
                "Est_Q_Agua_sc": round(float(getattr(V, "r_Q_W_sc_Estimado", 0.0)), 3),
                "Est_Q_gat_sc": round(float(getattr(V, "r_Q_gat_sc_Estimado", 0.0)) * 1000.0, 3),

                "Q_Liq":       round(V.r_Q_Liquido,           3),   # Q.Liq
                "Q_Crudo":     round(V.r_Q_Crudo,             3),   # Q.Crudo
                "Q_Neto":      round(V.r_Q_Crudo - (V.r_caudal_dil_BM if ((getattr(V, 'i_posicion_combo_box_2', 0) == 1) or (str(getattr(V, 'as_Codigo_pozo_18', '')).strip().upper() == 'SI')) else 0.0), 3), # Q.Neto
                "Q_Agua":      round(V.r_Q_W,                 3),   # Q.Agua (mapeado a r_Q_W)
                "Q_Dil":       round(V.r_caudal_dil_BM if ((getattr(V, 'i_posicion_combo_box_2', 0) == 1) or (str(getattr(V, 'as_Codigo_pozo_18', '')).strip().upper() == 'SI')) else 0.0, 3),   # Q.Diluente
                "Q_Gas":       round(V.r_Q_gas_STD,           3),   # Q.Gas (mapeado a r_Q_gas_STD)
                # ── Extras ───────────────────────────────────────────
                "Q_W":         round(V.r_Q_W,                 3),
                "Q_gas_STD":   round(V.r_Q_gas_STD,           3),
                "GOR":         round(V.r_GOR,                 2),
                "WC_sc":       round(V.r_WC_sc,               3),
                "GVF":         round(V.r_GVF,                 3),
                "b_Laminar":   bool(V.b_Laminar),
                "b_Wedge":     bool(V.b_Wedge),
                "i_Tipo_medidor": int(V.i_Tipo_medidor),
                "timestamp":   datetime.now().strftime("%H:%M:%S"),
                # ── Estado de Prueba ──────────────────────────────────
                "b_Prueba_en_Progreso": bool(getattr(V, "b_Prueba_en_Progreso", False)),
                "b_Parada_en_Progreso": bool(getattr(V, "b_Parada_en_Progreso", False)),
                "ad_TIEMPO_inicio_prueba": list(getattr(V, "ad_TIEMPO_inicio_prueba", [0]*8)),
                "ad_TIEMPO_prueba":        list(getattr(V, "ad_TIEMPO_prueba",        [0]*8)),
                # ── Tags Inicio Prueba (Pantalla) ─────────────────────
                "i_duracion_prueba_horas": int(getattr(V, "i_duracion_prueba_horas", 0)),
                "as_Codigo_pozo_16":  str(getattr(V, "as_Codigo_pozo_16",  "")),
                "as_Codigo_pozo_17":  str(getattr(V, "as_Codigo_pozo_17",  "")),
                "as_Codigo_pozo_03":  str(getattr(V, "as_Codigo_pozo_03",  "")),
                "as_Codigo_pozo_06":  str(getattr(V, "as_Codigo_pozo_06",  "")),
                "as_Codigo_pozo_08":  str(getattr(V, "as_Codigo_pozo_08",  "")),
                "as_Codigo_pozo_18":  str(getattr(V, "as_Codigo_pozo_18",  "")),
                "as_Codigo_pozo_19":  str(getattr(V, "as_Codigo_pozo_19",  "")),
                "r_T_Yac_C":          round(float(getattr(V, "r_T_Yac_C",         0.0)), 3),
                "r_API_formacion_BM":  round(float(getattr(V, "r_API_formacion_BM", 0.0)), 3),
                "r_API_2":             round(float(getattr(V, "r_API_2",            0.0)), 3),
                "r_API_1":             round(float(getattr(V, "r_API_1",            0.0)), 3),
                "r_caudal_dil_BM":     round(float(getattr(V, "r_caudal_dil_BM",   0.0)), 3),
                "r_API_MEZCLA_TEORICO": round(float(getattr(V, "r_API_MEZCLA_TEORICO", 0.0)), 3),
                "r_CAUDAL_NETO_TEORICO": round(float(getattr(V, "r_CAUDAL_NETO_TEORICO", 0.0)), 3),
                "r_Rso_PT": round(float(getattr(V, "r_Rso_PT", 0.0)), 3),
                "r_Bo_PT": round(float(getattr(V, "r_Bo", getattr(V, "r_Bo_PT", 0.0))), 3),
                "i_posicion_combo_box_1": int(getattr(V, "i_posicion_combo_box_1", 0)),
                "i_posicion_combo_box_2": int(getattr(V, "i_posicion_combo_box_2", 0)),
                "ar_TIEMPO_prueba_TOTAL": list(getattr(V, "ar_TIEMPO_prueba_TOTAL", [0]*8)),
                "ad_IHM_HORA_inicio":     list(getattr(V, "ad_IHM_HORA_inicio",    [0]*8)),
            }

            # ── Estado de los lazos PID del SoftPLC ──
            pid_nivel_data = {
                "instrumento": "LIC-01",
                "modo":        "Auto" if not V.b_MAN_LC else "Manual",
                "PV":          round(V.r_nivel_aux if V.b_PID_POSIC_SW else V.r_LIT_001, 2),
                "SP":          round(V.r_LEVEL_PID_SP, 2),
                "CV":          round(V.fb_LEVEL_PID_r_CVEU, 2),   # LCV-01: r_CVEU del PID de Nivel
                "CV_manual":   round(V.r_LEVEL_PID_03_CVOverride, 2),
                "Kp":          round(V.r_LEVEL_PID_03_KP, 4),
                "Ki":          round(V.r_LEVEL_PID_03_KI, 4),
                "Kd":          round(V.r_LEVEL_PID_03_KD, 4),
            }

            pid_presion_data = {
                "instrumento": "PIC-01",
                "modo":        "Auto" if not V.b_MAN_PC else "Manual",
                "PV":          round(V.r_PRESS_PID_PV, 2),
                "SP":          round(V.r_PRESS_PID_SP, 2),
                "CV":          round(V.fb_PRESS_PID_r_CVEU, 2),   # PCV-01: r_CVEU del PID de Presion
                "CV_manual":   round(V.r_PRESS_PID_03_CVOverride, 2),
                "Kp":          round(V.r_PRESS_PID_03_KP, 4),
                "Ki":          round(V.r_PRESS_PID_03_KI, 4),
                "Kd":          round(V.r_PRESS_PID_03_KD, 4),
            }

            instrument_selection_data = {
                "b_Control_PID_Gas": bool(V.b_Control_PID_Gas),
                "b_PID_POSIC_SW": bool(V.b_PID_POSIC_SW),
                "b_Sw_Wedge_Gas": bool(V.b_Sw_Wedge_Gas),
                "b_SW_DIL_MEDIDO_CALC": bool(V.b_SW_DIL_MEDIDO_CALC),
                "b_Sw_Wedge_Gas_2": bool(V.b_Sw_Wedge_Gas_2),
                "b_SEL_LAMINAR": bool(V.b_SEL_LAMINAR),
                "b_SEL_T_baja": bool(V.b_SEL_T_baja),
                "b_sw_AM_Laminar_Wedge_x": bool(V.b_sw_AM_Laminar_Wedge_x),
                "b_sw_AM_Laminar_Wedge_y": bool(V.b_sw_AM_Laminar_Wedge_y),
                "b_sel_tipo_instrum_dil": bool(V.b_sel_tipo_instrum_dil),
                "b_AUTO_GAS_01": bool(V.b_AUTO_GAS_01),
                "b_SEL_VLV_GAS_01": bool(V.b_SEL_VLV_GAS_01),
            }

            # ── Estado del Motor PLC ──
            plc_status = plc_engine.get_status()

            # ── Check real-time DB status cada ~2 segundos (4 loops de 500ms) ──
            if loop_count % 4 == 0:
                ping_db()

            # ── Emitir todo al frontend vía SocketIO ──
            socketio.emit("process_data", {
                "process":      process_data,
                "pid_nivel":    pid_nivel_data,
                "pid_presion":  pid_presion_data,
                "plc":          plc_status,
                "lazos_habilitados": not V.b_DESHABILITA_PID,
                "db_ok":        global_db_ok,
                "instrument_selection": instrument_selection_data,
            })

            # ── Guardar histórico en DB cada 10 s (10 × 1000 ms ciclo) ──
            loop_count += 1
            if loop_count >= 10:
                loop_count = 0
                _persist_lecturas(process_data)
                _check_and_record_alarmas()

        except Exception as e:
            logger.error(f"WebSocket Updater error: {e}")

        time.sleep(0.5)


# ── Estado de alarmas previo para detección de bordes ──────────────────────
_alarm_prev_state: dict = {}   # instrumento → nivel previo ('', 'HH', 'H', 'L', 'LL')

# Mapa instrumento → variable en V
_ALARM_VAR_MAP = {
    "FI-03":                  "r_Q_gas_STD",
    "GAS-01":                 "r_GVoidF",
    "LI-01":                  "r_LIT_001",
    "PDI-01":                 "r_PDT_01",
    "PDI-02":                 "r_PDT_02",
    "PDI-03":                 "r_PDT_03",
    "PDI-04":                 "r_Transmisor_Gas",
    "PI-01":                  "r_P_Gas",
    "PI-02":                  "r_P_Oil",
    "TI-01":                  "r_T_Oil_C",
    "TI-02":                  "r_T_Gas",
    "VI-01":                  "r_v_oil_medida",
    "WC":                     "r_WC",
    "NIV-AUX":                "r_nivel_aux",
    "Coriolis_Density":       "Coriolis_Density",
    "Coriolis_Temp":          "Coriolis_Temp",
    "Coriolis_Vol_flow_Rate": "Coriolis_Vol_flow_Rate",
    "Coriolis_Vol_flow_Ra":   "Coriolis_Vol_flow_Rate",
}

def _check_and_record_alarmas():
    """Detecta transiciones de alarma y las registra en historico_alarmas."""
    global _alarm_prev_state
    try:
        alarmas_cfg = db_exec(
            "SELECT instrumento, descripcion, unidad, SP_HH, SP_H, SP_L, SP_LL "
            "FROM tabla_configuracion_alarma"
        ) or []
        for cfg in alarmas_cfg:
            inst = cfg["instrumento"]
            var = _ALARM_VAR_MAP.get(inst)
            if not var:
                continue
            valor = float(getattr(V, var, 0.0) or 0.0)
            sp_hh = cfg.get("SP_HH")
            sp_h  = cfg.get("SP_H")
            sp_l  = cfg.get("SP_L")
            sp_ll = cfg.get("SP_LL")

            # Determinar nivel actual
            if sp_hh is not None and valor >= sp_hh:
                nivel = "HH"
                sp_act = sp_hh
            elif sp_h is not None and valor >= sp_h:
                nivel = "H"
                sp_act = sp_h
            elif sp_ll is not None and valor <= sp_ll:
                nivel = "LL"
                sp_act = sp_ll
            elif sp_l is not None and valor <= sp_l:
                nivel = "L"
                sp_act = sp_l
            else:
                nivel = ""
                sp_act = None

            prev = _alarm_prev_state.get(inst, "")
            if nivel != prev:  # transición detectada
                ts_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if nivel:  # entró a alarma
                    db_exec(
                        "INSERT INTO historico_alarmas "
                        "(instrumento, descripcion, unidad, valor, nivel, sp_activo) "
                        "VALUES (%s, %s, %s, %s, %s, %s)",
                        (inst, cfg.get("descripcion", ""), cfg.get("unidad", ""),
                         round(valor, 4), nivel, sp_act),
                        fetch=False
                    )
                    logger.warning(
                        f"🚨 ALARMA {nivel} | {inst} ({cfg.get('descripcion','')}): "
                        f"valor={valor:.3f} SP={sp_act}"
                    )
                    # ← Emitir a TODOS los clientes conectados en tiempo real
                    socketio.emit("new_alarm", {
                        "timestamp":   ts_now,
                        "instrumento": inst,
                        "descripcion": cfg.get("descripcion", ""),
                        "unidad":      cfg.get("unidad", ""),
                        "valor":       round(valor, 4),
                        "nivel":       nivel,
                        "sp_activo":   sp_act,
                    })
                else:  # salió de alarma – registrar retorno a normal
                    db_exec(
                        "INSERT INTO historico_alarmas "
                        "(instrumento, descripcion, unidad, valor, nivel, sp_activo) "
                        "VALUES (%s, %s, %s, %s, %s, %s)",
                        (inst, cfg.get("descripcion", ""), cfg.get("unidad", ""),
                         round(valor, 4), "OK", None),
                        fetch=False
                    )
                    socketio.emit("new_alarm", {
                        "timestamp":   ts_now,
                        "instrumento": inst,
                        "descripcion": cfg.get("descripcion", ""),
                        "unidad":      cfg.get("unidad", ""),
                        "valor":       round(valor, 4),
                        "nivel":       "OK",
                        "sp_activo":   None,
                    })
                _alarm_prev_state[inst] = nivel
    except Exception as exc:
        logger.error(f"Error en _check_and_record_alarmas: {exc}")



def _persist_lecturas(data: dict):
    """Inserta snapshot de proceso en lecturas_proceso (cada ~5 s)."""
    global _ACTIVE_PRUEBA_ID
    mapping = {
        "FI-03": "r_Q_gas_STD",
        "PI-01": "r_P_Gas",
        "TI-01": "r_T_Oil_C",
        "LI-01": "r_LIT_001",
        "PDI-01": "PDI_01",
        "PDI-03": "PDI_03",
        "PDI-02": "r_PDT_02",
        "TI-02": "r_T_Gas",
        "GAS-01": "r_GVoidF",
        "VI-01": "r_v_oil_medida",
        "WC": "r_WC",
        "NIV-AUX": "r_nivel_aux",
        "PDI-04": "r_Transmisor_Gas",
        "PI-02": "r_P_Oil",
        "Coriolis_Density": "Coriolis_Density",
        "Coriolis_Temp": "Coriolis_Temp",
        "Coriolis_Vol_flow_Rate": "Coriolis_Vol_flow_Rate",
        "Q_LIQ": "Q_Liq",
        "Q_CRUDO": "Q_Crudo",
        "Q_NETO": "Q_Neto",
        "Q_DIL": "Q_Dil",
        "Q_AGUA": "Q_Agua",
        "Q_GAS": "Q_Gas"
    }
    vals = []
    for db_tag, data_key in mapping.items():
        v = data.get(data_key, 0)
        try:
            v_f = float(v) if v is not None else 0.0
        except (ValueError, TypeError):
            v_f = 0.0
        vals.extend([db_tag, v_f, _ACTIVE_PRUEBA_ID])
    placeholders = ", ".join(["(%s, %s, %s)"] * len(mapping))
    db_exec(
        f"INSERT INTO lecturas_proceso (instrumento, valor, prueba_id) VALUES {placeholders}",
        tuple(vals), fetch=False
    )


# ─────────────────────────────────────────────────────────────
# Rutas HTTP estáticas
# ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "index.html")


@app.route("/api/status")
def api_status():
    return jsonify({
        "ok":  True,
        "ts":  datetime.now().isoformat(),
        "plc": plc_engine.get_status(),
    })


@app.route("/api/debug_state")
def api_debug_state():
    return jsonify({
        "b_BIT_PROTECTION": bool(V.b_BIT_PROTECTION),
        "b_Laminar": bool(V.b_Laminar),
        "b_Wedge": bool(V.b_Wedge),
        "i_Tipo_medidor": int(V.i_Tipo_medidor),
        "r_miu_Oil": float(V.r_miu_Oil),
        "r_v_oil_medida": float(V.r_v_oil_medida),
        "r_v_oil_calc": float(V.r_v_oil_calc),
        "r_Q_Mezcla_L": float(V.r_Q_Mezcla_L),
        "r_Q_Mezcla_W": float(V.r_Q_Mezcla_W),
        "b_sw_AM_Laminar_Wedge_x": bool(V.b_sw_AM_Laminar_Wedge_x),
        "b_sw_AM_Laminar_Wedge_y": bool(V.b_sw_AM_Laminar_Wedge_y),
        "b_SEL_LAMINAR": bool(V.b_SEL_LAMINAR),
        "r_DP_L": float(V.r_DP_L),
        "r_DP_W": float(V.r_DP_W),
        "r_RE_W": float(V.r_RE_W) if hasattr(V, 'r_RE_W') else 0.0,
        "r_RE_L": float(V.r_RE_L) if hasattr(V, 'r_RE_L') else 0.0,
        "r_PDT_01": float(V.r_PDT_01) if hasattr(V, 'r_PDT_01') else 0.0,
        "r_PDT_02": float(V.r_PDT_02) if hasattr(V, 'r_PDT_02') else 0.0,
        "r_PDT_03": float(V.r_PDT_03) if hasattr(V, 'r_PDT_03') else 0.0,
        "r_P_Oil": float(V.r_P_Oil) if hasattr(V, 'r_P_Oil') else 0.0,
        "r_T_Oil_C": float(V.r_T_Oil_C) if hasattr(V, 'r_T_Oil_C') else 0.0,
        "b_Sel_T_baja": bool(V.b_Sel_T_baja) if hasattr(V, 'b_Sel_T_baja') else False,
        "r_MAX_MIN_TRANSBAJA": float(V.r_MAX_MIN_TRANSBAJA) if hasattr(V, 'r_MAX_MIN_TRANSBAJA') else 0.5,
        
        # Nuevas variables agregadas para depuración e inspección de escenarios
        "r_Q_Liquido_L": float(getattr(V, "r_Q_Liquido_L", 0.0)),
        "r_Qb_Liquido_L": float(getattr(V, "r_Qb_Liquido_L", 0.0)),
        "r_Q_Crudo_L": float(getattr(V, "r_Q_Crudo_L", 0.0)),
        "r_Q_Crudo_sc_L": float(getattr(V, "r_Q_Crudo_sc_L", 0.0)),
        "r_Q_W_L": float(getattr(V, "r_Q_W_L", 0.0)),
        "r_Q_W_sc_L": float(getattr(V, "r_Q_W_sc_L", 0.0)),
        "r_Qb_Liquido_sc_L": float(getattr(V, "r_Qb_Liquido_sc_L", 0.0)),
        
        "r_Q_Liquido": float(getattr(V, "r_Q_Liquido", 0.0)),
        "r_Q_Crudo": float(getattr(V, "r_Q_Crudo", 0.0)),
        "r_Q_W": float(getattr(V, "r_Q_W", 0.0)),
        "r_Qb_Liquido_sc": float(getattr(V, "r_Qb_Liquido_sc", 0.0)),
        "r_Q_Crudo_sc": float(getattr(V, "r_Q_Crudo_sc", 0.0)),
        "r_Q_W_sc": float(getattr(V, "r_Q_W_sc", 0.0)),
        "r_Bo": float(getattr(V, "r_Bo", 1.0)),
        "r_d_m_PT": float(getattr(V, "r_d_m_PT", 0.0)),
        
        # Diluent debug variables
        "r_caudal_dil_BM": float(getattr(V, "r_caudal_dil_BM", 0.0)),
        "r_Q_DIL_MEDIDO": float(getattr(V, "r_Q_DIL_MEDIDO", 0.0)),
        "r_Qb_Dil_Estimado": float(getattr(V, "r_Qb_Dil_Estimado", 0.0)),
        "r_Qb_Dil_Estimado_sc": float(getattr(V, "r_Qb_Dil_Estimado_sc", 0.0)),
        "r_Q_Crudo_Estimado": float(getattr(V, "r_Q_Crudo_Estimado", 0.0)),
        "r_Q_Crudo_Neto_Estimado": float(getattr(V, "r_Q_Crudo_Neto_Estimado", 0.0)),
        "r_Q_Crudo_Neto_Estimado_sc": float(getattr(V, "r_Q_Crudo_Neto_Estimado_sc", 0.0)),
        "r_Vol_dil_total_real": float(getattr(V, "r_Vol_dil_total_real", 0.0)),
        "r_Vol_Dil_Total": float(getattr(V, "r_Vol_Dil_Total", 0.0)),
        "r_Vol_Dil_Total_sc": float(getattr(V, "r_Vol_Dil_Total_sc", 0.0)),
        "b_SW_DIL_MEDIDO_CALC": bool(getattr(V, "b_SW_DIL_MEDIDO_CALC", False)),
        "r_Qb_Liquido_sc_Estimado": float(getattr(V, "r_Qb_Liquido_sc_Estimado", 0.0)),
        "ad_TIEMPO_prueba_7": float(V.ad_TIEMPO_prueba[7]) if hasattr(V, "ad_TIEMPO_prueba") and len(V.ad_TIEMPO_prueba) > 7 else 0.0,
    })


# ─────────────────────────────────────────────────────────────
# API DE COMANDOS MANUALES — Escritura directa en V
# El operador web interactúa con el SoftPLC a través de estas
# rutas. Flask simplemente sobrescribe el valor en la memoria
# compartida; el ScanEngine lo leerá en el próximo ciclo.
# ─────────────────────────────────────────────────────────────

@app.route("/api/pid/<tag>", methods=["GET"])
def get_pid(tag):
    """Devuelve el estado actual de un lazo PID desde la memoria V."""
    tag = tag.upper().replace("_", "-")
    if tag == "LIC-01":
        return jsonify({
            "instrumento": "LIC-01",
            "modo":       "Auto" if not V.b_MAN_LC else "Manual",
            "PV":         round(V.r_LIT_001, 2),
            "SP":         round(V.r_LEVEL_PID_SP, 2),
            "CV":         round(V.fb_LEVEL_PID_r_CVEU, 2),
            "CV_manual":  round(V.r_LEVEL_PID_03_CVOverride, 2),
            "Kp":         V.r_LEVEL_PID_03_KP,
            "Ki":         V.r_LEVEL_PID_03_KI,
            "Kd":         V.r_LEVEL_PID_03_KD,
        })
    elif tag == "PIC-01":
        return jsonify({
            "instrumento": "PIC-01",
            "modo":       "Auto" if not V.b_MAN_PC else "Manual",
            "PV":         round(V.r_P_Gas, 2),
            "SP":         round(V.r_PRESS_PID_SP, 2),
            "CV":         round(V.fb_PRESS_PID_r_CVEU, 2),
            "CV_manual":  round(V.r_PRESS_PID_03_CVOverride, 2),
            "Kp":         V.r_PRESS_PID_03_KP,
            "Ki":         V.r_PRESS_PID_03_KI,
            "Kd":         V.r_PRESS_PID_03_KD,
        })
    return jsonify({"error": "Tag no encontrado. Use LIC-01 o PIC-01"}), 404


@app.route("/api/pid/<tag>", methods=["POST"])
def post_pid(tag):
    """
    Envía comandos de operador al SoftPLC sobrescribiendo V directamente.
    Payload JSON soportado:
      { "modo": "Auto"|"Manual", "SP": float, "CV_manual": float,
        "Kp": float, "Ki": float, "Kd": float }
    """
    tag = tag.upper().replace("_", "-")
    d = request.get_json() or {}

    if tag == "LIC-01":
        if "modo" in d:
            V.b_MAN_LC = (d["modo"] == "Manual")
        if "SP"       in d: V.r_LEVEL_PID_SP              = float(d["SP"])
        if "CV_manual" in d: 
            V.r_LEVEL_PID_03_CVOverride = float(d["CV_manual"])
            V.r_LEVEL_PID_03_CVOper     = float(d["CV_manual"])
        if "Kp"       in d: V.r_LEVEL_PID_03_KP           = float(d["Kp"])
        if "Ki"       in d: V.r_LEVEL_PID_03_KI           = float(d["Ki"])
        if "Kd"       in d: V.r_LEVEL_PID_03_KD           = float(d["Kd"])
        # Persistir en DB para recuperación tras reinicio
        db_exec(
            "UPDATE configuracion_actual SET modo=%s,SP=%s,CV_manual=%s,Kp=%s,Ki=%s,Kd=%s WHERE instrumento=%s",
            ("Manual" if V.b_MAN_LC else "Auto", V.r_LEVEL_PID_SP,
             V.r_LEVEL_PID_03_CVOverride, V.r_LEVEL_PID_03_KP,
             V.r_LEVEL_PID_03_KI, V.r_LEVEL_PID_03_KD, "LIC-01"),
            fetch=False
        )
        return jsonify({"ok": True, "instrumento": "LIC-01",
                        "b_MAN_LC": V.b_MAN_LC, "SP": V.r_LEVEL_PID_SP})

    elif tag == "PIC-01":
        if "modo" in d:
            V.b_MAN_PC = (d["modo"] == "Manual")
        if "SP"       in d: V.r_PRESS_PID_SP              = float(d["SP"])
        if "CV_manual" in d: 
            V.r_PRESS_PID_03_CVOverride = float(d["CV_manual"])
            V.r_PRESS_PID_03_CVOper     = float(d["CV_manual"])
        if "Kp"       in d: V.r_PRESS_PID_03_KP           = float(d["Kp"])
        if "Ki"       in d: V.r_PRESS_PID_03_KI           = float(d["Ki"])
        if "Kd"       in d: V.r_PRESS_PID_03_KD           = float(d["Kd"])
        db_exec(
            "UPDATE configuracion_actual SET modo=%s,SP=%s,CV_manual=%s,Kp=%s,Ki=%s,Kd=%s WHERE instrumento=%s",
            ("Manual" if V.b_MAN_PC else "Auto", V.r_PRESS_PID_SP,
             V.r_PRESS_PID_03_CVOverride, V.r_PRESS_PID_03_KP,
             V.r_PRESS_PID_03_KI, V.r_PRESS_PID_03_KD, "PIC-01"),
            fetch=False
        )
        return jsonify({"ok": True, "instrumento": "PIC-01",
                        "b_MAN_PC": V.b_MAN_PC, "SP": V.r_PRESS_PID_SP})

    return jsonify({"error": "Tag no encontrado. Use LIC-01 o PIC-01"}), 404


@app.route("/api/plc/lazos", methods=["POST"])
def toggle_lazos():
    """Habilita/deshabilita los lazos PID del SoftPLC.
    
    La Fase 4 (p05_main) usa lógica de latch con pushbuttons:
      - b_PB_HABILITA_PID = True  → habilita los lazos (latch set)
      - b_PB_DESHABILITA_PID = True → deshabilita los lazos (latch reset)
    Escribir directamente a b_DESHABILITA_PID no funciona porque
    la Fase 4 lo sobrescribe cada ciclo.
    """
    d = request.get_json(silent=True) or {}
    
    if "habilitar" in d:
        quiere_habilitar = bool(d["habilitar"])
    else:
        # Toggle: si están deshabilitados → habilitar, si habilitados → deshabilitar
        quiere_habilitar = V.b_DESHABILITA_PID  # True = están deshabilitados → quiere habilitar
    
    if quiere_habilitar:
        V.b_PB_HABILITA_PID = True       # Pulso: la Fase 4 lo procesa y lo limpia
        V.b_DESHABILITA_PID = False       # Efecto inmediato para la respuesta
        logger.info("🟢 Lazos PID: pulso HABILITAR enviado")
    else:
        V.b_PB_DESHABILITA_PID = True     # Pulso: la Fase 4 lo procesa y lo limpia
        V.b_DESHABILITA_PID = True        # Efecto inmediato para la respuesta
        logger.info("🔴 Lazos PID: pulso DESHABILITAR enviado")
        
    # Guardar en BD
    try:
        db_exec("UPDATE instrument_selection_config SET b_DESHABILITA_PID=%s WHERE id=1", (V.b_DESHABILITA_PID,), fetch=False)
    except Exception as e:
        logger.error(f"Error guardando b_DESHABILITA_PID en BD: {e}")

    # Guardar variables retenidas
    try:
        from fase1_sistema import save_retained_vars
        save_retained_vars()
    except Exception as e:
        logger.error(f"Error guardando variables retenidas al cambiar lazos: {e}")
    
    return jsonify({"lazos_habilitados": not V.b_DESHABILITA_PID})


@app.route("/api/plc/status", methods=["GET"])
def plc_status():
    """Estado en tiempo real del motor SoftPLC."""
    return jsonify(plc_engine.get_status())


@app.route("/api/instrument_selection", methods=["GET", "POST"])
def instrument_selection():
    if request.method == "POST":
        d = request.get_json() or {}
        if "b_Control_PID_Gas" in d:
            V.b_Control_PID_Gas = bool(d["b_Control_PID_Gas"])
        if "b_PID_POSIC_SW" in d:
            V.b_PID_POSIC_SW = bool(d["b_PID_POSIC_SW"])
        if "b_Sw_Wedge_Gas" in d:
            V.b_Sw_Wedge_Gas = bool(d["b_Sw_Wedge_Gas"])
        if "b_SW_DIL_MEDIDO_CALC" in d:
            V.b_SW_DIL_MEDIDO_CALC = bool(d["b_SW_DIL_MEDIDO_CALC"])
        if "b_Sw_Wedge_Gas_2" in d:
            V.b_Sw_Wedge_Gas_2 = bool(d["b_Sw_Wedge_Gas_2"])
        if "b_SEL_LAMINAR" in d:
            V.b_SEL_LAMINAR = bool(d["b_SEL_LAMINAR"])
        if "b_SEL_T_baja" in d:
            V.b_SEL_T_baja = bool(d["b_SEL_T_baja"])
            V.b_Sel_T_baja = bool(d["b_SEL_T_baja"])
        if "b_sw_AM_Laminar_Wedge_x" in d:
            V.b_sw_AM_Laminar_Wedge_x = bool(d["b_sw_AM_Laminar_Wedge_x"])
        if "b_sw_AM_Laminar_Wedge_y" in d:
            V.b_sw_AM_Laminar_Wedge_y = bool(d["b_sw_AM_Laminar_Wedge_y"])
        if "b_sel_tipo_instrum_dil" in d:
            V.b_sel_tipo_instrum_dil = bool(d["b_sel_tipo_instrum_dil"])
        if "b_AUTO_GAS_01" in d:
            V.b_AUTO_GAS_01 = bool(d["b_AUTO_GAS_01"])
        if "b_SEL_VLV_GAS_01" in d:
            V.b_SEL_VLV_GAS_01 = bool(d["b_SEL_VLV_GAS_01"])
        
        try:
            from fase1_sistema import save_retained_vars
            save_retained_vars()
        except Exception as e:
            logger.error(f"Error guardando variables retenidas en POST /api/instrument_selection: {e}")
            
        try:
            db_exec("""
                UPDATE instrument_selection_config 
                SET b_Control_PID_Gas=%s, b_PID_POSIC_SW=%s, b_Sw_Wedge_Gas=%s, 
                    b_SW_DIL_MEDIDO_CALC=%s, b_Sw_Wedge_Gas_2=%s, b_SEL_LAMINAR=%s, 
                    b_SEL_T_baja=%s, b_sw_AM_Laminar_Wedge_x=%s, b_sw_AM_Laminar_Wedge_y=%s,
                    b_sel_tipo_instrum_dil=%s, b_AUTO_GAS_01=%s, b_SEL_VLV_GAS_01=%s
                WHERE id=1
            """, (V.b_Control_PID_Gas, V.b_PID_POSIC_SW, V.b_Sw_Wedge_Gas, V.b_SW_DIL_MEDIDO_CALC,
                  V.b_Sw_Wedge_Gas_2, V.b_SEL_LAMINAR, V.b_SEL_T_baja, V.b_sw_AM_Laminar_Wedge_x, V.b_sw_AM_Laminar_Wedge_y,
                  V.b_sel_tipo_instrum_dil, V.b_AUTO_GAS_01, V.b_SEL_VLV_GAS_01), fetch=False)
        except Exception as e:
            logger.error(f"Error guardando en BD (instrument_selection_config): {e}")
            
        return jsonify({"ok": True})
        
    return jsonify({
        "b_Control_PID_Gas": bool(V.b_Control_PID_Gas),
        "b_PID_POSIC_SW": bool(V.b_PID_POSIC_SW),
        "b_Sw_Wedge_Gas": bool(V.b_Sw_Wedge_Gas),
        "b_SW_DIL_MEDIDO_CALC": bool(V.b_SW_DIL_MEDIDO_CALC),
        "b_Sw_Wedge_Gas_2": bool(V.b_Sw_Wedge_Gas_2),
        "b_SEL_LAMINAR": bool(V.b_SEL_LAMINAR),
        "b_SEL_T_baja": bool(V.b_SEL_T_baja),
        "b_sw_AM_Laminar_Wedge_x": bool(V.b_sw_AM_Laminar_Wedge_x),
        "b_sw_AM_Laminar_Wedge_y": bool(V.b_sw_AM_Laminar_Wedge_y),
        "b_sel_tipo_instrum_dil": bool(V.b_sel_tipo_instrum_dil),
        "b_AUTO_GAS_01": bool(V.b_AUTO_GAS_01),
        "b_SEL_VLV_GAS_01": bool(V.b_SEL_VLV_GAS_01),
    })


@app.route("/api/plc/simulacion", methods=["POST"])
def toggle_simulacion():
    """Activa/desactiva el modo simulación de entradas analógicas en V."""
    d = request.get_json() or {}
    V.b_simular_ai = bool(d.get("simular", not V.b_simular_ai))
    return jsonify({"b_simular_ai": V.b_simular_ai})



# ─────────────────────────────────────────────────────────────
# Calibración de Medidores
# ─────────────────────────────────────────────────────────────

# Tags editables por sección.  Los de solo-lectura se incluyen
# en la respuesta GET pero se ignoran en el POST.
_CALIB_WRITABLE = {
    # Wedge Gas
    "r_D_wedge_gas", "r_h_wedge_gas", "r_k_mp",
    # Wedge Crudo
    "r_D_Wedge", "r_m", "r_K_wedge",
    # Laminar
    "r_d_L", "r_L", "r_N_Tubos",
    "r_AK_L", "r_BK_L", "r_CK_L",
}

@app.route("/api/calibracion", methods=["GET"])
def get_calibracion():
    """Devuelve todos los parámetros de calibración de los medidores."""
    return jsonify({
        # ── Wedge Gas ──
        "r_D_wedge_gas":  getattr(V, "r_D_wedge_gas",  0.0),
        "r_h_wedge_gas":  getattr(V, "r_h_wedge_gas",  0.0),
        "r_k_mp":         getattr(V, "r_k_mp",          0.0),
        "r_DP_gas_PK":    round(getattr(V, "r_DP_gas_PK",   0.0), 3),  # solo lectura
        "r_Beta_mp":      round(getattr(V, "r_Beta_mp",     0.0), 4),  # solo lectura
        "r_Ao_cd":        round(getattr(V, "r_Ao_cd",       0.0), 6),  # solo lectura
        "r_Y1":           round(getattr(V, "r_Y1",          0.0), 4),  # solo lectura
        # ── Wedge Crudo ──
        "r_D_Wedge":      getattr(V, "r_D_Wedge",  0.0),
        "r_m":            getattr(V, "r_m",          0.0),
        "r_K_wedge":      getattr(V, "r_K_wedge",   1.0),
        "r_PDT_02":       round(getattr(V, "r_PDT_02",  0.0), 3),      # solo lectura
        "r_RE_W":         round(getattr(V, "r_RE_W",    0.0), 1),      # solo lectura
        "r_Qb_Liquido_W": round(getattr(V, "r_Qb_Liquido_W", 0.0), 1),# solo lectura
        # ── Laminar ──
        "r_d_L":          getattr(V, "r_d_L",    0.0),
        "r_L":            getattr(V, "r_L",      0.0),
        "r_N_Tubos":      getattr(V, "r_N_Tubos", 0.0),
        "r_AK_L":         getattr(V, "r_AK_L",   0.0),
        "r_BK_L":         getattr(V, "r_BK_L",   0.0),
        "r_CK_L":         getattr(V, "r_CK_L",   0.0),
        "r_PDT_03":       round(getattr(V, "r_PDT_03", 0.0), 3),       # solo lectura
        "r_PDT_01":       round(getattr(V, "r_PDT_01", 0.0), 3),       # solo lectura
        "r_RE_L":         round(getattr(V, "r_RE_L",   0.0), 1),       # solo lectura
        "r_Qb_Liquido_L": round(getattr(V, "r_Qb_Liquido_L", 0.0), 1),# solo lectura
    })


@app.route("/api/calibracion", methods=["POST"])
def set_calibracion():
    """Escribe los parámetros de calibración editables directamente en V."""
    d = request.get_json() or {}
    updated = []
    for key, val in d.items():
        if key in _CALIB_WRITABLE and hasattr(V, key):
            try:
                setattr(V, key, float(val))
                updated.append(key)
            except (TypeError, ValueError):
                pass
    if updated:
        try:
            from fase1_sistema import save_retained_vars
            save_retained_vars()
        except Exception as e:
            logger.error(f"Error saving calibration params to disk: {e}")
    return jsonify({"ok": True, "updated": updated})


# ─────────────────────────────────────────────────────────────
# Propiedades Físicas del Fluido
# ─────────────────────────────────────────────────────────────

@app.route("/api/propiedades", methods=["GET"])
def get_propiedades():
    return jsonify({
        "densidadRefDiluente": float(getattr(V, "r_d_D_ref",    0.0)),
        "densidadRefCrudo":    float(getattr(V, "r_d_Oil_ref",  0.0)),
        "gravEspGas":          float(getattr(V, "r_yg",          0.0)),
        "presionAtm":          float(getattr(V, "r_PA",          0.0)),
        "constanteGas":        float(getattr(V, "r_R_gas",       8.314)),
        "presionCriticaGas":   float(getattr(V, "r_Pc_Gas",      0.0)),
        "A": float(getattr(V, "r_A_ds",    0.0)),
        "B": float(getattr(V, "r_B_ds",    0.0)),
        "C": float(getattr(V, "r_C_ds",    0.0)),
        "D": float(getattr(V, "r_D_ds",    0.0)),
        "E": float(getattr(V, "r_E_ds",    1.0)),
        "Z": float(getattr(V, "r_Z_Gas_P", 1.0)),
        "densidadGas": float(getattr(V, "r_d_Gas",   0.0)),
        "laminar":     float(getattr(V, "r_RE_L_M",  0.0)),
        "wedge":       float(getattr(V, "r_RE_W_M",  0.0)),
    })


@app.route("/api/propiedades", methods=["POST"])
def set_propiedades():
    d = request.get_json() or {}
    MAP = {
        "densidadRefDiluente": "r_d_D_ref",
        "densidadRefCrudo":    "r_d_Oil_ref",
        "gravEspGas":          "r_yg",
        "presionAtm":          "r_PA",
        "constanteGas":        "r_R_gas",
        "presionCriticaGas":   "r_Pc_Gas",
        "A": "r_A_ds", "B": "r_B_ds", "C": "r_C_ds",
        "D": "r_D_ds", "E": "r_E_ds",
        "densidadGas": "r_d_Gas",
        "laminar": "r_RE_L_M",
        "wedge":    "r_RE_W_M",
    }
    updated = []
    for fe_key, v_key in MAP.items():
        if fe_key in d and hasattr(V, v_key):
            try:
                setattr(V, v_key, float(d[fe_key]))
                updated.append(v_key)
            except (TypeError, ValueError):
                pass

    # Sincronizar variables físicas dependientes para evitar sobrescritura en el ciclo del PLC
    if "densidadRefDiluente" in d:
        try:
            dens_dil = float(d["densidadRefDiluente"])
            if dens_dil > 0:
                s_d_ref = dens_dil / getattr(V, "r_d_W_ref", 0.9990121)
                api_1 = (141.5 / s_d_ref) - 131.5
                setattr(V, "r_S_D_ref", s_d_ref)
                setattr(V, "r_API_1", api_1)
                updated.append("r_S_D_ref")
                updated.append("r_API_1")
        except (TypeError, ValueError):
            pass

    if "densidadRefCrudo" in d:
        try:
            dens_oil = float(d["densidadRefCrudo"])
            if dens_oil > 0:
                s_oil_ref = dens_oil / 0.9990121
                api_2 = (141.5 / s_oil_ref) - 131.5
                setattr(V, "r_S_Oil_ref", s_oil_ref)
                setattr(V, "r_API_2", api_2)
                updated.append("r_S_Oil_ref")
                updated.append("r_API_2")
        except (TypeError, ValueError):
            pass

    if updated:
        # Eliminar posibles duplicados manteniendo el orden
        seen = set()
        updated = [x for x in updated if not (x in seen or seen.add(x))]
        try:
            from fase1_sistema import save_retained_vars
            save_retained_vars()
        except Exception as e:
            logger.error(f"Error saving properties: {e}")
        for key in updated:
            if hasattr(V, key):
                val = float(getattr(V, key))
                try:
                    db_exec(
                        "INSERT INTO propiedades_config (parametro, valor) VALUES (%s, %s) "
                        "ON DUPLICATE KEY UPDATE valor=%s",
                        (key, val, val), fetch=False
                    )
                except Exception as e:
                    logger.error(f"Error guardando propiedad {key} en BD: {e}")
    return jsonify({"ok": True, "updated": updated})


# ─────────────────────────────────────────────────────────────
# Selección de Fórmulas de Cálculo
# ─────────────────────────────────────────────────────────────

@app.route("/api/formulas", methods=["GET"])
def get_formulas():
    return jsonify({
        "b_IHM_PB_miu":  int(bool(getattr(V, "b_IHM_PB_miu",  False))),
        "b_externa":     int(bool(getattr(V, "b_externa",     False))),
        "b_SEL_LAMINAR": int(bool(getattr(V, "b_SEL_LAMINAR", False))),
        "b_PB_PVT":      int(bool(getattr(V, "b_PB_PVT",      False))),
    })


@app.route("/api/formulas", methods=["POST"])
def set_formulas():
    d = request.get_json() or {}
    updated = []
    for key in ["b_IHM_PB_miu", "b_externa", "b_SEL_LAMINAR", "b_PB_PVT"]:
        if key in d:
            try:
                setattr(V, key, bool(int(d[key])))
                updated.append(key)
                if key == "b_PB_PVT":
                    mode_val = int(d[key])
                    try:
                        db_exec("UPDATE pvt_balance_config SET pvt_mode = %s WHERE id = 1", (mode_val,), fetch=False)
                    except Exception as _e:
                        pass
                    if mode_val == 1:
                        V.r_Rso_PT = getattr(V, "r_Rso_PT2", V.r_Rso_PT)
                        V.r_Bo = getattr(V, "r_Bo2", V.r_Bo)
            except (TypeError, ValueError):
                pass
    if updated:
        try:
            from fase1_sistema import save_retained_vars
            save_retained_vars()
        except Exception as e:
            logger.error(f"Error saving formulas: {e}")
    return jsonify({"ok": True, "updated": updated})


# ─────────────────────────────────────────────────────────────
# Cálculos PVT y Balance de Masa (Persistencia en BD)
# ─────────────────────────────────────────────────────────────

@app.route("/api/pvt-balance", methods=["GET"])
def get_pvt_balance():
    """Retorna la configuración y cálculos guardados de PVT y Balance de Masa."""
    try:
        rows = db_exec("SELECT * FROM pvt_balance_config WHERE id = 1")
        if rows:
            r = rows[0]
            return jsonify({
                "ok": True,
                "pvtMode": int(r.get("pvt_mode", 0)),
                "tempYac": float(r.get("temp_yac", 0.0)),
                "rso": float(r.get("rso", 0.0)),
                "bo": float(r.get("bo", 1.0)),
                "apiForm_real": float(r.get("api_form_real", 0.0)),
                "apiForm_teo": float(r.get("api_form_teo", 0.0)),
                "apiMez_real": float(r.get("api_mez_real", 0.0)),
                "apiMez_teo": float(r.get("api_mez_teo", 0.0)),
                "apiDil_real": float(r.get("api_dil_real", 0.0)),
                "apiDil_teo": float(r.get("api_dil_teo", 0.0)),
                "qDil_real": float(r.get("q_dil_real", 0.0)),
                "qDil_teo": float(r.get("q_dil_teo", 0.0)),
                "qNet_real": float(r.get("q_net_real", 0.0)),
                "qNet_teo": float(r.get("q_net_teo", 0.0)),
                "qNetDil_real": float(r.get("q_net_dil_real", 0.0)),
                "qNetDil_teo": float(r.get("q_net_dil_teo", 0.0)),
                "qAgua_real": float(r.get("q_agua_real", 0.0)),
                "qAgua_teo": float(r.get("q_agua_teo", 0.0)),
                "qTotal_real": float(r.get("q_total_real", 0.0)),
                "qTotal_teo": float(r.get("q_total_teo", 0.0)),
            })
    except Exception as e:
        logger.error(f"Error reading pvt_balance_config from BD: {e}")

    # Fallback si no hay BD o error
    return jsonify({
        "ok": True,
        "pvtMode": 1 if getattr(V, "b_PB_PVT", False) else 0,
        "tempYac": float(getattr(V, "r_T_Yac_C", 0.0)),
        "rso": float(getattr(V, "r_Rso_PT2", 0.0)),
        "bo": float(getattr(V, "r_Bo2", 1.0)),
        "apiForm_real": float(getattr(V, "r_API_formacion_BM", 0.0)),
        "apiForm_teo": float(getattr(V, "r_API_formacion_BM", 0.0)),
        "apiMez_real": float(getattr(V, "r_API_2", 0.0)),
        "apiMez_teo": float(getattr(V, "r_API_MEZCLA_TEORICO", 0.0)),
        "apiDil_real": float(getattr(V, "r_API_1", 0.0)),
        "apiDil_teo": float(getattr(V, "r_API_1", 0.0)),
        "qDil_real": float(getattr(V, "r_caudal_dil_BM", 0.0)),
        "qDil_teo": float(getattr(V, "r_caudal_dil_BM", 0.0)),
        "qNet_real": float(getattr(V, "r_Q_Crudo", 0.0)),
        "qNet_teo": float(getattr(V, "r_CAUDAL_NETO_TEORICO", 0.0)),
        "qNetDil_real": float(getattr(V, "r_Q_Crudo", 0.0)),
        "qNetDil_teo": float(getattr(V, "r_Q_Crudo", 0.0)),
        "qAgua_real": float(getattr(V, "r_Q_W", 0.0)),
        "qAgua_teo": float(getattr(V, "r_Q_W", 0.0)),
        "qTotal_real": float(getattr(V, "r_Q_Liquido", 0.0)),
        "qTotal_teo": float(getattr(V, "r_Q_Liquido", 0.0)),
    })


@app.route("/api/pvt-balance", methods=["POST"])
def set_pvt_balance():
    """Guarda en BD y actualiza en memoria los cálculos y valores ingresados de PVT y Balance de Masa."""
    d = request.get_json() or {}
    try:
        field_map = {
            "pvtMode": "pvt_mode",
            "tempYac": "temp_yac",
            "rso": "rso",
            "bo": "bo",
            "apiForm_real": "api_form_real",
            "apiForm_teo": "api_form_teo",
            "apiMez_real": "api_mez_real",
            "apiMez_teo": "api_mez_teo",
            "apiDil_real": "api_dil_real",
            "apiDil_teo": "api_dil_teo",
            "qDil_real": "q_dil_real",
            "qDil_teo": "q_dil_teo",
            "qNet_real": "q_net_real",
            "qNet_teo": "q_net_teo",
            "qNetDil_real": "q_net_dil_real",
            "qNetDil_teo": "q_net_dil_teo",
            "qAgua_real": "q_agua_real",
            "qAgua_teo": "q_agua_teo",
            "qTotal_real": "q_total_real",
            "qTotal_teo": "q_total_teo",
        }

        updates = []
        params = []
        for json_key, db_col in field_map.items():
            if json_key in d and d[json_key] is not None:
                try:
                    val = int(d[json_key]) if json_key == "pvtMode" else float(d[json_key])
                    updates.append(f"{db_col} = %s")
                    params.append(val)
                except (ValueError, TypeError):
                    pass

        # Sincronizar variables de SoftPLC en memoria
        if "pvtMode" in d:
            try:
                mode_val = int(d["pvtMode"])
                V.b_PB_PVT = bool(mode_val == 1)
            except (ValueError, TypeError):
                pass

        if "rso" in d:
            try:
                V.r_Rso_PT2 = float(d["rso"])
                if V.b_PB_PVT:
                    V.r_Rso_PT = V.r_Rso_PT2
            except (ValueError, TypeError):
                pass

        if "bo" in d:
            try:
                val_bo = float(d["bo"])
                V.r_Bo2 = val_bo if val_bo > 0 else 1.0
                if V.b_PB_PVT:
                    V.r_Bo = V.r_Bo2
            except (ValueError, TypeError):
                pass

        if "tempYac" in d:
            try:
                ty = float(d["tempYac"])
                if ty > 0:
                    setattr(V, "r_T_Yac_C", ty)
            except (ValueError, TypeError):
                pass

        if updates:
            # Asegurar que el registro id=1 exista en la tabla
            db_exec("INSERT INTO pvt_balance_config (id) VALUES (1) ON DUPLICATE KEY UPDATE id=1", fetch=False)
            set_clause = ", ".join(updates)
            sql = f"UPDATE pvt_balance_config SET {set_clause} WHERE id = 1"
            db_exec(sql, tuple(params), fetch=False)
            logger.info("💾 [PVT / Balance] Datos actualizados y guardados exitosamente en BD MySQL")

            try:
                from fase1_sistema import save_retained_vars
                save_retained_vars()
            except Exception as e:
                logger.error(f"Error saving retained vars from pvt-balance: {e}")

        return jsonify({"ok": True, "message": "Datos de PVT y Balance guardados correctamente en BD"})
    except Exception as e:
        logger.error(f"Error in set_pvt_balance: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500



# ─────────────────────────────────────────────────────────────
# Control de Prueba
# ─────────────────────────────────────────────────────────────

@app.route("/api/plc/prueba/iniciar", methods=["POST"])
def iniciar_prueba():
    d = request.get_json() or {}
    V.b_PB_inicio_prueba  = True
    V.b_PB_parada_prueba  = False
    V.b_IHM_Abortar_Prueba = False
    if "codigo_pozo" in d:
        if hasattr(V, 'as_Codigo_pozo_16'):
            V.as_Codigo_pozo_16 = str(d["codigo_pozo"])
    if "duracion_horas" in d:
        try:
            horas = float(d["duracion_horas"])
            if hasattr(V, 'ad_IHM_TIEMPO_prueba'):
                V.ad_IHM_TIEMPO_prueba[3] = horas
                V.ad_IHM_TIEMPO_prueba[2] = horas * 3600.0
        except Exception:
            pass
    return jsonify({
        "ok": True,
        "b_Prueba_en_Progreso": bool(getattr(V, "b_Prueba_en_Progreso", False)),
    })


@app.route("/api/plc/prueba/parar", methods=["POST"])
def parar_prueba():
    V.b_PB_parada_prueba = True
    V.b_IHM_Validar_Prueba = True  # Valida automáticamente para cerrar la parada
    return jsonify({"ok": True})


@app.route("/api/plc/prueba/abortar", methods=["POST"])
def abortar_prueba():
    V.b_IHM_Abortar_Prueba = True
    if hasattr(V, 'b_Prueba_en_Progreso'):
        V.b_Prueba_en_Progreso = False
    V.b_PB_inicio_prueba = False
    V.b_Parada_en_Progreso = False  # Libera la parada para permitir nuevas pruebas
    return jsonify({"ok": True})


@app.route("/api/plc/prueba/cargar_datos", methods=["POST"])
def cargar_datos_prueba():
    """Escribe los datos del formulario 'Cargar Datos Prueba' al objeto V del SoftPLC."""
    d = request.get_json() or {}
    updated = []

    # ── Cadenas de texto (as_Codigo_pozo_XX) ────────────────
    str_map = {
        "lugar":   "as_Codigo_pozo_17",
        "pozo":    "as_Codigo_pozo_03",
        "metodo":  "as_Codigo_pozo_06",
        "rpm":     "as_Codigo_pozo_08",
        "inyeccion": "as_Codigo_pozo_18",
    }
    for key, tag in str_map.items():
        if key in d and hasattr(V, tag):
            setattr(V, tag, str(d[key]))
            updated.append(tag)

    # ── Valores flotantes ────────────────────────────────────
    float_map = {
        "tempYac":       "r_T_Yac_C",
        "apiFormacion":  "r_API_formacion_BM",
        "apiMezcla":     "r_API_2",
        "apiDiluente":   "r_API_1",
        "caudalDiluente":"r_caudal_dil_BM",
    }
    for key, tag in float_map.items():
        if key in d and hasattr(V, tag):
            try:
                setattr(V, tag, float(d[key]))
                updated.append(tag)
            except (TypeError, ValueError):
                pass

    # ── Enteros (duración y combo boxes) ────────────────────
    if "duracionHoras" in d and hasattr(V, "i_duracion_prueba_horas"):
        try:
            V.i_duracion_prueba_horas = int(float(d["duracionHoras"]))
            updated.append("i_duracion_prueba_horas")
        except (TypeError, ValueError):
            pass

    if "comboMetodo" in d and hasattr(V, "i_posicion_combo_box_1"):
        try:
            V.i_posicion_combo_box_1 = int(d["comboMetodo"])
            updated.append("i_posicion_combo_box_1")
        except (TypeError, ValueError):
            pass

    if "comboInyeccion" in d and hasattr(V, "i_posicion_combo_box_2"):
        try:
            V.i_posicion_combo_box_2 = int(d["comboInyeccion"])
            updated.append("i_posicion_combo_box_2")
        except (TypeError, ValueError):
            pass

    # ── Fecha/Hora de inicio (ad_IHM_HORA_inicio) ───────────
    # d["fechaDD"], d["fechaMM"], d["fechaAAAA"], d["horaHH"], d["horaMM"]
    if hasattr(V, "ad_IHM_HORA_inicio"):
        try:
            arr = list(getattr(V, "ad_IHM_HORA_inicio", [0]*8))
            if "fechaDD"   in d: arr[2] = int(d["fechaDD"])
            if "fechaMM"   in d: arr[1] = int(d["fechaMM"])
            if "fechaAAAA" in d: arr[0] = int(d["fechaAAAA"])
            if "horaHH"    in d: arr[3] = int(d["horaHH"])
            if "horaMM"    in d: arr[4] = int(d["horaMM"])
            V.ad_IHM_HORA_inicio = arr
            updated.append("ad_IHM_HORA_inicio")
        except (TypeError, ValueError, IndexError):
            pass

    # ── Guardar en base de datos (upsert id=1) ───────────────
    try:
        current_rows = db_exec("SELECT * FROM prueba_configuracion WHERE id = 1")
        current = current_rows[0] if current_rows else {}
        
        lugar = d.get("lugar", current.get("lugar", ""))
        pozo = d.get("pozo", current.get("pozo", ""))
        metodo = d.get("metodo", current.get("metodo", ""))
        rpm = d.get("rpm", current.get("rpm", ""))
        inyeccion = d.get("inyeccion", current.get("inyeccion", ""))
        combo_inyeccion = int(d.get("comboInyeccion", current.get("combo_inyeccion", 0)))
        
        temp_yac = float(d.get("tempYac", current.get("temp_yac", 0.0)))
        api_formacion = float(d.get("apiFormacion", current.get("api_formacion", 0.0)))
        api_mezcla = float(d.get("apiMezcla", current.get("api_mezcla", 0.0)))
        api_diluente = float(d.get("apiDiluente", current.get("api_diluente", 0.0)))
        if combo_inyeccion == 0 or str(inyeccion).strip().upper() == "NO":
            caudal_diluente = 0.0
            if hasattr(V, "r_caudal_dil_BM"):
                V.r_caudal_dil_BM = 0.0
        else:
            caudal_diluente = float(d.get("caudalDiluente", current.get("caudal_diluente", 0.0)))
        
        duracion_horas = int(float(d.get("duracionHoras", current.get("duracion_horas", 0))))
        combo_metodo = int(d.get("comboMetodo", current.get("combo_metodo", 0)))
        
        fecha_dd = int(d.get("fechaDD", current.get("fecha_dd", 0)))
        fecha_mm = int(d.get("fechaMM", current.get("fecha_mm", 0)))
        fecha_aaaa = int(d.get("fechaAAAA", current.get("fecha_aaaa", 0)))
        hora_hh = int(d.get("horaHH", current.get("hora_hh", 0)))
        hora_mm = int(d.get("horaMM", current.get("hora_mm", 0)))
        
        db_exec("""
            INSERT INTO prueba_configuracion (
                id, lugar, pozo, metodo, rpm, inyeccion, temp_yac, api_formacion,
                api_mezcla, api_diluente, caudal_diluente, duracion_horas,
                combo_metodo, combo_inyeccion, fecha_dd, fecha_mm, fecha_aaaa,
                hora_hh, hora_mm
            ) VALUES (1, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                lugar=%s, pozo=%s, metodo=%s, rpm=%s, inyeccion=%s,
                temp_yac=%s, api_formacion=%s, api_mezcla=%s, api_diluente=%s, caudal_diluente=%s,
                duracion_horas=%s, combo_metodo=%s, combo_inyeccion=%s,
                fecha_dd=%s, fecha_mm=%s, fecha_aaaa=%s, hora_hh=%s, hora_mm=%s
        """, (
            lugar, pozo, metodo, rpm, inyeccion, temp_yac, api_formacion,
            api_mezcla, api_diluente, caudal_diluente, duracion_horas,
            combo_metodo, combo_inyeccion, fecha_dd, fecha_mm, fecha_aaaa,
            hora_hh, hora_mm,
            
            lugar, pozo, metodo, rpm, inyeccion, temp_yac, api_formacion,
            api_mezcla, api_diluente, caudal_diluente, duracion_horas,
            combo_metodo, combo_inyeccion, fecha_dd, fecha_mm, fecha_aaaa,
            hora_hh, hora_mm
        ), fetch=False)
    except Exception as ex_db:
        logger.error(f"Error guardando configuracion de prueba en BD: {ex_db}")

    return jsonify({"ok": True, "updated": updated})

@app.route('/api/plc/prueba/vaciar', methods=['POST'])
def vaciar_datos_prueba():
    updated = []
    # Vaciar cadenas
    for k in ["as_Codigo_pozo_17", "as_Codigo_pozo_03", "as_Codigo_pozo_06", 
              "as_Codigo_pozo_08", "as_Codigo_pozo_18", "as_Codigo_pozo_19", "as_Codigo_pozo_16"]:
        if hasattr(V, k):
            setattr(V, k, "")
            updated.append(k)
    # Vaciar floats y enteros
    for k in ["r_T_Yac_C", "r_API_formacion_BM", "r_API_2", "r_API_1", 
              "r_caudal_dil_BM", "i_duracion_prueba_horas", 
              "i_posicion_combo_box_1", "i_posicion_combo_box_2"]:
        if hasattr(V, k):
            setattr(V, k, 0.0)
            updated.append(k)
    # Vaciar hora
    if hasattr(V, "ad_IHM_HORA_inicio"):
        V.ad_IHM_HORA_inicio = [0]*8
        updated.append("ad_IHM_HORA_inicio")
    if hasattr(V, "ad_TIEMPO_inicio_prueba"):
        V.ad_TIEMPO_inicio_prueba = [0]*8
        updated.append("ad_TIEMPO_inicio_prueba")
    if hasattr(V, "ar_TIEMPO_prueba_TOTAL"):
        V.ar_TIEMPO_prueba_TOTAL = [0]*10
        updated.append("ar_TIEMPO_prueba_TOTAL")
        
    try:
        db_exec("DELETE FROM prueba_configuracion WHERE id = 1", fetch=False)
    except Exception as ex:
        logger.error(f"Error vaciando configuracion de prueba en BD: {ex}")

    return jsonify({"ok": True, "updated": updated})



# ─────────────────────────────────────────────────────────────
# Alarmas
# ─────────────────────────────────────────────────────────────

@app.route("/api/alarmas")
def get_alarmas():
    rows = db_exec("SELECT * FROM tabla_configuracion_alarma ORDER BY instrumento")
    return jsonify(rows or [])


@app.route("/api/alarmas/<instrumento>", methods=["GET"])
def get_alarma(instrumento):
    rows = db_exec("SELECT * FROM tabla_configuracion_alarma WHERE instrumento=%s",
                   (instrumento.upper(),))
    return jsonify(rows[0] if rows else {}), (200 if rows else 404)


_INSTRUMENT_MAPPING = {
    "LI-01": {
        "daq": ["r_Local_2_I_Ch0Data"],
        "hart": ["NIVEL"]
    },
    "PDI-01": {
        "daq": ["r_Local_2_I_Ch1Data"],
        "hart": ["LAMINAR_A"]
    },
    "PDI-03": {
        "daq": ["r_Local_2_I_Ch3Data"],
        "hart": ["LAMINAR_B"]
    },
    "PDI-02": {
        "daq": ["r_Local_4_I_Ch0Data"],
        "hart": ["WEDGE_LIQ"]
    },
    "PI-01": {
        "daq": ["r_Local_4_I_Ch4Data"],
        "hart": ["WEDGE_GAS"]
    },
    "TI-01": {
        "daq": ["r_Local_4_I_Ch2Data"],
        "hart": ["WEDGE_LIQ"]
    },
    "TI-02": {
        "daq": ["r_Local_4_I_Ch5Data"],
        "hart": ["WEDGE_GAS"]
    },
    "FI-03": {
        "daq": ["r_Local_2_I_Ch2Data"],
        "hart": []
    },
    "WC": {
        "daq": ["r_Local_4_I_Ch7Data"],
        "hart": []
    }
}


def _sync_daq_hart_enabled_state(instrumento, modo_manual):
    """
    Sincroniza el estado de habilitación de canales DAQ y HART cuando un instrumento
    pasa a manual o a automático, dependiendo de la comunicación que esté activa.
    """
    mapping = _INSTRUMENT_MAPPING.get(instrumento)
    if not mapping:
        return

    enabled_val = 0 if modo_manual else 1

    # Determinar si el tag está usando HART o DAQ actualmente
    uses_hart = False

    # Caso 1: Instrumentos que solo existen en HART
    if instrumento in ["PI-01", "TI-02"]:
        uses_hart = True
    # Caso 2: Instrumentos que solo existen en DAQ
    elif instrumento in ["FI-03", "WC"]:
        uses_hart = False
    # Caso 3: Instrumentos híbridos que dependen de V.b_habilitar_F_HART
    else:
        uses_hart = bool(getattr(V, "b_habilitar_F_HART", False))

    if uses_hart:
        hart_types = mapping.get("hart", [])
        if hart_types:
            placeholders = ",".join(["%s"] * len(hart_types))
            db_exec(
                f"UPDATE hart_channel_config SET enabled = %s WHERE instrument_type IN ({placeholders})",
                [enabled_val] + hart_types,
                fetch=False
            )
    else:
        daq_vars = mapping.get("daq", [])
        if daq_vars:
            placeholders = ",".join(["%s"] * len(daq_vars))
            db_exec(
                f"UPDATE daq_channel_config SET enabled = %s WHERE v_name IN ({placeholders})",
                [enabled_val] + daq_vars,
                fetch=False
            )
            # Recargar DAQ en caliente
            _load_daq_channels_from_db()


@app.route("/api/alarmas/<instrumento>", methods=["POST"])
def post_alarma(instrumento):
    d = request.get_json() or {}
    inst_upper = instrumento.upper()
    modo_manual = int(bool(d.get("modo_manual", 0)))
    
    def _flt(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    val_man = _flt(d.get("valor_manual")) if modo_manual else None

    db_exec(
        """UPDATE tabla_configuracion_alarma
           SET minimo=%s,maximo=%s,SP_HH=%s,SP_H=%s,SP_L=%s,SP_LL=%s,
               DB=%s,RAW_H=%s,RAW_L=%s,modo_manual=%s,valor_manual=%s
           WHERE instrumento=%s OR UPPER(instrumento)=%s""",
        (_flt(d.get("minimo")), _flt(d.get("maximo")), _flt(d.get("SP_HH")), _flt(d.get("SP_H")),
         _flt(d.get("SP_L")), _flt(d.get("SP_LL")), _flt(d.get("DB")), _flt(d.get("RAW_H")), _flt(d.get("RAW_L")),
         modo_manual, val_man,
         instrumento, inst_upper),
        fetch=False
    )
    
    # Sincronizar habilitación en Config DAQ y Config HART
    _sync_daq_hart_enabled_state(inst_upper, modo_manual)
    
    # Si modo manual activo, escribir valor en V inmediatamente
    _apply_manual_override(instrumento, modo_manual, val_man)

    # Si es variable Modbus RTU / Coriolis, sincronizar también en modbus_rtu_variables
    if modo_manual and val_man is not None:
        db_exec(
            "UPDATE modbus_rtu_variables SET current_val=%s, updated_at=NOW() WHERE var_name=%s OR UPPER(var_name)=%s",
            (val_man, instrumento, inst_upper),
            fetch=False
        )
    
    # Sincronizar fallas de presion (r_falla_presion_gas, r_falla_presion_crudo)
    if inst_upper in ('PI-01', 'PI-02'):
        _sync_falla_presion()
        
    return jsonify({"ok": True})

def _sync_falla_presion():
    """Sincroniza V.r_falla_presion_gas y crudo con el SP_HH de PI-01 y PI-02"""
    try:
        rows = db_exec("SELECT instrumento, SP_HH FROM tabla_configuracion_alarma WHERE instrumento IN ('PI-01', 'PI-02')")
        for r in (rows or []):
            if r["instrumento"] == "PI-01" and r.get("SP_HH") is not None:
                V.r_falla_presion_gas = float(r["SP_HH"])
            elif r["instrumento"] == "PI-02" and r.get("SP_HH") is not None:
                V.r_falla_presion_crudo = float(r["SP_HH"])
    except Exception as e:
        logger.warning(f"Error sincronizando fallas de presion: {e}")


def _apply_manual_override(instrumento, modo_manual, valor_manual):
    """Guarda/Escribe el valor manual en V.instrument_overrides si modo_manual=1."""
    # Mapa instrumento → atributo en V
    _TAG_MAP = {
        "FI-03":                  "r_Q_gas_STD",
        "GAS-01":                 "r_GVoidF",
        "LI-01":                  "r_LIT_001",
        "PDI-01":                 "r_PDT_01",
        "PDI-02":                 "r_PDT_02",
        "PDI-03":                 "r_PDT_03",
        "PDI-04":                 "r_Transmisor_Gas",
        "PI-01":                  "r_P_Gas",
        "PI-02":                  "r_P_Oil",
        "TI-01":                  "r_T_Oil_C",
        "TI-02":                  "r_T_Gas",
        "VI-01":                  "r_v_oil_medida",
        "WC":                     "r_WC",
        "NIV-AUX":                "r_nivel_aux",
        "CORIOLIS_DENSITY":       "Coriolis_Density",
        "CORIOLIS_TEMP":          "Coriolis_Temp",
        "CORIOLIS_VOL_FLOW_RATE": "Coriolis_Vol_flow_Rate",
        "CORIOLIS_VOL_FLOW_RA":   "Coriolis_Vol_flow_Rate",
        "Coriolis_Density":       "Coriolis_Density",
        "Coriolis_Temp":          "Coriolis_Temp",
        "Coriolis_Vol_flow_Rate": "Coriolis_Vol_flow_Rate",
        "Coriolis_Vol_flow_Ra":   "Coriolis_Vol_flow_Rate",
    }
    tag = _TAG_MAP.get(instrumento) or _TAG_MAP.get(str(instrumento).upper()) or _TAG_MAP.get(str(instrumento).lower())
    if tag:
        if not hasattr(V, 'instrument_overrides'):
            V.instrument_overrides = {}
        if modo_manual and valor_manual is not None:
            try:
                val_f = float(valor_manual)
                V.instrument_overrides[tag] = val_f
                if hasattr(V, tag):
                    setattr(V, tag, val_f)
            except Exception:
                pass
        else:
            V.instrument_overrides.pop(tag, None)
            if hasattr(V, tag):
                setattr(V, tag, 0.0)



# ─────────────────────────────────────────────────────────────
# Histórico de Alarmas
# ─────────────────────────────────────────────────────────────

@app.route("/api/historico_alarmas", methods=["GET"])
def get_historico_alarmas():
    """Consulta el histórico de alarmas con filtro de fecha."""
    inicio = request.args.get("inicio")
    fin    = request.args.get("fin")
    inst   = request.args.get("instrumento")   # opcional
    nivel  = request.args.get("nivel")          # opcional: HH, H, L, LL, OK

    q = "SELECT * FROM historico_alarmas WHERE 1=1"
    params = []
    if inicio and fin:
        q += " AND timestamp BETWEEN %s AND %s"
        params += [inicio, fin]
    if inst:
        q += " AND instrumento = %s"
        params.append(inst)
    if nivel:
        q += " AND nivel = %s"
        params.append(nivel)
    q += " ORDER BY timestamp DESC LIMIT 5000"

    rows = db_exec(q, tuple(params)) or []
    for r in rows:
        if isinstance(r.get("timestamp"), datetime):
            r["timestamp"] = r["timestamp"].strftime("%Y-%m-%d %H:%M:%S")
    return jsonify(rows)


@app.route("/api/historico_alarmas/excel", methods=["GET"])
def exportar_historico_alarmas_excel():
    """Genera y descarga un Excel con el histórico de alarmas filtrado por fecha."""
    inicio = request.args.get("inicio")
    fin    = request.args.get("fin")
    inst   = request.args.get("instrumento")
    nivel  = request.args.get("nivel")

    q = "SELECT * FROM historico_alarmas WHERE 1=1"
    params = []
    if inicio and fin:
        q += " AND timestamp BETWEEN %s AND %s"
        params += [inicio, fin]
    if inst:
        q += " AND instrumento = %s"
        params.append(inst)
    if nivel:
        q += " AND nivel = %s"
        params.append(nivel)
    q += " ORDER BY timestamp ASC LIMIT 15000"

    rows = db_exec(q, tuple(params)) or []

    # ── Crear Excel ──────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Alarmas"

    font_title  = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_hh     = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_ok     = Font(name="Segoe UI", size=10, color="1A5C2B")
    font_reg    = Font(name="Segoe UI", size=10)

    fill_title  = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_header = PatternFill(start_color="244062", end_color="244062", fill_type="solid")
    fill_hh     = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_h      = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    fill_l      = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    fill_ll     = PatternFill(start_color="7E22CE", end_color="7E22CE", fill_type="solid")
    fill_ok     = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fill_zebra  = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    align_c     = Alignment(horizontal="center", vertical="center")
    align_l     = Alignment(horizontal="left",   vertical="center")
    thin        = Side(border_style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = "HISTÓRICO DE ALARMAS — MFM ORINOCO"
    ws["A1"].font      = font_title
    ws["A1"].fill      = fill_title
    ws["A1"].alignment = align_c

    # Rango consultado
    ws.merge_cells("A2:G2")
    rango_txt = f"Período: {inicio or 'inicio'} → {fin or 'ahora'}"
    if inst:  rango_txt += f"  |  Instrumento: {inst}"
    if nivel: rango_txt += f"  |  Nivel: {nivel}"
    ws["A2"] = rango_txt
    ws["A2"].font      = Font(name="Segoe UI", size=9, italic=True, color="666666")
    ws["A2"].alignment = align_c

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 16

    # Encabezados
    headers = ["#", "Fecha / Hora", "Instrumento", "Descripción", "Unidad", "Valor", "Nivel Alarma", "SP Activo"]
    cols    = ["A", "B", "C", "D", "E", "F", "G", "H"]
    ws.merge_cells("A1:H1")   # re-merge con 8 cols
    for i, (h, c) in enumerate(zip(headers, cols), 1):
        cell = ws[f"{c}3"]
        cell.value     = h
        cell.font      = font_header
        cell.fill      = fill_header
        cell.alignment = align_c
        cell.border    = border

    ws.row_dimensions[3].height = 20
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12

    # Filas de datos
    nivel_fills = {"HH": fill_hh, "H": fill_h, "L": fill_l, "LL": fill_ll, "OK": fill_ok}
    nivel_fonts = {"HH": font_hh, "H": font_hh, "L": font_reg, "LL": font_hh, "OK": font_ok}

    for idx, r in enumerate(rows, 1):
        row_num = idx + 3
        ts = r.get("timestamp")
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S") if isinstance(ts, datetime) else str(ts)
        niv = r.get("nivel", "")
        row_fill  = nivel_fills.get(niv, fill_zebra if idx % 2 == 0 else None)
        row_font  = nivel_fonts.get(niv, font_reg)

        data = [idx, ts_str, r.get("instrumento",""), r.get("descripcion",""),
                r.get("unidad",""), r.get("valor",""), niv, r.get("sp_activo","")]

        for ci, (c, v) in enumerate(zip(cols, data)):
            cell = ws[f"{c}{row_num}"]
            cell.value     = v
            cell.font      = row_font
            cell.border    = border
            cell.alignment = align_c if ci not in (3,) else align_l
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_num].height = 16

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{3 + len(rows)}"

    # Total alarmas summary al pie
    footer_row = 4 + len(rows) + 1
    ws[f"A{footer_row}"] = f"Total registros: {len(rows)}"
    ws[f"A{footer_row}"].font = Font(name="Segoe UI", size=9, bold=True, color="444444")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Historico_Alarmas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ─────────────────────────────────────────────────────────────
# Reportes / Histórico
# ─────────────────────────────────────────────────────────────

@app.route("/api/reportes/descargar", methods=["GET"])
def descargar_reporte():
    prueba_id = request.args.get("prueba_id")
    f_inicio = request.args.get("inicio")
    f_fin    = request.args.get("fin")
    
    prueba_meta = None
    query = "SELECT * FROM lecturas_proceso"
    params = []
    
    if prueba_id:
        if prueba_id == "active":
            if bool(getattr(V, "b_Prueba_en_Progreso", False)) and _ACTIVE_PRUEBA_ID:
                rows_meta = db_exec("SELECT * FROM historico_pruebas WHERE id = %s", (_ACTIVE_PRUEBA_ID,))
                if rows_meta:
                    prueba_meta = rows_meta[0]
                    # Actualizar valores del PLC en tiempo real
                    prueba_meta["codigo_pozo"] = str(getattr(V, "as_Codigo_pozo_16", ""))
                    prueba_meta["lugar"] = str(getattr(V, "as_Codigo_pozo_17", ""))
                    prueba_meta["pozo"] = str(getattr(V, "as_Codigo_pozo_03", ""))
                    prueba_meta["metodo"] = str(getattr(V, "as_Codigo_pozo_06", ""))
                    prueba_meta["rpm"] = str(getattr(V, "as_Codigo_pozo_08", ""))
                    prueba_meta["inyeccion"] = str(getattr(V, "as_Codigo_pozo_18", ""))
                    prueba_meta["temp_yac"] = float(getattr(V, "r_T_Yac_C", 0.0))
                    prueba_meta["api_formacion"] = float(getattr(V, "r_API_formacion_BM", 0.0))
                    prueba_meta["api_mezcla"] = float(getattr(V, "r_API_2", 0.0))
                    prueba_meta["api_diluente"] = float(getattr(V, "r_API_1", 0.0))
                    prueba_meta["caudal_diluente"] = float(getattr(V, "r_caudal_dil_BM", 0.0))
                    prueba_meta["duracion_horas"] = float(getattr(V, "i_duracion_prueba_horas", 0.0))
                
                query += " WHERE prueba_id = %s"
                params.append(_ACTIVE_PRUEBA_ID)
            else:
                rows_meta = db_exec("SELECT * FROM historico_pruebas ORDER BY id DESC LIMIT 1")
                if rows_meta:
                    prueba_meta = rows_meta[0]
                    query += " WHERE prueba_id = %s"
                    params.append(prueba_meta["id"])
                else:
                    return jsonify({"error": "No hay pruebas registradas ni activas"}), 404
        else:
            try:
                p_id = int(prueba_id)
                rows_meta = db_exec("SELECT * FROM historico_pruebas WHERE id = %s", (p_id,))
                if rows_meta:
                    prueba_meta = rows_meta[0]
                    query += " WHERE prueba_id = %s"
                    params.append(p_id)
                    
                    # Si no hay lecturas directas por prueba_id (para retrocompatibilidad)
                    # podemos intentar por rango de fechas
                    rows_test = db_exec(query, tuple(params))
                    if not rows_test and prueba_meta.get("fecha_inicio"):
                        query = "SELECT * FROM lecturas_proceso WHERE timestamp BETWEEN %s AND %s"
                        params = [prueba_meta["fecha_inicio"], prueba_meta["fecha_fin"] or datetime.now()]
                else:
                    return jsonify({"error": f"No se encontró la prueba con ID {p_id}"}), 404
            except ValueError:
                return jsonify({"error": "ID de prueba inválido"}), 400
    else:
        if f_inicio and f_fin:
            query += " WHERE timestamp BETWEEN %s AND %s"
            params.extend([f_inicio, f_fin])
            
    query += " ORDER BY timestamp ASC LIMIT 15000"
    rows = db_exec(query, tuple(params))
    
    # Si no se filtró por prueba id específica pero una prueba está en progreso
    if not prueba_meta and bool(getattr(V, "b_Prueba_en_Progreso", False)):
        prueba_meta = {
            "id": _ACTIVE_PRUEBA_ID or "En progreso",
            "codigo_pozo": str(getattr(V, "as_Codigo_pozo_16", "")),
            "lugar": str(getattr(V, "as_Codigo_pozo_17", "")),
            "pozo": str(getattr(V, "as_Codigo_pozo_03", "")),
            "metodo": str(getattr(V, "as_Codigo_pozo_06", "")),
            "rpm": str(getattr(V, "as_Codigo_pozo_08", "")),
            "inyeccion": str(getattr(V, "as_Codigo_pozo_18", "")),
            "temp_yac": float(getattr(V, "r_T_Yac_C", 0.0)),
            "api_formacion": float(getattr(V, "r_API_formacion_BM", 0.0)),
            "api_mezcla": float(getattr(V, "r_API_2", 0.0)),
            "api_diluente": float(getattr(V, "r_API_1", 0.0)),
            "caudal_diluente": float(getattr(V, "r_caudal_dil_BM", 0.0)),
            "duracion_horas": float(getattr(V, "i_duracion_prueba_horas", 0.0)),
            "fecha_inicio": "En progreso",
            "fecha_fin": "",
            "estado": "En progreso"
        }
        
    pivoted = {}
    for r in (rows or []):
        ts = r["timestamp"]
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S") if isinstance(ts, datetime) else str(ts)
        if ts_str not in pivoted:
            pivoted[ts_str] = {}
        pivoted[ts_str][r["instrumento"]] = r["valor"]
        
    sorted_ts = sorted(pivoted.keys())
    instrumentos_orden = [
        "FI-03", "PI-01", "PI-02", "TI-01", "TI-02", "LI-01", "NIV-AUX",
        "PDI-01", "PDI-03", "PDI-02", "PDI-04", "GAS-01", "VI-01", "WC",
        "Coriolis_Density", "Coriolis_Temp", "Coriolis_Vol_flow_Rate",
        "Q_LIQ", "Q_CRUDO", "Q_NETO", "Q_DIL", "Q_AGUA", "Q_GAS"
    ]
    
    descripciones = {
        "FI-03": "Flujo Gas Vortex",
        "PI-01": "Presión de Entrada",
        "PI-02": "Presión Líquido / Wedge",
        "TI-01": "Temperatura de Entrada",
        "TI-02": "Temperatura Proceso",
        "LI-01": "Nivel del Separador",
        "NIV-AUX": "Nivel Auxiliar",
        "PDI-01": "Diferencial de Presión Lam. A",
        "PDI-03": "Diferencial de Presión Lam. B",
        "PDI-02": "Diferencial de Presión Wedge",
        "PDI-04": "Diferencial de Presión Wedge Gas",
        "GAS-01": "Porcentaje de Gas",
        "VI-01": "Viscosidad del Crudo",
        "WC": "Corte de Agua",
        "Coriolis_Density": "Densidad Coriolis",
        "Coriolis_Temp": "Temperatura Coriolis",
        "Coriolis_Vol_flow_Rate": "Caudal Coriolis",
        "Q_LIQ": "Caudal Líquido Medido",
        "Q_CRUDO": "Caudal Crudo Medido",
        "Q_NETO": "Caudal Neto Medido",
        "Q_DIL": "Caudal Diluente Medido",
        "Q_AGUA": "Caudal Agua Medido",
        "Q_GAS": "Caudal Gas Medido"
    }
    unidades = {
        "FI-03": "MSCFD",
        "PI-01": "PSIG",
        "PI-02": "psia",
        "TI-01": "°C",
        "TI-02": "°C",
        "LI-01": "%",
        "NIV-AUX": "%",
        "PDI-01": "inH2O",
        "PDI-03": "inH2O",
        "PDI-02": "inH2O",
        "PDI-04": "inH2O",
        "GAS-01": "%",
        "VI-01": "CP",
        "WC": "%",
        "Coriolis_Density": "gr/cm³",
        "Coriolis_Temp": "°F",
        "Coriolis_Vol_flow_Rate": "BB/D",
        "Q_LIQ": "BBLD",
        "Q_CRUDO": "BBLD",
        "Q_NETO": "BBLD",
        "Q_DIL": "BBLD",
        "Q_AGUA": "BBLD",
        "Q_GAS": "MCFD"
    }
    try:
        db_descs = db_exec("SELECT instrumento, descripcion, unidad FROM tabla_configuracion_alarma")
        for d in (db_descs or []):
            inst = d["instrumento"]
            if inst in descripciones:
                if d.get("descripcion"):
                    descripciones[inst] = d["descripcion"]
                if d.get("unidad"):
                    unidades[inst] = d["unidad"]
    except Exception:
        pass
        
    # ── OBTENER VALORES PARA CÁLCULO DE CONDICIONES Y VALORES ACTUALES ──
    vals_calc = {
        "r_WC": float(getattr(V, "r_WC", 0.0)),
        "r_GVoidF": float(getattr(V, "r_GVoidF", 0.0)),
        "r_T_Gas": float(getattr(V, "r_T_Gas", 0.0)),
        "r_T_Oil_C": float(getattr(V, "r_T_Oil_C", 0.0)),
        "r_P_Gas": float(getattr(V, "r_P_Gas", 0.0)),
        "r_v_oil_medida": float(getattr(V, "r_v_oil_medida", 0.0)),
        "r_Q_gas_STD": float(getattr(V, "r_Q_gas_STD", 0.0)),
        "r_LIT_001": float(getattr(V, "r_LIT_001", 0.0)),
    }
    
    # Si es una prueba finalizada, extraer la última lectura registrada
    if prueba_meta and prueba_meta.get("estado") != "En progreso":
        try:
            last_rows = db_exec("""
                SELECT instrumento, valor FROM lecturas_proceso 
                WHERE prueba_id = %s 
                AND timestamp = (SELECT MAX(timestamp) FROM lecturas_proceso WHERE prueba_id = %s)
            """, (prueba_meta["id"], prueba_meta["id"]))
            
            tag_to_var = {
                "WC": "r_WC",
                "GAS-01": "r_GVoidF",
                "TI-02": "r_T_Gas",
                "TI-01": "r_T_Oil_C",
                "PI-01": "r_P_Gas",
                "VI-01": "r_v_oil_medida",
                "FI-03": "r_Q_gas_STD",
                "LI-01": "r_LIT_001"
            }
            for lr in (last_rows or []):
                var_name = tag_to_var.get(lr["instrumento"])
                if var_name:
                    vals_calc[var_name] = float(lr["valor"])
        except Exception as ex_db_calc:
            logger.error(f"Error recuperando lecturas finales: {ex_db_calc}")

    # Estructurar arrays de datos para la segunda hoja
    val_actuales = [
        ("Corte de Agua (%)", vals_calc["r_WC"], "0.000"),
        ("GVF (%)", vals_calc["r_GVoidF"], "0.000"),
        ("Temp. Gas (ºC)", vals_calc["r_T_Gas"], "0.000"),
        ("Temp. Mezcla (ºC)", vals_calc["r_T_Oil_C"], "0.000"),
        ("Presión en Línea (PSI)", vals_calc["r_P_Gas"], "0.0"),
        ("Viscosidad (cP)", vals_calc["r_v_oil_medida"], "0.0"),
        ("RGP", vals_calc["r_Q_gas_STD"] * 12.5, "0.000"),
        ("RGP NETO", vals_calc["r_Q_gas_STD"] * 11.2, "0.000"),
    ]
    
    v_liq = vals_calc["r_LIT_001"]
    v_crudo = vals_calc["r_P_Gas"]
    v_gas = vals_calc["r_Q_gas_STD"]
    v_gvf = vals_calc["r_GVoidF"]
    
    cond_linea = [
        ("Vol. Líquido (BBLS)", v_liq * 0.1, "0.000"),
        ("Vol. Crudo (BBLS)", v_crudo * 0.05, "0.000"),
        ("Vol. Crudo Neto (BBLS)", v_crudo * 0.045, "0.000"),
        ("Vol. Diluente (BBLS)", 0.0, "0.000"),
        ("Vol. Agua (BBLS)", v_liq * 0.02, "0.000"),
        ("Vol. Gas Arrastrado (CF)", v_gvf * 1.2, "0.000"),
        ("Vol. Gas Total (MCF)", v_gas * 0.8, "0.000"),
        ("Tasa Est. Líquido (BPD)", v_liq * 2.4, "0.000"),
        ("Tasa Est. Crudo (BPD)", v_crudo * 1.2, "0.000"),
        ("Tasa Est. Crudo Neto (BPD)", v_crudo * 1.08, "0.000"),
        ("Tasa Est. Diluente (BPD)", 0.0, "0.000"),
        ("Tasa Est. Agua (BPD)", v_liq * 0.48, "0.000"),
        ("Tasa Est. Gas Arrastrado (CFD)", v_gvf * 28.8, "0.000"),
        ("Tasa Est. Gas Total (MCFD)", v_gas * 19.2, "0.000"),
    ]
    
    cond_estandar = [
        ("Vol. Líquido (BBLS)", v_liq * 0.098, "0.000"),
        ("Vol. Crudo (BBLS)", v_crudo * 0.049, "0.000"),
        ("Vol. Crudo Neto (BBLS)", v_crudo * 0.044, "0.000"),
        ("Vol. Diluente (BBLS)", 0.0, "0.000"),
        ("Vol. Agua (BBLS)", v_liq * 0.019, "0.000"),
        ("Vol. Gas Arrastrado (CF)", v_gvf * 1.15, "0.000"),
        ("Vol. Gas Total (MCF)", v_gas * 0.76, "0.000"),
        ("Tasa Est. Líquido (BPD)", v_liq * 2.35, "0.000"),
        ("Tasa Est. Crudo (BPD)", v_crudo * 1.17, "0.000"),
        ("Tasa Est. Crudo Neto (BPD)", v_crudo * 1.05, "0.000"),
        ("Tasa Est. Diluente (BPD)", 0.0, "0.000"),
        ("Tasa Est. Agua (BPD)", v_liq * 0.46, "0.000"),
        ("Tasa Est. Gas Arrastrado (CFD)", v_gvf * 27.6, "0.000"),
        ("Tasa Est. Gas Total (MCFD)", v_gas * 18.2, "0.000"),
    ]
        
    # Crear Libro de Excel nativo (.xlsx) con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas Históricas"
    
    # Asegurar visibilidad de líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True
    
    # Definición de Estilos Premium
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1F497D")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    
    fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_header = PatternFill(start_color="244062", end_color="244062", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_meta_lbl = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_thick_bottom = Side(border_style="medium", color="1F497D")
    
    grid_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_line = Border(bottom=border_thick_bottom)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Índice de filas actuales
    r_idx = 1
    
    # Fila 1: Título unificado
    ws.merge_cells("A1:R1")
    ws["A1"] = "REPORTE DE PROCESO Y PRUEBAS - MFM ORINOCO"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40
    r_idx += 2
    
    # Fila 3: Información General
    ws.cell(row=r_idx, column=1, value="Fecha de exportación:").font = font_bold
    ws.cell(row=r_idx, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = font_regular
    r_idx += 2
    
    # Datos de la Prueba si corresponde
    if prueba_meta:
        ws.cell(row=r_idx, column=1, value="DATOS DE LA PRUEBA DE POZO").font = font_section
        ws.row_dimensions[r_idx].height = 24
        for col_idx in range(1, 19):
            ws.cell(row=r_idx, column=col_idx).border = bottom_line
        r_idx += 1
        
        meta_fields = [
            ("ID de Prueba", prueba_meta.get("id", "-")),
            ("Código de Pozo", prueba_meta.get("codigo_pozo", "-")),
            ("Lugar", prueba_meta.get("lugar", "-")),
            ("Pozo", prueba_meta.get("pozo", "-")),
            ("Método de Producción", prueba_meta.get("metodo", "-")),
            ("RPM", prueba_meta.get("rpm", "-")),
            ("Inyección Diluente", prueba_meta.get("inyeccion", "-")),
            ("Temperatura Yacimiento (°C)", prueba_meta.get("temp_yac", 0.0)),
            ("API Formación", prueba_meta.get("api_formacion", 0.0)),
            ("API Mezcla", prueba_meta.get("api_mezcla", 0.0)),
            ("API Diluente", prueba_meta.get("api_diluente", 0.0)),
            ("Caudal Diluente (BPD)", prueba_meta.get("caudal_diluente", 0.0)),
            ("Duración de Prueba (Horas)", prueba_meta.get("duracion_horas", 0.0)),
            ("Fecha Inicio", prueba_meta.get("fecha_inicio", "-")),
            ("Fecha Fin", prueba_meta.get("fecha_fin") or "En progreso"),
            ("Estado de la Prueba", prueba_meta.get("estado", "-"))
        ]
        
        for i, (lbl, val) in enumerate(meta_fields):
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d %H:%M:%S")
            c_offset = 1 if i % 2 == 0 else 4
            curr_row = r_idx + (i // 2)
            
            c_lbl = ws.cell(row=curr_row, column=c_offset, value=lbl)
            c_lbl.font = font_bold
            c_lbl.fill = fill_meta_lbl
            c_lbl.border = grid_border
            
            c_val = ws.cell(row=curr_row, column=c_offset + 1, value=val)
            c_val.font = font_regular
            c_val.border = grid_border
            
            if isinstance(val, float):
                c_val.number_format = "0.0"
            elif isinstance(val, int):
                c_val.number_format = "0"
                
        r_idx += (len(meta_fields) + 1) // 2
        r_idx += 2
    elif f_inicio and f_fin:
        ws.cell(row=r_idx, column=1, value="RANGO DE FECHAS SELECCIONADO").font = font_section
        ws.row_dimensions[r_idx].height = 24
        for col_idx in range(1, 19):
            ws.cell(row=r_idx, column=col_idx).border = bottom_line
        r_idx += 1
        
        ws.cell(row=r_idx, column=1, value="Fecha de Inicio:").font = font_bold
        ws.cell(row=r_idx, column=2, value=f_inicio).font = font_regular
        r_idx += 1
        ws.cell(row=r_idx, column=1, value="Fecha de Fin:").font = font_bold
        ws.cell(row=r_idx, column=2, value=f_fin).font = font_regular
        r_idx += 2

    # Tabla de Tags/Instrumentos
    ws.cell(row=r_idx, column=1, value="INFORMACIÓN DE INSTRUMENTOS (TAGS)").font = font_section
    ws.row_dimensions[r_idx].height = 24
    for col_idx in range(1, 19):
        ws.cell(row=r_idx, column=col_idx).border = bottom_line
    r_idx += 1
    
    tags_headers = ["Tag", "Nombre del Instrumento", "Unidad"]
    for col_c, h in enumerate(tags_headers, 1):
        cell = ws.cell(row=r_idx, column=col_c, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = grid_border
    ws.row_dimensions[r_idx].height = 22
    r_idx += 1
    
    for inst in instrumentos_orden:
        c1 = ws.cell(row=r_idx, column=1, value=inst)
        c1.font = font_bold
        c1.border = grid_border
        c1.alignment = align_center
        
        c2 = ws.cell(row=r_idx, column=2, value=descripciones[inst])
        c2.font = font_regular
        c2.border = grid_border
        
        c3 = ws.cell(row=r_idx, column=3, value=unidades[inst])
        c3.font = font_regular
        c3.border = grid_border
        c3.alignment = align_center
        r_idx += 1
    r_idx += 2
    
    # Histórico de Lecturas
    ws.cell(row=r_idx, column=1, value="HISTÓRICO DE LECTURAS (DATOS DE PROCESO)").font = font_section
    ws.row_dimensions[r_idx].height = 24
    for col_idx in range(1, 19):
        ws.cell(row=r_idx, column=col_idx).border = bottom_line
    r_idx += 1
    
    headers_row = ["Timestamp"] + [f"{inst} ({unidades[inst]})" for inst in instrumentos_orden]
    for col_c, h in enumerate(headers_row, 1):
        cell = ws.cell(row=r_idx, column=col_c, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = grid_border
    ws.row_dimensions[r_idx].height = 22
    r_idx += 1
    
    if sorted_ts:
        for idx_ts, ts in enumerate(sorted_ts):
            ws.row_dimensions[r_idx].height = 20
            c_ts = ws.cell(row=r_idx, column=1, value=ts)
            c_ts.font = font_regular
            c_ts.border = grid_border
            c_ts.alignment = align_center
            
            row_fill = fill_zebra if idx_ts % 2 != 0 else None
            if row_fill:
                c_ts.fill = row_fill
                
            for col_i, inst in enumerate(instrumentos_orden, 2):
                val = pivoted[ts].get(inst, "")
                cell_v = ws.cell(row=r_idx, column=col_i)
                cell_v.border = grid_border
                cell_v.alignment = align_right
                if row_fill:
                    cell_v.fill = row_fill
                    
                if isinstance(val, (int, float)):
                    cell_v.value = round(val, 3)
                    cell_v.number_format = "0.000"
                else:
                    cell_v.value = val
            r_idx += 1
    else:
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(headers_row))
        cell = ws.cell(row=r_idx, column=1, value="Sin datos para el rango o prueba seleccionada")
        cell.font = font_regular
        cell.alignment = align_center
        for col_idx in range(1, len(headers_row) + 1):
            ws.cell(row=r_idx, column=col_idx).border = grid_border
        r_idx += 1
        
    # Auto-ajustar el ancho de las columnas (Hoja 1)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # ── SEGUNDA HOJA: CONDICIONES Y VALORES ACTUALES ──
    ws2 = wb.create_sheet(title="Condiciones y Valores")
    ws2.views.sheetView[0].showGridLines = True
    
    # Fila 1: Título unificado
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "VALORES ACTUALES Y CONDICIONES DE LA PRUEBA"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_title
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 40
    
    # Headers
    # Col A-B: Valores Actuales
    ws2.merge_cells("A3:B3")
    ws2["A3"] = "VALORES ACTUALES"
    ws2["A3"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws2["A3"].fill = fill_header
    ws2["A3"].alignment = align_center
    ws2.row_dimensions[3].height = 24
    
    ws2.cell(row=4, column=1, value="Parámetro").font = font_bold
    ws2.cell(row=4, column=1).border = grid_border
    ws2.cell(row=4, column=1).fill = fill_meta_lbl
    ws2.cell(row=4, column=2, value="Valor").font = font_bold
    ws2.cell(row=4, column=2).border = grid_border
    ws2.cell(row=4, column=2).fill = fill_meta_lbl
    ws2.cell(row=4, column=2).alignment = align_right
    ws2.row_dimensions[4].height = 20
    
    for row_idx, (lbl, val, fmt) in enumerate(val_actuales, 5):
        cell_lbl = ws2.cell(row=row_idx, column=1, value=lbl)
        cell_lbl.font = font_regular
        cell_lbl.border = grid_border
        
        cell_val = ws2.cell(row=row_idx, column=2, value=val)
        cell_val.font = font_bold
        cell_val.border = grid_border
        cell_val.alignment = align_right
        cell_val.number_format = fmt
        
        if (row_idx - 5) % 2 != 0:
            cell_lbl.fill = fill_zebra
            cell_val.fill = fill_zebra
            
    # Col D-E: Condiciones de Línea
    ws2.merge_cells("D3:E3")
    ws2["D3"] = "CONDICIONES DE LÍNEA"
    ws2["D3"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws2["D3"].fill = fill_header
    ws2["D3"].alignment = align_center
    
    ws2.cell(row=4, column=4, value="Parámetro").font = font_bold
    ws2.cell(row=4, column=4).border = grid_border
    ws2.cell(row=4, column=4).fill = fill_meta_lbl
    ws2.cell(row=4, column=5, value="Valor").font = font_bold
    ws2.cell(row=4, column=5).border = grid_border
    ws2.cell(row=4, column=5).fill = fill_meta_lbl
    ws2.cell(row=4, column=5).alignment = align_right
    
    for row_idx, (lbl, val, fmt) in enumerate(cond_linea, 5):
        cell_lbl = ws2.cell(row=row_idx, column=4, value=lbl)
        cell_lbl.font = font_regular
        cell_lbl.border = grid_border
        
        cell_val = ws2.cell(row=row_idx, column=5, value=val)
        cell_val.font = font_bold
        cell_val.border = grid_border
        cell_val.alignment = align_right
        cell_val.number_format = fmt
        
        if (row_idx - 5) % 2 != 0:
            cell_lbl.fill = fill_zebra
            cell_val.fill = fill_zebra
            
    # Col G-H: Condiciones Estándar
    ws2.merge_cells("G3:H3")
    ws2["G3"] = "CONDICIONES ESTÁNDAR"
    ws2["G3"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header_green = PatternFill(start_color="1E4D2B", end_color="1E4D2B", fill_type="solid")
    ws2["G3"].fill = fill_header_green
    ws2["G3"].alignment = align_center
    
    ws2.cell(row=4, column=7, value="Parámetro").font = font_bold
    ws2.cell(row=4, column=7).border = grid_border
    ws2.cell(row=4, column=7).fill = fill_meta_lbl
    ws2.cell(row=4, column=8, value="Valor").font = font_bold
    ws2.cell(row=4, column=8).border = grid_border
    ws2.cell(row=4, column=8).fill = fill_meta_lbl
    ws2.cell(row=4, column=8).alignment = align_right
    
    for row_idx, (lbl, val, fmt) in enumerate(cond_estandar, 5):
        cell_lbl = ws2.cell(row=row_idx, column=7, value=lbl)
        cell_lbl.font = font_regular
        cell_lbl.border = grid_border
        
        cell_val = ws2.cell(row=row_idx, column=8, value=val)
        cell_val.font = font_bold
        cell_val.border = grid_border
        cell_val.alignment = align_right
        cell_val.number_format = fmt
        
        if (row_idx - 5) % 2 != 0:
            cell_lbl.fill = fill_zebra
            cell_val.fill = fill_zebra
            
    # Altura de filas en ws2
    for r_num in range(5, 20):
        ws2.row_dimensions[r_num].height = 20
        
    # Espaciadores de columnas en ws2
    ws2.column_dimensions['C'].width = 4
    ws2.column_dimensions['F'].width = 4
    
    # Auto-ajustar anchos ws2
    for col in ws2.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ['C', 'F']:
            continue
        max_len = 0
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Guardar en memoria y retornar como send_file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "Reporte_MFM_Orinoco.xlsx"
    if prueba_meta:
        cod_pozo_clean = str(prueba_meta.get("codigo_pozo", "")).replace(" ", "_").replace("/", "-")
        filename = f"Reporte_Prueba_{prueba_meta.get('id')}_{cod_pozo_clean}.xlsx"
        
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/api/reportes/pruebas", methods=["GET"])
def listar_pruebas():
    rows = db_exec("SELECT * FROM historico_pruebas ORDER BY id DESC")
    for r in (rows or []):
        if r.get("fecha_inicio"):
            r["fecha_inicio"] = r["fecha_inicio"].strftime("%Y-%m-%d %H:%M:%S")
        if r.get("fecha_fin"):
            r["fecha_fin"] = r["fecha_fin"].strftime("%Y-%m-%d %H:%M:%S")
    return jsonify(rows or [])


@app.route("/api/valores_agregados", methods=["GET"])
def get_valores_agregados():
    instrumento = request.args.get("instrumento")
    limit       = min(int(request.args.get("limit", 200)), 2000)
    query  = "SELECT * FROM valores_agregados"
    params = []
    if instrumento:
        query  += " WHERE instrumento=%s"
        params  = [instrumento.upper()]
    query += " ORDER BY timestamp DESC LIMIT %s"
    params.append(limit)
    rows = db_exec(query, tuple(params))
    return jsonify(rows or [])


@app.route("/api/modbus/status", methods=["GET"])
def get_modbus_status():
    row = db_exec(
        "SELECT instrumento, valor_promedio, fuente, timestamp "
        "FROM valores_agregados ORDER BY timestamp DESC LIMIT 1"
    )
    if not row:
        return jsonify({"activo": False, "mensaje": "Sin datos del DAQ aún"})
    r   = row[0]
    ts  = r["timestamp"]
    diff = (datetime.now() - ts).total_seconds() if ts else 9999
    return jsonify({
        "activo":       diff < 30,
        "ultimo_dato":  str(ts),
        "fuente":       r["fuente"],
        "antiguedad_s": round(diff, 1),
    })



# ─────────────────────────────────────────────────────────────
# API DAQ — Configuración y Estado en Tiempo Real
# ─────────────────────────────────────────────────────────────

@app.route("/api/daq/live", methods=["GET"])
def daq_live():
    """
    Retorna el snapshot en vivo de los 6 canales AI de la DAQ.
    Detecta 'stale data': si no hubo read exitoso en los últimos 5 s
    se reporta connected=False aunque el cliente Modbus no haya fallado.
    """
    import time as _time
    import modbus_daq as _mdaq
    import fase2_entradas as _f2

    # IMPORTANTE: leer el snapshot desde el módulo del PLC directamente.
    # 'V' en app.py es 'python_migration.global_vars.V' mientras que
    # el scan engine usa 'global_vars.V' (distinto por sys.path dual).
    # Los atributos _daq_channel_snapshot y _daq_last_success viven en
    # el V del scan engine, accesible vía _f2.V
    _V_plc      = _f2.V
    snapshot     = getattr(_V_plc, "_daq_channel_snapshot", None)
    last_success = getattr(_V_plc, "_daq_last_success", None)
    STALE_LIMIT  = 5.0

    if last_success is None:
        data_age_s = 0.0
        stale      = False
    else:
        data_age_s = _time.monotonic() - last_success
        stale      = data_age_s > STALE_LIMIT

    connected = (not _V_plc.b_Error_DAQ) and (not stale)

    # Si no hay snapshot todavía, construirlo desde el mapa de canales
    if not snapshot:
        snapshot = [
            {"ch": addr, "var": var_name, "desc": desc,
             "raw": None, "ma": None, "open_wire": True}
            for (var_name, addr, _escala, desc) in _f2._INPUT_MAP
        ]

    # Si los datos son stale, marcar todos como open-wire
    if stale:
        snapshot = [dict(ch=c["ch"], var=c.get("var",""), desc=c.get("desc",""),
                         raw=None, ma=None, open_wire=True) for c in snapshot]
        _V_plc.b_Error_DAQ = True

    cooldown_left = max(0.0, _mdaq.RECONNECT_COOLDOWN -
                        (_time.monotonic() - _mdaq._last_attempt))

    last_error = _mdaq._last_error
    if stale and not last_error:
        last_error = f"Sin datos hace {round(data_age_s, 1)} s"

    import fase8_salidas as _f8
    ao_channels = []
    for (var_name, modbus_addr, scale_min, scale_max, desc) in getattr(_f8, '_OUTPUT_MAP', []):
        val_raw = float(getattr(_V_plc, var_name, scale_min) or scale_min)
        val_eu = 0.0
        if scale_max > scale_min:
            val_eu = (val_raw - scale_min) / (scale_max - scale_min) * 100.0
            
        ao_channels.append({
            "ch": modbus_addr,
            "var": var_name,
            "desc": desc,
            "val_eu": round(val_eu, 2),
            "val_raw": round(val_raw, 1),
            "addr": modbus_addr
        })

    return jsonify({
        "connected":   connected,
        "stale":       stale,
        "data_age_s":  round(data_age_s, 1),
        "port":        _mdaq.DAQ_PORT,
        "baudrate":    _mdaq.DAQ_BAUDRATE,
        "slave_id":    _mdaq.DAQ_SLAVE_ID,
        "simulating":  V.b_simular_ai,
        "channels":    snapshot,
        "ao_channels": ao_channels,
        "last_error":  last_error,
        "retry_in_s":  round(cooldown_left, 1),
        "ts":          datetime.now().strftime("%H:%M:%S"),
    })


@app.route("/api/instrument_selection", methods=["POST"])
def save_instrument_selection():
    """Guarda la configuración de selección de instrumentos en la BD y recarga a memoria."""
    d = request.get_json() or {}
    try:
        db_exec(
            """UPDATE instrument_selection_config SET
               b_Control_PID_Gas = %s, b_PID_POSIC_SW = %s,
               b_Sw_Wedge_Gas = %s, b_SW_DIL_MEDIDO_CALC = %s,
               b_Sw_Wedge_Gas_2 = %s, b_SEL_LAMINAR = %s, b_SEL_T_baja = %s,
               b_sw_AM_Laminar_Wedge_x = %s, b_sw_AM_Laminar_Wedge_y = %s,
               b_sel_tipo_instrum_dil = %s, b_AUTO_GAS_01 = %s,
               b_SEL_VLV_GAS_01 = %s, b_DESHABILITA_PID = %s
               WHERE id = 1""",
            (
                bool(d.get("b_Control_PID_Gas", False)),
                bool(d.get("b_PID_POSIC_SW", False)),
                bool(d.get("b_Sw_Wedge_Gas", False)),
                bool(d.get("b_SW_DIL_MEDIDO_CALC", False)),
                bool(d.get("b_Sw_Wedge_Gas_2", False)),
                bool(d.get("b_SEL_LAMINAR", False)),
                bool(d.get("b_SEL_T_baja", False)),
                bool(d.get("b_sw_AM_Laminar_Wedge_x", True)),
                bool(d.get("b_sw_AM_Laminar_Wedge_y", False)),
                bool(d.get("b_sel_tipo_instrum_dil", False)),
                bool(d.get("b_AUTO_GAS_01", False)),
                bool(d.get("b_SEL_VLV_GAS_01", False)),
                bool(d.get("b_DESHABILITA_PID", False)),
            )
        )
        _load_instrument_selection_from_db()  # Recargar a V global
        return jsonify({"ok": True})
    except Exception as e:
        print("Error saving instrument selection:", e)
        return jsonify({"error": str(e)}), 500


@app.route("/api/daq/config", methods=["GET"])
def daq_get_config():
    """Lee la configuración de canales desde la BD (tabla daq_channel_config)."""
    rows = db_exec("SELECT * FROM daq_channel_config ORDER BY channel_addr")
    return jsonify(rows or [])


@app.route("/api/daq/config", methods=["POST"])
def daq_save_config():
    """
    Guarda (UPSERT) la configuración de un canal en la BD.
    Payload: { channel_addr, v_name, description, scale, eu_min, eu_max, enabled, modbus_addr }
    """
    d = request.get_json() or {}
    required = ["channel_addr", "v_name", "description"]
    if not all(k in d for k in required):
        return jsonify({"error": "Faltan campos requeridos"}), 400

    db_exec(
        """INSERT INTO daq_channel_config
             (channel_addr, v_name, description, scale, eu_min, eu_max, enabled, modbus_addr)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
           ON DUPLICATE KEY UPDATE
             v_name=%s, description=%s, scale=%s,
             eu_min=%s, eu_max=%s, enabled=%s, modbus_addr=%s, updated_at=NOW()""",
        (
            int(d["channel_addr"]),
            d["v_name"], d["description"],
            float(d.get("scale", 1000.0)),
            float(d.get("eu_min", 4.0)),
            float(d.get("eu_max", 20.0)),
            bool(d.get("enabled", True)),
            int(d.get("modbus_addr", d["channel_addr"])),
            # ON DUPLICATE KEY values:
            d["v_name"], d["description"],
            float(d.get("scale", 1000.0)),
            float(d.get("eu_min", 4.0)),
            float(d.get("eu_max", 20.0)),
            bool(d.get("enabled", True)),
            int(d.get("modbus_addr", d["channel_addr"])),
        ),
        fetch=False,
    )
    
    # Recargar la configuración en caliente
    _load_daq_channels_from_db()
    
    return jsonify({"ok": True, "channel_addr": d["channel_addr"]})

@app.route("/api/daq/ao_config", methods=["GET"])
def daq_get_ao_config():
    """Lee la configuración de canales AO desde la BD (tabla daq_ao_config)."""
    rows = db_exec("SELECT * FROM daq_ao_config ORDER BY channel_addr")
    return jsonify(rows or [])

@app.route("/api/daq/ao_config", methods=["POST"])
def daq_save_ao_config():
    """
    Guarda (UPSERT) la configuración de un canal AO en la BD.
    Payload: { channel_addr, v_name, description, scale_min, scale_max, enabled, modbus_addr }
    """
    d = request.get_json() or {}
    required = ["channel_addr", "v_name", "description"]
    if not all(k in d for k in required):
        return jsonify({"error": "Faltan campos requeridos"}), 400

    db_exec(
        """INSERT INTO daq_ao_config
             (channel_addr, v_name, description, scale_min, scale_max, enabled, modbus_addr)
           VALUES (%s,%s,%s,%s,%s,%s,%s)
           ON DUPLICATE KEY UPDATE
             v_name=%s, description=%s, scale_min=%s,
             scale_max=%s, enabled=%s, modbus_addr=%s, updated_at=NOW()""",
        (
            int(d["channel_addr"]),
            d["v_name"], d["description"],
            float(d.get("scale_min", 4000.0)),
            float(d.get("scale_max", 20000.0)),
            bool(d.get("enabled", True)),
            int(d.get("modbus_addr", d["channel_addr"])),
            # ON DUPLICATE KEY values:
            d["v_name"], d["description"],
            float(d.get("scale_min", 4000.0)),
            float(d.get("scale_max", 20000.0)),
            bool(d.get("enabled", True)),
            int(d.get("modbus_addr", d["channel_addr"])),
        ),
        fetch=False,
    )
    
    # Recargar la configuración en caliente
    _load_daq_ao_from_db()
    
    return jsonify({"ok": True, "channel_addr": d["channel_addr"]})


@app.route("/api/daq/connection", methods=["GET"])
def daq_get_connection():
    """Devuelve la configuración de conexión guardada en BD."""
    import python_migration.modbus_daq as _mdaq
    rows = db_exec("SELECT * FROM daq_connection_config WHERE id=1")
    if rows:
        return jsonify(rows[0])
    # Fallback: valores actuales del módulo
    return jsonify({
        "id": 1, "port": _mdaq.DAQ_PORT, "baudrate": _mdaq.DAQ_BAUDRATE,
        "slave_id": _mdaq.DAQ_SLAVE_ID, "bytesize": 8, "parity": "N",
        "stopbits": 1, "timeout_ms": int(_mdaq.DAQ_TIMEOUT * 1000),
    })


@app.route("/api/daq/connection", methods=["POST"])
def daq_save_connection():
    """
    1. Guarda la configuración en BD (UPSERT fila id=1)
    2. Aplica los parámetros al módulo modbus_daq en tiempo real
    3. Fuerza reconexion limpia con cooldown = 0
    """
    import sys
    import python_migration.modbus_daq as _mdaq1
    d = request.get_json() or {}

    port     = str(d.get("port",     _mdaq1.DAQ_PORT)).upper()
    baudrate = int(d.get("baudrate", _mdaq1.DAQ_BAUDRATE))
    slave_id = int(d.get("slave_id", _mdaq1.DAQ_SLAVE_ID))
    timeout_ms = int(d.get("timeout_ms", int(_mdaq1.DAQ_TIMEOUT * 1000)))

    # 1. Persistir en BD (UPSERT sobre la fila única id=1)
    db_exec(
        """INSERT INTO daq_connection_config
               (id, port, baudrate, slave_id, timeout_ms)
           VALUES (1, %s, %s, %s, %s)
           ON DUPLICATE KEY UPDATE
               port=%s, baudrate=%s, slave_id=%s, timeout_ms=%s, updated_at=NOW()""",
        (port, baudrate, slave_id, timeout_ms,
         port, baudrate, slave_id, timeout_ms),
        fetch=False,
    )

    # 2. Aplicar a todas las instancias del módulo en tiempo real
    mdaq_instances = [_mdaq1]
    _mdaq2 = sys.modules.get('modbus_daq')
    if _mdaq2:
        mdaq_instances.append(_mdaq2)

    for mdaq in mdaq_instances:
        mdaq.DAQ_PORT     = port
        mdaq.DAQ_BAUDRATE = baudrate
        mdaq.DAQ_SLAVE_ID = slave_id
        mdaq.DAQ_TIMEOUT  = timeout_ms / 1000.0
        # 3. Forzar reconexión limpia inmediata
        mdaq.mark_disconnected()
        mdaq._last_attempt = 0.0

    logger.info(f"📡 DAQ conexión actualizada: {port} @ {baudrate} baud, slave={slave_id}")
    return jsonify({
        "ok":       True,
        "port":     port,
        "baudrate": baudrate,
        "slave_id": slave_id,
        "msg":      f"Guardado en BD y reconectando a {port}...",
    })

@app.route("/api/daq/reboot", methods=["POST"])
def post_daq_reboot():
    """
    Fuerza el reinicio de la comunicación con la DAQ M-7026.
    Libera el puerto COM y restablece el cooldown para reconectar de inmediato.
    """
    try:
        import sys
        mdaq_instances = []
        import modbus_daq as _mdaq1
        mdaq_instances.append(_mdaq1)
        
        _mdaq2 = sys.modules.get('modbus_daq')
        if _mdaq2:
            mdaq_instances.append(_mdaq2)
            
        for mdaq in mdaq_instances:
            # 1. Marcar como desconectado (cierra el puerto serial COM activo)
            mdaq.mark_disconnected()
            # 2. Forzar reconexión inmediata en el próximo ciclo del ScanEngine (cooldown = 0)
            mdaq._last_attempt = 0.0
        
        logger.info("[DAQ] Comando de reinicio de puerto/conexión ejecutado. Reconectando...")
        return jsonify({
            "ok": True,
            "message": "Comunicación con la DAQ M-7026 reiniciada. Puerto serial liberado y reconectando..."
        })
    except Exception as e:
        logger.error(f"[DAQ] Error al reiniciar la conexión de la DAQ: {e}")
        return jsonify({
            "ok": False,
            "error": f"Error al reiniciar la conexión: {e}"
        }), 500




# ─────────────────────────────────────────────────────────────
# API HART — Configuración y Estado en Tiempo Real
# ─────────────────────────────────────────────────────────────

@app.route("/api/hart/config", methods=["GET"])
def get_hart_config():
    return jsonify(HART_CONFIG)

@app.route("/api/hart/config", methods=["POST"])
def post_hart_config():
    d = request.get_json() or {}
    HART_CONFIG.update(d)
    try:
        with open(HART_CONFIG_FILE, "w") as f:
            json.dump(HART_CONFIG, f)
    except Exception:
        pass
    return jsonify({"ok": True, "config": HART_CONFIG})

@app.route("/api/hart/reboot", methods=["POST"])
def post_hart_reboot():
    """
    Realiza login y solicita reboot al gateway ICP DAS HRT-711.
    También desconecta la conexión persistente actual para forzar reconexión posterior.
    """
    mode = HART_CONFIG.get("mode", "tcp")
    if mode != "tcp":
        return jsonify({"ok": False, "error": "El reinicio por software solo está disponible en modo TCP/IP."}), 400

    ip = HART_CONFIG.get("ip", "192.168.255.1")
    password = "admin123"  # Contraseña estándar del gateway

    # 1. Forzar desconexión local del poller para que no intente usar el socket mientras se reinicia
    try:
        from comunicacion_hart import force_disconnect
        force_disconnect()
    except Exception as e:
        logger.error(f"Error forzando desconexión HART: {e}")

    # 2. Hacer la solicitud HTTP al gateway
    import urllib.request
    import http.cookiejar
    
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    
    try:
        login_url = f"http://{ip}/login.cgi?webpwd={password}"
        logger.info(f"[HART] Iniciando login de reinicio en {login_url}")
        req = urllib.request.Request(login_url)
        with opener.open(req, timeout=5) as resp:
            resp.read()
            
        reboot_url = f"http://{ip}/reboot.cgi?mysubmit2=Reboot"
        logger.info(f"[HART] Enviando comando de reboot en {reboot_url}")
        req_reboot = urllib.request.Request(reboot_url)
        # Se usa un timeout corto de 5s; es común que falle si la conexión se cierra inmediatamente tras recibir el comando
        try:
            with opener.open(req_reboot, timeout=5) as resp:
                resp.read()
        except Exception as re:
            # Ignorar si es un reset de conexión normal al reiniciarse
            logger.info(f"[HART] Envío de reboot completado (respuesta recibida/conexión reseteada: {re})")
            
        return jsonify({"ok": True, "message": "Comando de reinicio enviado correctamente al gateway. Guardando cambios y reconectando en 15s."})
    except Exception as e:
        logger.error(f"[HART] Error al reiniciar gateway via HTTP: {e}")
        return jsonify({"ok": False, "error": f"Error de comunicación con el gateway web: {e}"}), 502

@app.route("/api/hart/live", methods=["GET"])
def get_hart_live():
    """Devuelve el estado global y canales HART desde la caché (no bloquea)."""
    import time as _time
    from datetime import datetime
    with _HART_CACHE_LOCK:
        results_list = [
            {"channel_idx": idx, **res}
            for idx, res in sorted(_HART_LATEST_RESULTS.items())
        ]
        ls = _HART_GLOBAL["last_success"]
        data_age_s = (_time.monotonic() - ls) if ls else 0.0
        stale = data_age_s > 10.0 if ls else True
        connected = (ls is not None) and not stale
        cooldown_left = max(0.0, _HART_POLL_INTERVAL - (_time.monotonic() - _HART_GLOBAL["last_attempt"]))
        last_error = _HART_GLOBAL["last_error"]
        
    return jsonify({
        "connected": connected,
        "stale": stale,
        "data_age_s": round(data_age_s, 1),
        "last_error": last_error,
        "retry_in_s": round(cooldown_left, 1),
        "ts": datetime.now().strftime("%H:%M:%S"),
        "mode": HART_CONFIG.get("mode", "tcp"),
        "ip": HART_CONFIG.get("ip", "192.168.255.1"),
        "port": HART_CONFIG.get("port", 502),
        "com_port": HART_CONFIG.get("com_port", "COM3"),
        "baudrate": HART_CONFIG.get("baudrate", 9600),
        "channels": results_list
    })

@app.route("/api/hart/config/channels", methods=["GET"])
def get_hart_channels_config():
    """Retorna la configuración de los 15 canales HART desde la BD."""
    rows = db_exec("SELECT * FROM hart_channel_config ORDER BY channel_idx")
    return jsonify(rows or [])

@app.route("/api/hart/config/channels", methods=["POST"])
def save_hart_channel_config():
    """
    Guarda (UPSERT) la configuracion de un canal HART en la BD.
    Payload: { channel_idx, v_name, description, hart_device_index, hart_device_address, enabled }

    hart_device_index   = N en 'HART Device N' del HG Tool (0-based).
                          Determina la direccion Modbus: 1300 + N x 10.
    hart_device_address = Direccion HART fisica en el bus (1-15, informativo).
    """
    d = request.get_json() or {}
    required = ["channel_idx", "v_name", "description"]
    if not all(k in d for k in required):
        return jsonify({"error": "Faltan campos requeridos"}), 400

    # hart_device_index = N en 'HART Device N' -> determina direccion Modbus
    dev_index = d.get('hart_device_index')
    if dev_index is None:
        dev_index = int(d["channel_idx"])  # fallback
    dev_index = max(0, int(dev_index))

    # hart_device_address = direccion HART fisica (informativo)
    hart_addr = d.get('hart_device_address')
    if hart_addr is None:
        hart_addr = d.get('slave_id', dev_index + 1)
    hart_addr = max(0, int(hart_addr))

    modbus_addr = 1300 + dev_index * 10

    db_exec(
        """INSERT INTO hart_channel_config
             (channel_idx, v_name, description,
              hart_device_index, hart_device_address, slave_id, enabled, instrument_type)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
           ON DUPLICATE KEY UPDATE
             description=%s,
             hart_device_index=%s, hart_device_address=%s, slave_id=%s,
             enabled=%s, updated_at=NOW()""",
        (
            int(d["channel_idx"]),
            d["v_name"],
            d["description"],
            dev_index,
            hart_addr,
            hart_addr,
            bool(d.get("enabled", True)),
            d.get("instrument_type", "NONE"),
            # ON DUPLICATE KEY: solo actualiza descripcion, device, enabled (NO instrument_type)
            d["description"],
            dev_index,
            hart_addr,
            hart_addr,
            bool(d.get("enabled", True)),
        ),
        fetch=False,
    )
    logger.info(
        f"[HART] Canal {d['channel_idx']} guardado: "
        f"Device{dev_index} modbus_addr={modbus_addr} HART_bus={hart_addr}"
    )
    return jsonify({
        "ok": True,
        "channel_idx":         d["channel_idx"],
        "hart_device_index":   dev_index,
        "hart_device_address": hart_addr,
        "modbus_addr":         modbus_addr,
    })

# ─────────────────────────────────────────────────────────────
# API Modbus RTU Devices — Gestión de múltiples dispositivos
# ─────────────────────────────────────────────────────────────

# Caché de estado de conexión en memoria (id_dispositivo -> dict)
_MODBUS_RTU_STATUS = {}
_MODBUS_RTU_STATUS_LOCK = threading.Lock()

def _ensure_modbus_rtu_table():
    """Crea la tabla modbus_rtu_devices si no existe, y añade columnas faltantes (parity, stopbits)."""
    try:
        db_exec(
            """CREATE TABLE IF NOT EXISTS modbus_rtu_devices (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                name         VARCHAR(100) NOT NULL DEFAULT 'Dispositivo',
                port         VARCHAR(20)  NOT NULL DEFAULT 'COM3',
                baudrate     INT          NOT NULL DEFAULT 9600,
                slave_id     INT          NOT NULL DEFAULT 1,
                parity       VARCHAR(2)   NOT NULL DEFAULT 'N',
                stopbits     TINYINT      NOT NULL DEFAULT 1,
                enabled      TINYINT(1)   NOT NULL DEFAULT 1,
                created_at   DATETIME     DEFAULT NOW(),
                updated_at   DATETIME     DEFAULT NOW() ON UPDATE NOW()
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            fetch=False
        )
        # Auto-migración: añadir columnas faltantes en tablas ya existentes
        _cols = {r.get('Field') for r in (db_exec("SHOW COLUMNS FROM modbus_rtu_devices") or [])}
        if 'parity' not in _cols:
            db_exec("ALTER TABLE modbus_rtu_devices ADD COLUMN parity VARCHAR(2) NOT NULL DEFAULT 'N' AFTER slave_id", fetch=False)
            logger.info("[ModbusRTU] Columna 'parity' agregada a modbus_rtu_devices")
        if 'stopbits' not in _cols:
            db_exec("ALTER TABLE modbus_rtu_devices ADD COLUMN stopbits TINYINT NOT NULL DEFAULT 1 AFTER parity", fetch=False)
            logger.info("[ModbusRTU] Columna 'stopbits' agregada a modbus_rtu_devices")
    except Exception as e:
        logger.warning(f"[ModbusRTU] No se pudo crear/migrar tabla: {e}")

# Crear la tabla al importar
try:
    _ensure_modbus_rtu_table()
except Exception:
    pass


@app.route("/api/modbus_rtu/devices", methods=["GET"])
def modbus_rtu_get_devices():
    """Retorna todos los dispositivos Modbus RTU configurados."""
    try:
        _ensure_modbus_rtu_table()
        rows = db_exec("SELECT * FROM modbus_rtu_devices ORDER BY id")
        return jsonify(rows or [])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/modbus_rtu/devices", methods=["POST"])
def modbus_rtu_save_device():
    """
    Crea o actualiza un dispositivo Modbus RTU.
    Payload: { id (opcional), name, port, baudrate, slave_id, parity, stopbits, enabled }
    """
    d = request.get_json() or {}
    name     = str(d.get("name",     "Dispositivo"))[:100]
    port     = str(d.get("port",     "COM3")).upper()[:20]
    baudrate = int(d.get("baudrate", 9600))
    slave_id = int(d.get("slave_id", 1))
    parity   = str(d.get("parity",   "N")).upper()[:2]
    stopbits = int(d.get("stopbits", 1))
    enabled  = bool(d.get("enabled", True))
    dev_id   = d.get("id")

    try:
        _ensure_modbus_rtu_table()
        if dev_id:
            db_exec(
                """UPDATE modbus_rtu_devices
                   SET name=%s, port=%s, baudrate=%s, slave_id=%s,
                       parity=%s, stopbits=%s, enabled=%s, updated_at=NOW()
                   WHERE id=%s""",
                (name, port, baudrate, slave_id, parity, stopbits, enabled, int(dev_id)),
                fetch=False
            )
            new_id = int(dev_id)
        else:
            db_exec(
                """INSERT INTO modbus_rtu_devices (name, port, baudrate, slave_id, parity, stopbits, enabled)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (name, port, baudrate, slave_id, parity, stopbits, enabled),
                fetch=False
            )
            rows = db_exec("SELECT LAST_INSERT_ID() AS id")
            new_id = rows[0]["id"] if rows else None

        logger.info(f"[ModbusRTU] Dispositivo {'actualizado' if dev_id else 'creado'}: id={new_id} {name} {port}@{baudrate} {parity}-8-{stopbits} slave={slave_id}")
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        logger.error(f"[ModbusRTU] Error guardando dispositivo: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/modbus_rtu/devices/<int:dev_id>", methods=["DELETE"])
def modbus_rtu_delete_device(dev_id):
    """Elimina un dispositivo Modbus RTU por ID."""
    try:
        db_exec("DELETE FROM modbus_rtu_devices WHERE id=%s", (dev_id,), fetch=False)
        with _MODBUS_RTU_STATUS_LOCK:
            _MODBUS_RTU_STATUS.pop(dev_id, None)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/modbus_rtu/status", methods=["GET"])
def modbus_rtu_get_status():
    """
    Retorna el estado de conexión de todos los dispositivos.
    El estado es actualizado por el endpoint /test o por un poller futuro.
    """
    with _MODBUS_RTU_STATUS_LOCK:
        return jsonify(dict(_MODBUS_RTU_STATUS))


@app.route("/api/modbus_rtu/test", methods=["POST"])
def modbus_rtu_test_connection():
    """
    Verifica la conexión a un dispositivo Modbus RTU.
    Payload: { port, baudrate, slave_id, parity?, stopbits?, id? }
    Reutiliza el cliente del pool persistente (no abre un puerto nuevo)
    y adquiere el lock de transacción para no interferir con el poller.
    """
    d        = request.get_json() or {}
    port     = str(d.get("port",     "COM3")).upper()
    baudrate = int(d.get("baudrate", 9600))
    slave_id = int(d.get("slave_id", 1))
    parity   = str(d.get("parity",   "N")).upper()
    stopbits = int(d.get("stopbits", 1))
    dev_id   = d.get("id")

    result = {"connected": False, "error": "", "latency_ms": None, "port": port,
              "baudrate": baudrate, "slave_id": slave_id, "parity": parity, "stopbits": stopbits}
    try:
        import time as _time
        port_lock = _get_port_lock(port, baudrate, parity, stopbits)
        # Obtener (o crear) el cliente persistente del pool
        client = _get_or_create_modbus_client(port, baudrate, parity, stopbits)
        if client is None:
            result["error"] = f"No se pudo abrir puerto {port}"
        else:
            # Adquirir el lock de transacción (espera máx 2s para no bloquear Flask)
            if not port_lock.acquire(timeout=2.0):
                result["error"] = f"Puerto {port} ocupado por otro proceso — intenta de nuevo"
            else:
                try:
                    t0 = _time.monotonic()
                    resp = None
                    for addr in [0, 1]:
                        try:
                            resp = client.read_holding_registers(addr, count=2, slave=slave_id)
                            if resp and not resp.isError():
                                break
                            resp = client.read_input_registers(addr, count=2, slave=slave_id)
                            if resp and not resp.isError():
                                break
                        except Exception:
                            pass
                    latency = round((_time.monotonic() - t0) * 1000, 1)
                    result["connected"]  = True
                    result["latency_ms"] = latency
                    if resp and not resp.isError():
                        result["registers"] = list(getattr(resp, 'registers', []))
                        result["error"]     = ""
                    else:
                        result["registers"] = []
                        result["error"]     = (f"Puerto {port} abierto OK ({latency:.0f} ms). "
                                               f"Slave {slave_id} sin respuesta en addr 0-1 — "
                                               f"verifica el Node Address del instrumento.")
                finally:
                    port_lock.release()
    except Exception as e:
        result["error"] = str(e)

    if dev_id is not None:
        with _MODBUS_RTU_STATUS_LOCK:
            _MODBUS_RTU_STATUS[dev_id] = {
                "connected":  result["connected"],
                "error":      result["error"],
                "latency_ms": result["latency_ms"],
                "last_check": datetime.now().strftime("%H:%M:%S")
            }
    return jsonify(result)


@app.route("/api/modbus_rtu/scan", methods=["POST"])
def modbus_rtu_scan():
    """
    Escanea un puerto COM probando combinaciones de baudrate, paridad y slave IDs.
    Payload: { port, baudrates?, parities?, slave_start?, slave_end? }
    """
    d = request.get_json() or {}
    port = str(d.get("port", "COM3")).upper()
    baudrates = d.get("baudrates", [9600, 19200, 38400, 4800])
    parities = d.get("parities", ["N", "E", "O"])
    slave_start = max(1, int(d.get("slave_start", 1)))
    slave_end = min(247, max(slave_start, int(d.get("slave_end", 16))))
    addrs_to_test = [0, 250, 249, 1, 100]

    found = []
    try:
        from pymodbus.client import ModbusSerialClient as _MbClient
        for baud in baudrates:
            for parity in parities:
                for stopbits in ([1] if parity in ('E', 'O') else [1, 2]):
                    try:
                        client = _MbClient(port=port, baudrate=baud, bytesize=8, parity=parity, stopbits=stopbits, timeout=0.15)
                        if not client.connect():
                            continue
                        for s_id in range(slave_start, slave_end + 1):
                            answered = False
                            for fc in [3, 4]:
                                if answered: break
                                for test_addr in addrs_to_test:
                                    try:
                                        if fc == 3:
                                            resp = client.read_holding_registers(test_addr, count=2, slave=s_id)
                                        else:
                                            resp = client.read_input_registers(test_addr, count=2, slave=s_id)
                                        if resp and not resp.isError() and hasattr(resp, 'registers'):
                                            found.append({
                                                "port": port,
                                                "baudrate": baud,
                                                "parity": parity,
                                                "stopbits": stopbits,
                                                "slave_id": s_id,
                                                "fc": fc,
                                                "test_addr": test_addr,
                                                "registers": list(resp.registers)
                                            })
                                            answered = True
                                            break
                                    except Exception:
                                        pass
                        client.close()
                        if found: break
                    except Exception:
                        pass
                if found: break
            if found: break
        return jsonify({"ok": True, "found": found, "count": len(found)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
# API Modbus RTU — Comandos y Variables
# ─────────────────────────────────────────────────────────────

_MODBUS_CMD_VALUES = {}   # { cmd_id: { values:[], connected:bool, error:'', ts:'' } }
_MODBUS_CMD_LOCK   = threading.Lock()

def _ensure_modbus_cmd_tables():
    """Crea las tablas de comandos y variables si no existen."""
    try:
        db_exec("""CREATE TABLE IF NOT EXISTS modbus_rtu_commands (
            id               INT AUTO_INCREMENT PRIMARY KEY,
            device_id        INT NOT NULL,
            cmd_name         VARCHAR(100)  NOT NULL DEFAULT '',
            enabled          TINYINT(1)   NOT NULL DEFAULT 1,
            internal_address VARCHAR(100) NOT NULL DEFAULT '',
            poll_interval    INT          NOT NULL DEFAULT 1,
            reg_count        INT          NOT NULL DEFAULT 2,
            swap_code        VARCHAR(30)  NOT NULL DEFAULT 'No Change',
            node_address     INT          NOT NULL DEFAULT 1,
            modbus_function  VARCHAR(60)  NOT NULL DEFAULT 'FC 3 - Read Holding Registers (4X)',
            mb_address       INT          NOT NULL DEFAULT 0,
            num_variables    INT          NOT NULL DEFAULT 1,
            sort_order       INT          NOT NULL DEFAULT 0,
            created_at       DATETIME     DEFAULT NOW(),
            updated_at       DATETIME     DEFAULT NOW() ON UPDATE NOW()
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""", fetch=False)

        db_exec("""CREATE TABLE IF NOT EXISTS modbus_rtu_variables (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            command_id  INT          NOT NULL,
            var_index   INT          NOT NULL DEFAULT 0,
            var_name    VARCHAR(100) NOT NULL DEFAULT '',
            var_label   VARCHAR(100) NOT NULL DEFAULT '',
            current_val DOUBLE       DEFAULT NULL,
            updated_at  DATETIME     DEFAULT NOW() ON UPDATE NOW()
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""", fetch=False)
    except Exception as e:
        logger.warning(f"[ModbusRTU-CMD] No se pudo crear tabla: {e}")

try:
    _ensure_modbus_cmd_tables()
except Exception:
    pass



def _apply_swap_code(registers, swap_code, num_regs=2):
    """
    Aplica el swap code a los registros crudos Modbus y devuelve float32 o int.

    Convenciones de byte order estándar ProSoft / Prolink:
      - No Change         (0): ABCD (Big-Endian, MSW first)
      - Word Swap         (1): CDAB (LSW first)
      - Word and Byte Swap(2): DCBA (Little-Endian completo)
      - Byte Swap         (3): BADC (Swap de bytes dentro de cada palabra de 16-bit)

    Para registros de 16-bit (num_regs=1):
      - No Change / Word Swap:          MSB primero (A B)
      - Byte Swap / Word and Byte Swap: LSB primero (B A)
    """
    import struct as _struct

    if not registers:
        return None

    try:
        if num_regs == 1 or len(registers) == 1:
            raw = registers[0] & 0xFFFF
            if swap_code in ('Byte Swap', 'Word and Byte Swap'):
                raw = ((raw & 0xFF) << 8) | ((raw >> 8) & 0xFF)
            return _struct.unpack('>h', _struct.pack('>H', raw))[0]

        # 2 registros = 32 bits
        r0 = registers[0] & 0xFFFF
        r1 = registers[1] & 0xFFFF
        A = (r0 >> 8) & 0xFF; B = r0 & 0xFF
        C = (r1 >> 8) & 0xFF; D = r1 & 0xFF

        if swap_code == 'No Change':
            raw_bytes = bytes([A, B, C, D])
        elif swap_code == 'Word Swap':
            raw_bytes = bytes([C, D, A, B])
        elif swap_code == 'Word and Byte Swap':
            raw_bytes = bytes([D, C, B, A])
        elif swap_code == 'Byte Swap':
            raw_bytes = bytes([B, A, D, C])
        else:
            raw_bytes = bytes([A, B, C, D])

        val = _struct.unpack('>f', raw_bytes)[0]
        # Evitar valores NaN / Inf
        import math as _math
        if _math.isnan(val) or _math.isinf(val):
            return None
        return round(val, 6)
    except Exception as e:
        logger.debug(f"[SwapCode] Error interpretando registros {registers}: {e}")
        return None


@app.route("/api/modbus_rtu/devices/<int:dev_id>/commands", methods=["GET"])
def modbus_rtu_get_commands(dev_id):
    """Lista todos los comandos del dispositivo, con sus variables."""
    try:
        _ensure_modbus_cmd_tables()
        cmds = db_exec(
            "SELECT * FROM modbus_rtu_commands WHERE device_id=%s ORDER BY sort_order, id",
            (dev_id,)
        ) or []
        for cmd in cmds:
            cmd_id = cmd.get("id") or cmd.get("cmd_id")
            vars_rows = db_exec(
                "SELECT * FROM modbus_rtu_variables WHERE command_id=%s ORDER BY var_index",
                (cmd_id,)
            ) or []
            cmd["variables"] = vars_rows
        return jsonify(cmds)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/modbus_rtu/commands", methods=["POST"])
def modbus_rtu_save_command():
    """
    Crea o actualiza un comando Modbus RTU.
    Payload: { id?, device_id, cmd_name, enabled, internal_address, poll_interval,
               reg_count, swap_code, node_address, modbus_function, mb_address,
               num_variables, sort_order }
    """
    d = request.get_json() or {}
    required = ["device_id"]
    if not all(k in d for k in required):
        return jsonify({"error": "Falta device_id"}), 400

    cmd_id      = d.get("id")
    device_id   = int(d["device_id"])
    cmd_name    = str(d.get("cmd_name", ""))[:100]
    enabled     = bool(d.get("enabled", True))
    int_addr    = str(d.get("internal_address", ""))[:100]
    poll_int    = int(d.get("poll_interval", 1))
    reg_count   = max(1, int(d.get("reg_count", 2)))
    swap_code   = str(d.get("swap_code", "No Change"))[:30]
    node_addr   = int(d.get("node_address", 1))
    mb_func     = str(d.get("modbus_function", "FC 3 - Read Holding Registers (4X)"))[:60]
    mb_addr     = int(d.get("mb_address", 0))
    num_vars    = max(1, int(d.get("num_variables", 1)))
    sort_order  = int(d.get("sort_order", 0))

    try:
        _ensure_modbus_cmd_tables()
        if cmd_id:
            db_exec(
                """UPDATE modbus_rtu_commands SET
                   cmd_name=%s, enabled=%s, internal_address=%s, poll_interval=%s,
                   reg_count=%s, swap_code=%s, node_address=%s, modbus_function=%s,
                   mb_address=%s, num_variables=%s, sort_order=%s, updated_at=NOW()
                   WHERE id=%s""",
                (cmd_name, enabled, int_addr, poll_int, reg_count, swap_code,
                 node_addr, mb_func, mb_addr, num_vars, sort_order, int(cmd_id)),
                fetch=False
            )
            new_id = int(cmd_id)
        else:
            db_exec(
                """INSERT INTO modbus_rtu_commands
                   (device_id, cmd_name, enabled, internal_address, poll_interval,
                    reg_count, swap_code, node_address, modbus_function, mb_address,
                    num_variables, sort_order)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (device_id, cmd_name, enabled, int_addr, poll_int, reg_count,
                 swap_code, node_addr, mb_func, mb_addr, num_vars, sort_order),
                fetch=False
            )
            rows   = db_exec("SELECT LAST_INSERT_ID() AS id")
            new_id = rows[0]["id"] if rows else None

        # Sincronizar variables: asegura que haya exactamente num_vars filas
        if new_id:
            existing = db_exec(
                "SELECT id, var_index FROM modbus_rtu_variables WHERE command_id=%s ORDER BY var_index",
                (new_id,)
            ) or []
            existing_idx = {r["var_index"] for r in existing}

            # Insertar las que faltan
            for i in range(num_vars):
                if i not in existing_idx:
                    db_exec(
                        """INSERT INTO modbus_rtu_variables (command_id, var_index, var_name, var_label)
                           VALUES (%s,%s,%s,%s)""",
                        (new_id, i, f"var_{new_id}_{i}", f"Variable {i+1}"),
                        fetch=False
                    )
            # Eliminar las sobrantes
            if len(existing) > num_vars:
                for r in existing:
                    if r["var_index"] >= num_vars:
                        db_exec("DELETE FROM modbus_rtu_variables WHERE id=%s", (r["id"],), fetch=False)

        logger.info(f"[ModbusRTU-CMD] Comando {'actualizado' if cmd_id else 'creado'}: id={new_id} dev={device_id}")
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        logger.error(f"[ModbusRTU-CMD] Error guardando comando: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/modbus_rtu/commands/<int:cmd_id>", methods=["DELETE"])
def modbus_rtu_delete_command(cmd_id):
    """Elimina un comando y sus variables."""
    try:
        db_exec("DELETE FROM modbus_rtu_variables WHERE command_id=%s", (cmd_id,), fetch=False)
        db_exec("DELETE FROM modbus_rtu_commands WHERE id=%s", (cmd_id,), fetch=False)
        with _MODBUS_CMD_LOCK:
            _MODBUS_CMD_VALUES.pop(cmd_id, None)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/modbus_rtu/commands/<int:cmd_id>/variables", methods=["POST"])
def modbus_rtu_save_variables(cmd_id):
    """
    Guarda los nombres/etiquetas de las variables de un comando.
    Payload: [ { var_index, var_name, var_label }, ... ]
    """
    vars_list = request.get_json() or []
    try:
        for v in vars_list:
            db_exec(
                """UPDATE modbus_rtu_variables
                   SET var_name=%s, var_label=%s, updated_at=NOW()
                   WHERE command_id=%s AND var_index=%s""",
                (str(v.get("var_name",""))[:100], str(v.get("var_label",""))[:100],
                 cmd_id, int(v.get("var_index", 0))),
                fetch=False
            )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/modbus_rtu/live_values", methods=["GET"])
def modbus_rtu_live_values():
    """Retorna los últimos valores leídos de todos los comandos."""
    with _MODBUS_CMD_LOCK:
        return jsonify(dict(_MODBUS_CMD_VALUES))


def _read_modbus_command_sync(port, baudrate, cmd, parity='N', stopbits=1):
    """Ejecuta una lectura Modbus para un comando y devuelve el resultado dict."""
    slave_id     = int(cmd.get("node_address", 1))
    mb_addr      = int(cmd.get("mb_address", 0))
    reg_count    = max(1, int(cmd.get("reg_count", 2)))
    swap_code    = str(cmd.get("swap_code", "No Change"))
    num_vars     = max(1, int(cmd.get("num_variables", 1)))
    mb_func_str  = str(cmd.get("modbus_function", "FC 3"))
    regs_per_var = max(1, reg_count // num_vars) if num_vars > 0 else reg_count

    fc_num = 3
    if "FC 1" in mb_func_str or "FC1" in mb_func_str:  fc_num = 1
    elif "FC 2" in mb_func_str or "FC2" in mb_func_str: fc_num = 2
    elif "FC 4" in mb_func_str or "FC4" in mb_func_str: fc_num = 4

    result = {"connected": False, "error": "", "values": [], "ts": datetime.now().strftime("%H:%M:%S")}
    import time as _time

    try:
        from pymodbus.client import ModbusSerialClient as _MbClient
        client = _MbClient(port=port, baudrate=baudrate, bytesize=8, parity=parity, stopbits=stopbits, timeout=1.5)
        if client.connect():
            parsed_values = []
            t0 = _time.monotonic()
            any_success = False
            for v_idx in range(num_vars):
                start_addr = mb_addr + v_idx * regs_per_var
                try:
                    try:
                        if fc_num == 1:
                            resp = client.read_coils(start_addr, count=regs_per_var, slave=slave_id)
                        elif fc_num == 2:
                            resp = client.read_discrete_inputs(start_addr, count=regs_per_var, slave=slave_id)
                        elif fc_num == 4:
                            resp = client.read_input_registers(start_addr, count=regs_per_var, slave=slave_id)
                        else:
                            resp = client.read_holding_registers(start_addr, count=regs_per_var, slave=slave_id)
                    except TypeError:
                        # Fallback pymodbus < 3.0 / unit parameter
                        if fc_num == 1:
                            resp = client.read_coils(start_addr, count=regs_per_var, unit=slave_id)
                        elif fc_num == 2:
                            resp = client.read_discrete_inputs(start_addr, count=regs_per_var, unit=slave_id)
                        elif fc_num == 4:
                            resp = client.read_input_registers(start_addr, count=regs_per_var, unit=slave_id)
                        else:
                            resp = client.read_holding_registers(start_addr, count=regs_per_var, unit=slave_id)

                    if resp is None:
                        result["error"] = f"Timeout en {port} Slave {slave_id} Addr {start_addr}"
                        parsed_values.append(None)
                    elif hasattr(resp, 'isError') and resp.isError():
                        result["error"] = f"Modbus Error: {resp}"
                        logger.warning(f"[ModbusRTU] Error: {port} slave={slave_id} addr={start_addr} -> {resp}")
                        parsed_values.append(None)
                    elif fc_num in (1, 2):
                        bits = getattr(resp, 'bits', [])
                        val = int(bits[0]) if bits else None
                        parsed_values.append(val)
                        any_success = True
                    else:
                        regs = list(getattr(resp, 'registers', []))
                        val = _apply_swap_code(regs, swap_code, regs_per_var)
                        logger.info(f"[ModbusRTU] ✅ Lectura OK {port} slave={slave_id} addr={start_addr} regs={regs} -> val={val} ({swap_code})")
                        parsed_values.append(val)
                        any_success = True
                except Exception as e_v:
                    parsed_values.append(None)
                    result["error"] = str(e_v)
                    logger.debug(f"[ModbusRTU] Error leyendo var {v_idx}: {e_v}")

            client.close()
            latency = round((_time.monotonic() - t0) * 1000, 1)
            result["connected"] = any_success or (not result["error"])
            result["values"]    = parsed_values
            result["latency_ms"] = latency

            # Persistir valores en BD
            cmd_id = cmd.get("id")
            if cmd_id:
                try:
                    vars_rows = db_exec(
                        "SELECT id, var_index, var_name FROM modbus_rtu_variables WHERE command_id=%s ORDER BY var_index",
                        (cmd_id,)
                    ) or []
                    overrides = getattr(V, "instrument_overrides", {})
                    for vr in vars_rows:
                        v_name = vr.get("var_name")
                        if v_name and v_name in overrides:
                            continue
                        vi = vr["var_index"]
                        val = parsed_values[vi] if vi < len(parsed_values) else None
                        if val is not None:
                            db_exec(
                                "UPDATE modbus_rtu_variables SET current_val=%s, updated_at=NOW() WHERE id=%s",
                                (val, vr["id"]), fetch=False
                            )
                except Exception as e_db:
                    logger.debug(f"[ModbusRTU] Error persistiendo a BD: {e_db}")
        else:
            result["error"] = f"No se pudo abrir puerto {port}"
    except ModuleNotFoundError:
        result["error"] = "pymodbus no instalado"
    except Exception as e:
        result["error"] = str(e)

    return result


@app.route("/api/modbus_rtu/commands/<int:cmd_id>/poll", methods=["POST"])
def modbus_rtu_poll_command(cmd_id):
    """
    Lee registros del instrumento para un comando específico bajo demanda.
    Payload: { device: { port, baudrate, parity?, stopbits? }, cmd: { node_address, modbus_function,
               mb_address, reg_count, swap_code, num_variables } }
    """
    d   = request.get_json() or {}
    dev = d.get("device", {})
    cmd = d.get("cmd", {})
    cmd["id"] = cmd_id

    port     = str(dev.get("port", "COM3")).upper()
    baudrate = int(dev.get("baudrate", 9600))
    parity   = str(dev.get("parity", "N"))
    stopbits = int(dev.get("stopbits", 1))

    result = _read_modbus_command_sync(port, baudrate, cmd, parity=parity, stopbits=stopbits)

    # Actualizar caché
    with _MODBUS_CMD_LOCK:
        _MODBUS_CMD_VALUES[cmd_id] = {
            "connected":  result["connected"],
            "error":      result["error"],
            "values":     result["values"],
            "latency_ms": result.get("latency_ms"),
            "ts":         result["ts"]
        }

    return jsonify(result)


from modbus_pool import (
    get_port_lock as _get_port_lock,
    get_or_create_modbus_client as _get_or_create_modbus_client,
    invalidate_modbus_client as _invalidate_modbus_client
)


def _read_modbus_command_persistent(port, baudrate, cmd, parity='N', stopbits=1):
    """
    Versión del lector Modbus que usa el cliente persistente del pool.
    Si el cliente falla, lo invalida para que se reconecte en el próximo ciclo.
    """
    slave_id     = int(cmd.get("node_address", 1))
    mb_addr      = int(cmd.get("mb_address", 0))
    reg_count    = max(1, int(cmd.get("reg_count", 2)))
    swap_code    = str(cmd.get("swap_code", "No Change"))
    num_vars     = max(1, int(cmd.get("num_variables", 1)))
    mb_func_str  = str(cmd.get("modbus_function", "FC 3"))
    regs_per_var = max(1, reg_count // num_vars) if num_vars > 0 else reg_count

    fc_num = 3
    if "FC 1" in mb_func_str or "FC1" in mb_func_str:   fc_num = 1
    elif "FC 2" in mb_func_str or "FC2" in mb_func_str:  fc_num = 2
    elif "FC 4" in mb_func_str or "FC4" in mb_func_str:  fc_num = 4

    result = {"connected": False, "error": "", "values": [], "ts": datetime.now().strftime("%H:%M:%S")}
    import time as _time

    client = _get_or_create_modbus_client(port, baudrate, parity, stopbits)
    if client is None:
        result["error"] = f"No se pudo abrir puerto {port}"
        return result

    # Adquirir el lock de transacción — serializa lecturas concurrentes (poller + /test + /poll)
    port_lock = _get_port_lock(port, baudrate, parity, stopbits)
    if not port_lock.acquire(timeout=3.0):
        result["error"] = f"Timeout esperando lock de {port} (bus ocupado)"
        return result

    try:
        parsed_values = []
        t0 = _time.monotonic()
        any_success = False
        for v_idx in range(num_vars):
            start_addr = mb_addr + v_idx * regs_per_var
            try:
                try:
                    if fc_num == 1:
                        resp = client.read_coils(start_addr, count=regs_per_var, slave=slave_id)
                    elif fc_num == 2:
                        resp = client.read_discrete_inputs(start_addr, count=regs_per_var, slave=slave_id)
                    elif fc_num == 4:
                        resp = client.read_input_registers(start_addr, count=regs_per_var, slave=slave_id)
                    else:
                        resp = client.read_holding_registers(start_addr, count=regs_per_var, slave=slave_id)
                except TypeError:
                    # Fallback pymodbus < 3.0 (parámetro 'unit')
                    if fc_num == 1:
                        resp = client.read_coils(start_addr, count=regs_per_var, unit=slave_id)
                    elif fc_num == 2:
                        resp = client.read_discrete_inputs(start_addr, count=regs_per_var, unit=slave_id)
                    elif fc_num == 4:
                        resp = client.read_input_registers(start_addr, count=regs_per_var, unit=slave_id)
                    else:
                        resp = client.read_holding_registers(start_addr, count=regs_per_var, unit=slave_id)

                if resp is None:
                    result["error"] = f"Timeout en {port} Slave {slave_id} Addr {start_addr}"
                    parsed_values.append(None)
                elif hasattr(resp, 'isError') and resp.isError():
                    result["error"] = f"Modbus Error: {resp}"
                    logger.debug(f"[ModbusRTU] Error: {port} slave={slave_id} addr={start_addr} -> {resp}")
                    parsed_values.append(None)
                elif fc_num in (1, 2):
                    bits = getattr(resp, 'bits', [])
                    val = int(bits[0]) if bits else None
                    parsed_values.append(val)
                    any_success = True
                else:
                    regs = list(getattr(resp, 'registers', []))
                    val = _apply_swap_code(regs, swap_code, regs_per_var)
                    logger.info(f"[ModbusRTU] ✅ {port} slave={slave_id} addr={start_addr} regs={regs} -> {val} ({swap_code})")
                    parsed_values.append(val)
                    any_success = True
            except Exception as e_v:
                parsed_values.append(None)
                result["error"] = str(e_v)
                logger.debug(f"[ModbusRTU] Error leyendo var {v_idx}: {e_v}")
                # Si hay excepción de comunicación, invalidar cliente para reconexión
                if any(kw in str(e_v).lower() for kw in ('permission', 'access', 'closed', 'not connected', 'invalid')):
                    _invalidate_modbus_client(port, baudrate, parity, stopbits)
                    result["error"] = f"Conexión perdida en {port} — reconectando en próximo ciclo"
                    break

        latency = round((_time.monotonic() - t0) * 1000, 1)
        result["connected"]   = any_success
        result["values"]      = parsed_values
        result["latency_ms"]  = latency

        # Persistir valores en BD
        cmd_id = cmd.get("id")
        if cmd_id and any_success:
            try:
                vars_rows = db_exec(
                    "SELECT id, var_index, var_name FROM modbus_rtu_variables WHERE command_id=%s ORDER BY var_index",
                    (cmd_id,)
                ) or []
                overrides = getattr(V, "instrument_overrides", {})
                for vr in vars_rows:
                    v_name = vr.get("var_name")
                    if v_name and v_name in overrides:
                        continue
                    vi  = vr["var_index"]
                    val = parsed_values[vi] if vi < len(parsed_values) else None
                    if val is not None:
                        db_exec(
                            "UPDATE modbus_rtu_variables SET current_val=%s, updated_at=NOW() WHERE id=%s",
                            (val, vr["id"]), fetch=False
                        )
            except Exception as e_db:
                logger.debug(f"[ModbusRTU] Error persistiendo a BD: {e_db}")

    except Exception as e:
        result["error"] = str(e)
        logger.warning(f"[ModbusRTU-Pool] Error inesperado en {port}: {e}")
        _invalidate_modbus_client(port, baudrate, parity, stopbits)
    finally:
        port_lock.release()

    return result


def _ping_modbus_device(port, baudrate, slave_id, parity='N', stopbits=1):
    """
    Verifica si el puerto COM abre y el slave responde a una lectura básica.
    Retorna (connected: bool, error: str, latency_ms: float|None).
    Usa el pool de clientes persistentes y el lock de transacción.
    """
    import time as _time
    client = _get_or_create_modbus_client(port, baudrate, parity, stopbits)
    if client is None:
        return False, f"No se pudo abrir puerto {port}", None

    port_lock = _get_port_lock(port, baudrate, parity, stopbits)
    if not port_lock.acquire(timeout=2.0):
        return False, "Bus ocupado (timeout lock)", None

    try:
        t0 = _time.monotonic()
        resp = None
        for addr in [0, 1, 100, 250]:
            try:
                resp = client.read_holding_registers(addr, count=2, slave=slave_id)
                if resp and not resp.isError():
                    break
                resp = client.read_input_registers(addr, count=2, slave=slave_id)
                if resp and not resp.isError():
                    break
            except Exception:
                pass
        latency = round((_time.monotonic() - t0) * 1000, 1)
        if resp and not resp.isError():
            return True, "", latency
        else:
            # Puerto abre OK aunque slave no responda en esas direcciones
            return True, "Puerto abierto — slave sin respuesta en addr 0,1,100,250", latency
    except Exception as e:
        _invalidate_modbus_client(port, baudrate, parity, stopbits)
        return False, str(e), None
    finally:
        port_lock.release()


def _modbus_rtu_background_poller():
    """
    Hilo daemon que consulta periódicamente los dispositivos y comandos Modbus RTU
    habilitados. Usa conexiones persistentes por puerto para evitar el PermissionError
    de Windows al abrir/cerrar el puerto serial en cada ciclo.

    Comportamiento:
    - Si hay comandos: ejecuta cada comando y reporta conectado si al menos uno tuvo éxito.
    - Si NO hay comandos: hace un ping básico al slave_id del dispositivo para reportar
      si el puerto/slave está accesible (evita el "desconectado" falso al no tener comandos).
    """
    import time
    logger.info("[ModbusRTU-Poller] Iniciando hilo de sondeo continuo Modbus RTU (conexiones persistentes)")
    while True:
        try:
            time.sleep(1.0)
            _ensure_modbus_cmd_tables()
            devices = db_exec("SELECT * FROM modbus_rtu_devices WHERE enabled=1") or []
            if not devices:
                continue

            for dev in devices:
                dev_id   = dev["id"]
                port     = str(dev.get("port", "COM3")).upper()
                baudrate = int(dev.get("baudrate", 9600))
                parity   = str(dev.get("parity",   "N"))
                stopbits = int(dev.get("stopbits", 1))
                slave_id = int(dev.get("slave_id", 1))

                commands = db_exec(
                    "SELECT * FROM modbus_rtu_commands WHERE device_id=%s AND enabled=1 ORDER BY sort_order, id",
                    (dev_id,)
                ) or []

                if not commands:
                    # Sin comandos configurados: ping básico para verificar que el puerto responde
                    connected, error, latency = _ping_modbus_device(
                        port, baudrate, slave_id, parity=parity, stopbits=stopbits
                    )
                    with _MODBUS_RTU_STATUS_LOCK:
                        _MODBUS_RTU_STATUS[dev_id] = {
                            "connected":  connected,
                            "error":      error,
                            "latency_ms": latency,
                            "last_check": datetime.now().strftime("%H:%M:%S")
                        }
                    continue

                any_dev_connected = False
                last_error = ""
                for cmd in commands:
                    cmd_id = cmd["id"]
                    # Usar lector con cliente persistente (no abre/cierra el puerto)
                    res = _read_modbus_command_persistent(port, baudrate, cmd, parity=parity, stopbits=stopbits)
                    if res.get("connected"):
                        any_dev_connected = True
                    else:
                        last_error = res.get("error", "")
                    with _MODBUS_CMD_LOCK:
                        _MODBUS_CMD_VALUES[cmd_id] = {
                            "connected":  res["connected"],
                            "error":      res["error"],
                            "values":     res["values"],
                            "latency_ms": res.get("latency_ms"),
                            "ts":         res["ts"]
                        }

                with _MODBUS_RTU_STATUS_LOCK:
                    _MODBUS_RTU_STATUS[dev_id] = {
                        "connected":  any_dev_connected,
                        "error":      "" if any_dev_connected else (last_error or "Sin respuesta"),
                        "last_check": datetime.now().strftime("%H:%M:%S")
                    }

        except Exception as e:
            logger.debug(f"[ModbusRTU-Poller] Excepción en ciclo: {e}")
            time.sleep(2.0)



# ─────────────────────────────────────────────────────────────
# WebSocket Events
# ─────────────────────────────────────────────────────────────

@socketio.on("connect")
def on_connect():
    logger.info(f"🔌 Cliente conectado: {request.sid}")
    # Enviar estado inicial del PID al conectar
    emit("pid_config", {
        "PIC-01": {
            "instrumento": "PIC-01",
            "modo":       "Auto" if not V.b_MAN_PC else "Manual",
            "SP":         V.r_PRESS_PID_SP,
            "Kp":         V.r_PRESS_PID_03_KP,
            "Ki":         V.r_PRESS_PID_03_KI,
            "Kd":         V.r_PRESS_PID_03_KD,
        },
        "LIC-01": {
            "instrumento": "LIC-01",
            "modo":       "Auto" if not V.b_MAN_LC else "Manual",
            "SP":         V.r_LEVEL_PID_SP,
            "Kp":         V.r_LEVEL_PID_03_KP,
            "Ki":         V.r_LEVEL_PID_03_KI,
            "Kd":         V.r_LEVEL_PID_03_KD,
        },
    })


@socketio.on("disconnect")
def on_disconnect():
    logger.info(f"🔌 Cliente desconectado: {request.sid}")


@socketio.on("update_pid")
def ws_update_pid(data):
    """Recibe comandos de ajuste PID desde el frontend vía WebSocket."""
    tag = str(data.get("instrumento", "")).upper()
    if tag == "LIC-01":
        if "modo" in data:       V.b_MAN_LC                  = (data["modo"] == "Manual")
        if "SP" in data:         V.r_LEVEL_PID_SP             = float(data["SP"])
        if "CV_manual" in data:  
            V.r_LEVEL_PID_03_CVOverride  = float(data["CV_manual"])
            V.r_LEVEL_PID_03_CVOper      = float(data["CV_manual"])
        if "Kp" in data:         V.r_LEVEL_PID_03_KP          = float(data["Kp"])
        if "Ki" in data:         V.r_LEVEL_PID_03_KI          = float(data["Ki"])
        if "Kd" in data:         V.r_LEVEL_PID_03_KD          = float(data["Kd"])
        # Persistir en DB
        try:
            db_exec(
                "UPDATE configuracion_actual SET modo=%s,SP=%s,CV_manual=%s,Kp=%s,Ki=%s,Kd=%s WHERE instrumento=%s",
                ("Manual" if V.b_MAN_LC else "Auto", V.r_LEVEL_PID_SP,
                 V.r_LEVEL_PID_03_CVOverride, V.r_LEVEL_PID_03_KP,
                 V.r_LEVEL_PID_03_KI, V.r_LEVEL_PID_03_KD, "LIC-01"),
                fetch=False
            )
        except Exception as e:
            logger.error(f"Error guardando PID LIC-01 en BD (WS): {e}")

    elif tag == "PIC-01":
        if "modo" in data:       V.b_MAN_PC                   = (data["modo"] == "Manual")
        if "SP" in data:         V.r_PRESS_PID_SP             = float(data["SP"])
        if "CV_manual" in data:  
            V.r_PRESS_PID_03_CVOverride  = float(data["CV_manual"])
            V.r_PRESS_PID_03_CVOper      = float(data["CV_manual"])
        if "Kp" in data:         V.r_PRESS_PID_03_KP          = float(data["Kp"])
        if "Ki" in data:         V.r_PRESS_PID_03_KI          = float(data["Ki"])
        if "Kd" in data:         V.r_PRESS_PID_03_KD          = float(data["Kd"])
        # Persistir en DB
        try:
            db_exec(
                "UPDATE configuracion_actual SET modo=%s,SP=%s,CV_manual=%s,Kp=%s,Ki=%s,Kd=%s WHERE instrumento=%s",
                ("Manual" if V.b_MAN_PC else "Auto", V.r_PRESS_PID_SP,
                 V.r_PRESS_PID_03_CVOverride, V.r_PRESS_PID_03_KP,
                 V.r_PRESS_PID_03_KI, V.r_PRESS_PID_03_KD, "PIC-01"),
                fetch=False
            )
        except Exception as e:
            logger.error(f"Error guardando PID PIC-01 en BD (WS): {e}")

    emit("pid_updated", {"instrumento": tag, "ok": True}, broadcast=True)


@socketio.on("toggle_lazos")
def ws_toggle_lazos(_data=None):
    if V.b_DESHABILITA_PID:
        # Quiere habilitar
        V.b_PB_HABILITA_PID = True
        V.b_DESHABILITA_PID = False
    else:
        # Quiere deshabilitar
        V.b_PB_DESHABILITA_PID = True
        V.b_DESHABILITA_PID = True
        
    # Guardar en BD
    try:
        db_exec("UPDATE instrument_selection_config SET b_DESHABILITA_PID=%s WHERE id=1", (V.b_DESHABILITA_PID,), fetch=False)
    except Exception as e:
        logger.error(f"Error guardando b_DESHABILITA_PID en BD (WS): {e}")

    # Guardar variables retenidas
    try:
        from fase1_sistema import save_retained_vars
        save_retained_vars()
    except Exception as e:
        logger.error(f"Error guardando variables retenidas al cambiar lazos (WS): {e}")
        
    emit("lazos_status", {"lazos_habilitados": not V.b_DESHABILITA_PID}, broadcast=True)


# ─────────────────────────────────────────────────────────────
# Entrypoint
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    # Forzar stdout a UTF-8 para evitar UnicodeEncodeError en Windows (cp1252)
    if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    print("\n" + "=" * 60)
    print("  MFM ORINOCO -- SoftPLC + Flask 'Todo en Uno'")
    print("  Arquitectura: Memoria Compartida (objeto V)")
    print("=" * 60)

    # -- Restaurar configuracion DAQ desde BD --
    _load_daq_connection_from_db()

    # -- Restaurar mapeo de canales DAQ desde BD --
    _load_daq_channels_from_db()
    _load_daq_ao_from_db()

    # -- Restaurar seleccion de instrumentos y estado de lazos desde BD --
    _load_instrument_selection_from_db()

    # -- Restaurar parametros de lazos PID (configuracion_actual) desde BD --
    _load_configuracion_actual_from_db()

    # -- Restaurar overrides manuales de instrumentos desde BD --
    try:
        rows = db_exec("SELECT instrumento, modo_manual, valor_manual FROM tabla_configuracion_alarma")
        for row in (rows or []):
            if row.get("modo_manual") and row.get("valor_manual") is not None:
                _apply_manual_override(row["instrumento"], row["modo_manual"], row["valor_manual"])
        print("  [OK] Overrides manuales restaurados desde BD")
    except Exception as _em:
        print(f"  [WARN] No se pudieron restaurar overrides manuales: {_em}")

    # -- Sincronizar fallas de presion con limites de PI-01 y PI-02 --
    _sync_falla_presion()

    # -- Restaurar configuracion de prueba de pozo desde BD --
    _init_and_load_prueba_config()

    # -- Restaurar configuracion PVT y Balance de Masa desde BD --
    _init_and_load_pvt_balance_config()

    # -- Arrancar Hilo 1: ScanEngine (100 ms, daemon) --
    plc_engine.start()
    print(f"  [OK] SoftPLC ScanEngine activo ({len(PHASE_REGISTRY)} fases, 100 ms)")

    # -- Arrancar Hilo 2: WebSocket Updater (500 ms, daemon) --
    ws_thread = threading.Thread(target=websocket_updater, daemon=True, name="WSUpdater")
    ws_thread.start()
    print("  [OK] WebSocket Updater activo (500 ms)")

    # -- Arrancar Hilo 3: HART Background Poller (daemon) --
    hart_thread = threading.Thread(target=_hart_background_poller, daemon=True, name="HARTPoller")
    hart_thread.start()
    print(f"  [OK] HART Poller activo (cada {_HART_POLL_INTERVAL}s) -> {HART_CONFIG.get('ip')}:{HART_CONFIG.get('port')} (multi-drop, HART Device Address por canal)")

    # -- Arrancar Hilo 4: Modbus RTU Background Poller (daemon) --
    modbus_rtu_thread = threading.Thread(target=_modbus_rtu_background_poller, daemon=True, name="ModbusRTUPoller")
    modbus_rtu_thread.start()
    print("  [OK] Modbus RTU Poller activo (sondeo continuo de instrumentos RTU)")

    print("  -> http://localhost:5000")
    print("=" * 60 + "\n")

    try:
        socketio.run(app, host="0.0.0.0", port=5000,
                     debug=False, allow_unsafe_werkzeug=True)
    except KeyboardInterrupt:
        print("\nDeteniendo SoftPLC...")
        plc_engine.stop()
        print("  Motor detenido. Adios.")

